import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import numpy as np
from datetime import datetime
import base64
import os
import pickle
import tempfile
try:
    import chardet
except ImportError:
    chardet = None

# ─── Chemin de cache persistant (survit aux déconnexions) ────────────────────
_CACHE_DIR  = os.path.join(tempfile.gettempdir(), "saham_app_cache")
os.makedirs(_CACHE_DIR, exist_ok=True)

def _cache_path(key: str) -> str:
    return os.path.join(_CACHE_DIR, f"{key}.pkl")

def save_cache(key: str, obj) -> None:
    try:
        with open(_cache_path(key), "wb") as f:
            pickle.dump(obj, f)
    except Exception:
        pass

def load_cache(key: str):
    p = _cache_path(key)
    if os.path.exists(p):
        try:
            with open(p, "rb") as f:
                return pickle.load(f)
        except Exception:
            pass
    return None

def clear_cache(key: str) -> None:
    p = _cache_path(key)
    if os.path.exists(p):
        try:
            os.remove(p)
        except Exception:
            pass
# ─────────────────────────────────────────────────────────────────────────────

# Configuration de la page
st.set_page_config(
    page_title="Application de Structuration de Données",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personnalisé
st.markdown("""
<style>
    .main-header {
        text-align: center;
        background: linear-gradient(135deg, #1a4d3e 0%, #2d5f4e 100%);
        padding: 2.5rem;
        border-radius: 15px;
        color: white;
        margin-bottom: 2rem;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    
    .metric-card {
        background: white;
        padding: 2rem;
        border-radius: 15px;
        text-align: center;
        box-shadow: 0 4px 12px rgba(0,0,0,0.1);
        border-left: 5px solid #ff6b35;
    }
    
    .metric-value {
        font-size: 3em;
        font-weight: bold;
        color: #1a4d3e !important;
        margin: 0;
    }
    
    .metric-label {
        font-size: 1.1em;
        color: #6c757d !important;
        margin-top: 0.5rem;
    }
    
    .login-box {
        max-width: 450px;
        margin: 5rem auto;
        background: white;
        padding: 3rem;
        border-radius: 15px;
        box-shadow: 0 10px 30px rgba(0,0,0,0.15);
    }
    
    .module-card {
        background: white;
        padding: 2rem;
        border-radius: 15px;
        text-align: center;
        box-shadow: 0 4px 12px rgba(0,0,0,0.1);
        cursor: pointer;
        transition: all 0.3s;
        border: 3px solid transparent;
    }
    
    .module-card:hover {
        border-color: #ff6b35;
        transform: translateY(-5px);
        box-shadow: 0 8px 20px rgba(255, 107, 53, 0.3);
    }
    
    .module-card.active {
        border-color: #1a4d3e;
        background: #f0f8f5;
    }
    
    .warning-box {
        background: #fff3cd !important;
        border-left: 5px solid #ffc107;
        padding: 1.5rem;
        border-radius: 8px;
        margin: 1rem 0;
        color: #212529 !important;
    }
    
    .warning-box * { color: #212529 !important; }
    
    .info-box {
        background: #d1ecf1 !important;
        border-left: 5px solid #17a2b8;
        padding: 1.5rem;
        border-radius: 8px;
        margin: 1rem 0;
        color: #212529 !important;
    }
    
    .info-box * { color: #212529 !important; }
    
    .success-box {
        background: #d4edda !important;
        border-left: 5px solid #28a745;
        padding: 1.5rem;
        border-radius: 8px;
        margin: 1rem 0;
        color: #212529 !important;
    }
    
    .success-box * { color: #212529 !important; }
    
    .metric-card {
        background: white !important;
        padding: 2rem;
        border-radius: 15px;
        text-align: center;
        box-shadow: 0 4px 12px rgba(0,0,0,0.1);
        border-left: 5px solid #ff6b35;
        color: #212529 !important;
    }
    
    .metric-card * { color: inherit !important; }
    
    .stButton>button {
        width: 100%;
        background: linear-gradient(135deg, #ff6b35 0%, #ff8c61 100%);
        color: white;
        border-radius: 25px;
        padding: 0.75rem 2rem;
        font-weight: bold;
        border: none;
    }
    
    .stButton>button:hover {
        background: linear-gradient(135deg, #ff8c61 0%, #ff6b35 100%);
        transform: translateY(-2px);
    }
    
    h1, h2, h3 {
        color: #1a4d3e;
    }
    
    .logo-container {
        text-align: center;
        margin-bottom: 2rem;
    }
</style>
""", unsafe_allow_html=True)

# Initialisation des variables de session
if 'show_welcome' not in st.session_state:
    st.session_state.show_welcome = True
if 'selected_module' not in st.session_state:
    st.session_state.selected_module = None          # module en cours d'utilisation
if 'pending_module' not in st.session_state:
    st.session_state.pending_module = None           # module en attente de login
if 'module_auth' not in st.session_state:
    st.session_state.module_auth = {}                # {module_name: True/False}
# Compat : certaines fonctions utilisent encore authenticated / username
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False
if 'username' not in st.session_state:
    st.session_state.username = ''
if 'uploaded_files_bam' not in st.session_state:
    st.session_state.uploaded_files_bam = []

# ── Restauration des données persistantes au démarrage ───────────────────────
if 'combined_data_bam' not in st.session_state:
    cached = load_cache("combined_data_bam")
    st.session_state.combined_data_bam = cached   # None si rien en cache
if 'saham_pdm_localites' not in st.session_state:
    cached_pdm = load_cache("saham_pdm_localites")
    st.session_state.saham_pdm_localites = cached_pdm
if 'saham_credits_depots_raw' not in st.session_state:
    cached_raw = load_cache("saham_credits_depots_raw")
    st.session_state.saham_credits_depots_raw = cached_raw
# ─────────────────────────────────────────────────────────────────────────────

if 'cleaning_report_bam' not in st.session_state:
    st.session_state.cleaning_report_bam = None
if 'processing_done' not in st.session_state:
    st.session_state.processing_done = False

# Variables pour Visualisations Saham Bank
if 'saham_referentiel' not in st.session_state:
    st.session_state.saham_referentiel = None
if 'saham_financial' not in st.session_state:
    st.session_state.saham_financial = None
if 'saham_aggregated' not in st.session_state:
    st.session_state.saham_aggregated = None

# Variables pour les totaux BAM (référence pour PDM)
if 'total_depots_bam' not in st.session_state:
    st.session_state.total_depots_bam = None
if 'total_credits_bam' not in st.session_state:
    st.session_state.total_credits_bam = None

# Variables pour les 3 nouveaux fichiers Saham (module PDM)
if 'saham_credits_depots_raw' not in st.session_state:
    st.session_state.saham_credits_depots_raw = None
if 'saham_mapping_chapitre' not in st.session_state:
    cached_mc = load_cache("saham_mapping_chapitre")
    st.session_state.saham_mapping_chapitre = cached_mc
if 'saham_mapping_localites_agences' not in st.session_state:
    cached_mla = load_cache("saham_mapping_localites_agences")
    st.session_state.saham_mapping_localites_agences = cached_mla
if 'saham_pdm_localites' not in st.session_state:
    st.session_state.saham_pdm_localites = None
# Import Saham multi-années
if 'saham_years_data' not in st.session_state:
    st.session_state.saham_years_data = {}   # {annee: DataFrame}


# Fonctions de calcul des totaux
def calculate_total_depots(df):
    """Calcule le total des dépôts"""
    if 'Montant_Depots' in df.columns:
        return df['Montant_Depots'].sum()
    return 0

def calculate_total_credits(df):
    """Calcule le total des crédits"""
    if 'Montant_Credits' in df.columns:
        return df['Montant_Credits'].sum()
    return 0

def calculate_total_guichets(df):
    """Calcule le total des guichets"""
    if 'Nombre_Guichets' in df.columns:
        return df['Nombre_Guichets'].sum()
    return 0

# Logo Saham encodé en base64
SAHAM_LOGO_BASE64 = "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCAHGAekDASIAAhEBAxEB/8QAHAABAAIDAQEBAAAAAAAAAAAAAAEGAgMEBQcI/8QAMhAAAgEDBAEEAQMDBAMBAQAAAAECAwQRBRIhMUEGEyJRYRQycSNCgTNScpFlYsEkU//EABwBAQABBQEBAAAAAAAAAAAAAAABAgQFBgcDCP/EAC0RAAICAgICAgICAgICAwEAAAABAgMEEQUhBhITMSJBFFEjMkJhM4EHFTRi/9oADAMBAAIRAxEAPwD8/gA9TCAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABKAABUiH30AGmu+AH0UgAyUCuKTIBLIKGyWtAAAAAAAADQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABKAABUh/RCJBHy+itIr0SACdAAAgkAApYAAIIYYSb46BDG9BfZvt7SnUTTqNzzwmb5aVeRhuVPdF9YPFCU11LB6KF9dUf2VX/nk8X7N9F7VKv6maalOpST305Qx9oxhlrODsUNXjU+N5ShNfeD2U46PdR2xqe3J9LBQ7XEyEONqv7gV0HdudArOLnbyVRfycm4srqg/nSaIWTH9lN3E2Vr8Vs0APKeGsA942Rl9GJnTOL76AAPX1etni5JfZDIJIKHLQAAJT2SkAACQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACSlPb0AAh1+yt1+wcEyFFLoz7eDdSt60lxbzl9cFMrqo/bPenAuseoxPOTk6FPTLyssU6Oz/AJI2PQr/AGvO3rwWzzq/7L5cFmv6iclkI98tPr26fu05NfweTbHe1FP/ACe1V8LFtMtreKyav91oxYIqZTwQj3Mfpp6ZkAClnowACCAGAEQQlgkBE9FLi2OFywpx7jwyVj8B7fGCZRi19HvVfdV/qz2WGp3dnLNOpJp9pvJ2rfXqNd7LqjFL+CsEp+GW1mLGS6Mvi83OvqzsuMdM0zUI7qTSf1k5l96erUpP2luXg5Nvd1beWadRr/J3dL9RuLUa/wAl+TG2V219xNjquwMxf5OmcG4tatB/OLRpL46lhqNPDjHL/By73025pytZxx9ZJr5CcXqw8MnxyN0fensqxDaPbeafc2T/AKlNy/hHhk5OWNrX+DK131TXTNWyuOvxpalEkDDBX9/RZNa+wABop0wAwuSRvS2AAQQ2AACoAAAAAAAAAAAAAAAAAAAAAAAAAAAAAq0AACjZDIbwTBOXRKx5DhUk8Q/b5EpxitnrTRObSiTODisvH+Hk9em6dXvJKUYP20/k3xwdTQdGjUpqpXXHhMsEFTpUnSpU1FPgwGZyvr+MToHBeJW3tWWLo5lto9lSUZyk5TXjB04OlCKjCnHj8GtxaJzg1+3PnN9M6pieO4lFa0uzOU5PhLAzLHZhuMslg7bN7bMisampaUScxxunBTzxhnF1/TKfs+/brbLto7E+hUpqvazg+MLsyWFlzUktmv8APcXTZjys12fPsybe7tPAPReUlTuZxXhmnBveNJSqTPnrKqcMiQAaBUUN7AABAAAADAH0U7MGvslLjgywCdjYABGyJQ9jBrkknAwJdoivdb2mbaNzcUpJ0ptY8Z7LFo+vJJQuouLf9yZWYiTeeyxuwq7V39myYHkN2M1H9H0aEqF1T+E6c448pNnK1DRKFfM6K2z+nwVeyvrm1knTqPC8Fr0bWqN1DFw1Gp4MPbi248tr6N7xeVweSgoWrTKzf6bdWsvnCTj9pZPGfRKlWnNbXTU4s5eo6JaV4Opb/CX0e9PK+v4ssOS8R6+WnvZTyM8nsvLC5t5NOm2vtHhaafJmMfIjb9HP8vArl9FoVnUlBp9rDPPv9un5PD9Q6wof9Ig+u2XODiOclvtGK5zmqsOtkZk5SYmTGz7NJMS0vHJxNR1xPNC2+KXGTgzk5ScpPLfk26hbyt7qVKSxzwedm8YVEHBNHz9z/I5U8mUJPol4wQQmSZb6NaT/ZD7AfYJKWwACnYAAIYAACIb0A/lwuMAiXC58kN6RXCDlJJI9OiWU728UF+1csu7UaVGFvFL4+TnemrRWtgq8lzLydFJtObNP5PL9n6HcPEOJ/j1K6SMZPa1jvyzi+rbvdttabSS7SOxf1Y0LGVRvEn0Ui7qyrV5VG+Wz04nEf+zRYea8y4f4kYriGF0RjAX+nFfk32ttUua0adOLll8/wbPc41V72cvxMeWbb2j1aFYyvLpPD2xfJcJx2RjSXSNGl0adjbunHiTwemTSjvb4NSzcid09I7L43xUOPo+SX2jwazdK0s5RT+T6KdKbk5Tk+ZM6XqO7VxcbYyyonKM3xeH6w2znHlPLvMyZQT+jHySiQZlrSNOgvySOh6bt3W1DbN5T8M9/qXTZ209umm4fg2ek7fNX3scqPB3JydWOypHcvBrV/IOu3R1Pi/GFm4bZQMY7Hg7etaTOi3Xox3RfePBw8OTePDMziZcbEmaJynDzwZuCQbMU3ntmWOcEwjyi82m+jBVRlvTMqjy0k/B3vR0X+pqPHgr8Ibazl9otXpOnKFOpOSxFvgxHJz1Bm5eHYjszdnVb5f8k5RhLmbwQkaLL/Zn0LWvjjFGXDM6c5Q/azCKMsEqxx+iHUp/wC30Zu4lnPT/Brq1JVP3PKIxySl9k/LJlPwUx7SMecE54ElwYxjkp+/s9ox93s220cTnUb4USl69VdW+qSz/cXVONO0qyk8LBQb6W+6qNcrcbPw9O3tHJvP8pacNmlNmRik8k4NrcfVI5EpeyRIAKSQAAAAAAACpMhpv6DM6dJ1JRiouTf14FGlOrNKCzyWvRNKo2dP9TcZlN9LwjG5uSqU+zZeA8fuzbU9dDSNLo2tONSqt82v+jpVJ1G1tfx6wTNqSUuE/peDGPDf5NPzcqdjO5cTwVWHCK/ownnc8snPAx8l9eSJqcZ4xwY7tmadkHNxTJYi2pLBktvkh7dywSm0ymUfaDicj1Tab3SuIrpfJorc1jJfa1JV7apRks7lwUi9oOjXlTkumbbxGZ7fjI4r5fxEq7HbFfZ5UZRIaJRssmn9HN1v9kPsjkl9gjY7AAIJQAAJABjJ4JKXHfZkezSqH6mts2N4Z4qbcm/4LZ6OtVTo/qai4ksox/I5Kphs2XxjBeZkrr6OtXhsoQpQWElyZUecLx5FeSlM1Vaqo29SpJ4xE0iPtffs7tZH+LhvfWkcH1VeJ1Vb03wV9rBnc1HXvJ1G8rIowq16mymss3bGiqqezhHKZc+RypRMqFCdWsqcFlly0XT6dnQ9ySW9rk8ug6X7DVetxNcpfZ0pSdSq85X0kYXPz9pxR0LxTxnpXWL6GHKfXB4tfvI2lqqKeZz4TXg6FarGjZznJpSiUrU7uV7cOb/auEW3FUSvl7NF75Ty6xaHVW9HlqJ5bby88s25FSeODGL5Nzrh6LSOJX3+8nJ/bMyYRc5KK7bMYczSO56a0517j3qqeyL4PDKyIVRe2ZPheMsy7l6rZ3NMoSs9Jg3H5s9UeOTO4qSknSwsRfBrW7HJoOXa5zcon0Tw2EsWhRn0ZpwnmE1mL7OPquhQk3UtMRzzg6knjozpSk+X0emJmyq6LXmOBozItxXZRbq1qUqjhUi4NcZaPMoyVRLcmmfQbqnp9xHZWpuT+znrQ7GU8rdFfhmxVcvFLs5jm+C3uftD6K1Z21S5qKMYvvBdrelG3tYUo4ylyzTQoWtm1KjBuaN9KMp5k3jPgw/I8grPo3PxzxtcdBTn9mOOTOUYwoupKSSJccfk5Pqi5lRoxpR43Fhj4ztl0bDzPKQw6tyOhb1adbOySyvBtwvLKfpV7OzuFNtyi+0y1U6sLqiqlKX+C6y8D4o7SMVw/kdWfL42zY8E9I1xyuw5MxiWjbUk3r9EykKcsvCQSybqSVLdUl1FFcK3OSRb5mRHDg22cr1HcOja+zGXL7Ki++Toa7duveSa6yc5PJvnGY3wVpnz/wCXcj/JvaRIAMnLtGowjpAAHmVAAEgAAABJtpLlsJZjnwdf03aK4rqq+VBlrlX/ABGV4bAlm3qCOl6d0tUKfu3Ec55WTsTfwwlnHSJrVZVEotYS4MM+DS+RyZTn0z6A4TjYYlSSXZjGMd258TfaNkUY5MovLwY+MvaXZmcibjA8GvXX6alBUntbfOPJ7qFVXNrCsvKKv6inOep7Gmoro63pe4dazdu+4ftRnJ4UVT7o0fF5Xee65PR75kRM5RaWWuOiF0YGXTN8Uk47RlTbycT1ZabmrmisJfuSO5AxlThVhKE+mZDCv+KWzX+c4v8AmYsnFdooT6IPVq1CVteShJYXaPKjeMa1WQ2j555GiWNc4NEPsEsjkuVos9AAE+oAAI0yGPBi+Q8iCcpYIfRVBe79Ub7Kg61anTj3KWH/AAXylSjbW8beH7YLCSK56UtY1K8q0+qfn8ljqVN8d65yalzWV7P1Oy+FcYqalJrs1c7sZOb6muMWKpQ+M2+f4OtSW6cX9FT9RXLr6k34isFtxVW5dmZ8z5F4+MoJnLVOqkoQjulJ9lt9PabTt6EbirFKb5OZodfT7Z77nbOfa/B2oaxp0pucquPpGbzHOK1FGicBVgxn8tr7PbXk8xSzhcoxhJKeZLLOfca7aRhNUYylJrhmzRr+jqWdy2Shw19muW4du/ZnScXnMWa+Kt6MtXsa93FOlUcVjlLyV240q5oxcYxcn94Lk5+2pKKwn0jUqkvKLrD5NYq9Gix5Xxqvkdy39lClZXKniVJ/zg3UdPvKs1CnQ4f9zRd24t8wX/RP6j2/jTpr+S6nzia0ka/X/wDHsE+2V/TdBk6i95Ljl5O9SjTtqKpU0kka5VakpZkzbRpe4s5MVmZ0shG1cT43Hjn7IxcnFOb67PJZ61b17p27iljtmfqOurSw29NrBSqUpJqpB4aMlh8YradmD5vyaVGTGuL6Rf5pOTa5j4GWljPB4fT2oU7q39mrJKaXB7sS3NSWOeDEZWNKqX0bdw/I15dakn2a8cmcVlNPolxwRnksnsy83KS0S4pLoJtDIyFt9Eykkvy/RlGWeypa9dyurp7ZYUXg7ut3StbZPOJMptWe6UmvLybfxeN6pSaOR+c8rCT+OBluPdpWpSsqq3ZcH4OfFkz5M5kY0LY60c4wM+zEmrIvsvdtcUbukqlJrP0ZbKninko1rXuaEk6NVxO5Zeo69COKsHJ47Nby+He/xOqcN5zBRSt+zv0ISdRRfx/k5HqXU1Szb0mk/LTObqGvV7lvEXHJyKilUk6k5ZbPbB4r1aciw8h8uqvi1Wxu9xuT7YxglYS4IZsij6rRy7Kt+efvsAAnZ5IAAN7JAAIAAABstlhRi12y5aJbU7Sw9yS5kVOyxUuKSx0+S73UVCjTiutqMBztjj9HTfA8KM7PkaMZSTWfswXYim4oYNNlJt9nZYuMW0iTOmsSzkwMZZJh09kTrVq0cn1NYVpVI3VJJxXeDwaFd/ptQhuUopvDLRCutvt1Ipr8nnu9Ntrlbqa2T+0ZevOfxuEjRszxz60GjUL+k3w8rrBgzY4pp5i0zCUE/JrspeWI/lKSiSYJmSKVW1ssj0RfxWQkmY/k8tt0XTD7Njq1Ixcl+Tywco9lgpyeMSWMfZ5qjy2/spJ4pJx02J8kdY/hGSIJ4J6Nx/QAAhFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAB6NAAAGwAASSgAASDIAAhAAAkpCDYARsyAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACdgHgjJJC5eQSXRsAAJ7GiJ/Y+gB1H/2Q=="

# Page d'accueil
def welcome_page():
    """Page d'accueil sans logo"""
    # Masquer la sidebar
    st.markdown("""
        <style>
        [data-testid="stSidebar"] {
            display: none;
        }
        .welcome-header {
            background: linear-gradient(135deg, #1a4d3e 0%, #2d5f4e 100%);
            padding: 3rem;
            border-radius: 20px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.2);
            text-align: center;
        }
        .welcome-title {
            color: white !important;
            font-size: 3em;
            margin: 0;
            font-weight: bold;
        }
        .welcome-subtitle {
            color: #ff6b35;
            font-size: 1.2em;
            margin-top: 1rem;
        }
        .info-section {
            text-align: center;
            margin: 2rem 0;
        }
        </style>
        """, unsafe_allow_html=True)
    
    # Centrer le contenu
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        # En-tête avec fond vert
        st.markdown("""
        <div class="welcome-header">
            <h1 class="welcome-title">SAHAM BANK</h1>
            <p class="welcome-subtitle">سهام بنك</p>
            <p style="color: white; font-size: 1em; margin-top: 1rem; font-style: italic;">
                Accélérateur de vos ambitions
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("<br><br>", unsafe_allow_html=True)
        
        # Titre principal
        st.markdown("""
        <div class="info-section">
            <h2 style="color: #ff6b35;">Application de Structuration de Données</h2>
        </div>
        """, unsafe_allow_html=True)
        
        # Description
        st.markdown("""
        <div class="info-section">
            <p style="color: #6c757d; font-size: 1.1em;">Direction Financière - PFE 2026</p>
        </div>
        """, unsafe_allow_html=True)
        
        # Section modules
        st.markdown("""
        <div class="info-section">
            <p style="color: #1a4d3e; font-size: 1.2em; font-weight: bold;">
                Structuration des bases de données bancaires
            </p>
            <p style="color: #6c757d; font-size: 1.1em;">
                BAM • GPBM • Balance • Établissements de Crédit
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("<br><br>", unsafe_allow_html=True)
        
        # Bouton Commencer
        btn_col1, btn_col2, btn_col3 = st.columns([1, 2, 1])
        with btn_col2:
            if st.button("Commencer", use_container_width=True, type="primary", key="welcome_start"):
                st.session_state.show_welcome = False
                st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# CONNEXION PAR MODULE
# ─────────────────────────────────────────────────────────────────────────────

# Identifiants spécifiques à chaque module
MODULE_CREDENTIALS = {
    "BAM": {
        "users": {"bam": "bam2026", "admin": "admin"},
        "label": "Module BAM — Données Banque Al-Maghrib",
        "icon": "🏦"
    },
    "Balance": {
        "users": {"balance": "balance2026", "admin": "admin"},
        "label": "Module Balance — Produits & Charges",
        "icon": "📊"
    },
}

def module_login_page(module_name):
    """Affiche la page de connexion propre à un module."""
    creds = MODULE_CREDENTIALS.get(module_name, {})
    label = creds.get("label", module_name)
    icon  = creds.get("icon", "🔐")

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown(f"""
        <div class="login-box">
            <div style="text-align: center; margin-bottom: 2rem;">
                <div style="background: #1a4d3e; padding: 1.5rem; border-radius: 10px; margin-bottom: 1.5rem;">
                    <h1 style="color: white; margin: 0; font-size: 2em;">SAHAM BANK</h1>
                    <div style="color: #ff6b35; font-size: 0.8em; margin-top: 0.5rem;">سهام بنك</div>
                </div>
                <p style="color: #ff6b35; font-size: 1.2em; font-weight: bold;">
                    {icon} {label}
                </p>
                <p style="color: #6c757d; font-size: 0.9em;">Accès sécurisé — veuillez vous authentifier</p>
            </div>
        """, unsafe_allow_html=True)

        st.markdown("### Connexion")
        username = st.text_input("Identifiant", placeholder="Entrez votre identifiant",
                                 key=f"login_user_{module_name}")
        password = st.text_input("Mot de passe", type="password",
                                 placeholder="Entrez votre mot de passe",
                                 key=f"login_pwd_{module_name}")

        col_a, col_b, col_c = st.columns([1, 2, 1])
        with col_b:
            if st.button("Se Connecter", use_container_width=True,
                         key=f"login_btn_{module_name}"):
                users = creds.get("users", {})
                if username.lower() in users and password == users[username.lower()]:
                    st.session_state.module_auth[module_name] = True
                    st.session_state.selected_module = module_name
                    st.session_state.authenticated = True
                    st.session_state.username = username.capitalize()
                    st.session_state.pending_module = None
                    st.rerun()
                else:
                    st.error("❌ Identifiant ou mot de passe incorrect.")

        # ── Identifiants de test ──────────────────────────────────────────
        st.markdown("<br>", unsafe_allow_html=True)
        users_test = creds.get("users", {})
        # Afficher tous les comptes sauf "admin"
        non_admin = {u: p for u, p in users_test.items() if u != "admin"}
        rows_html = "".join(
            f"<tr>"
            f"<td style='padding:3px 10px;'><code style='background:#eef;border-radius:4px;padding:2px 6px;'>{u}</code></td>"
            f"<td style='padding:3px 10px;'><code style='background:#eef;border-radius:4px;padding:2px 6px;'>{p}</code></td>"
            f"</tr>"
            for u, p in non_admin.items()
        )
        st.markdown(f"""
        <div style="margin-top:0.5rem; background:#f8f9fa; border:1px dashed #adb5bd;
                    border-radius:8px; padding:0.9rem 1.2rem; text-align:center;">
            <p style="color:#6c757d; font-size:0.82em; margin:0 0 0.5rem; font-weight:600;
                      letter-spacing:0.04em; text-transform:uppercase;">
                🔑 Identifiants de test
            </p>
            <table style="margin:0 auto; border-collapse:collapse; font-size:0.88em;">
                <thead>
                    <tr>
                        <th style="color:#495057; padding:2px 10px; font-weight:600;">Identifiant</th>
                        <th style="color:#495057; padding:2px 10px; font-weight:600;">Mot de passe</th>
                    </tr>
                </thead>
                <tbody>
                    {rows_html}
                </tbody>
            </table>
        </div>
        """, unsafe_allow_html=True)
        # ─────────────────────────────────────────────────────────────────

        # Bouton retour
        st.markdown("<br>", unsafe_allow_html=True)
        col_x, col_y, col_z = st.columns([1, 2, 1])
        with col_y:
            if st.button("← Retour à la sélection", use_container_width=True,
                         key=f"back_btn_{module_name}"):
                st.session_state.pending_module = None
                st.rerun()

        st.markdown("</div>", unsafe_allow_html=True)



# Page de sélection du module
def module_selection_page():
    st.markdown("""
    <div class="main-header">
        <h1>Application de Structuration de Données</h1>
        <p style="font-size: 0.9em; opacity: 0.9;">Direction Financière — SAHAM BANK — PFE 2026</p>
    </div>
    """, unsafe_allow_html=True)

    st.header("Sélectionnez un Module")

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.markdown("""
        <div class="module-card active">
            <h3 style="color: #1a4d3e; margin-top: 1.5rem;">BAM</h3>
            <p style="color: #6c757d;">Données bancaires</p>
            <p style="color: #6c757d; font-size: 0.9em;">Dépôts, crédits et guichets</p>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Accéder au module BAM", use_container_width=True, type="primary", key="btn_bam"):
            if st.session_state.module_auth.get("BAM"):
                st.session_state.selected_module = "BAM"
            else:
                st.session_state.pending_module = "BAM"
            st.rerun()

    with col2:
        st.markdown("""
        <div class="module-card" style="opacity:0.6;">
            <h3 style="color: #6c757d; margin-top: 1.5rem;">GPBM</h3>
            <p style="color: #6c757d; font-size: 0.9em;">À venir prochainement</p>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        st.button("Accéder au module GPBM", use_container_width=True, disabled=True, key="btn_gpbm")

    with col3:
        st.markdown("""
        <div class="module-card active">
            <h3 style="color: #1a4d3e; margin-top: 1.5rem;">Balance</h3>
            <p style="color: #6c757d;">Produits & Charges</p>
            <p style="color: #6c757d; font-size: 0.9em;">Import et analyse balance</p>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Accéder au module Balance", use_container_width=True, type="primary", key="btn_balance"):
            if st.session_state.module_auth.get("Balance"):
                st.session_state.selected_module = "Balance"
            else:
                st.session_state.pending_module = "Balance"
            st.rerun()

    with col4:
        st.markdown("""
        <div class="module-card" style="opacity:0.6;">
            <h3 style="color: #6c757d; margin-top: 1.5rem;">Établissements<br>de Crédit</h3>
            <p style="color: #6c757d; font-size: 0.9em;">À venir prochainement</p>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        st.button("Accéder au module EC", use_container_width=True, disabled=True, key="btn_ec")



def get_month_from_filename(filename):
    months_fr = {
        'janvier': 1, 'fevrier': 2, 'mars': 3, 'avril': 4, 'mai': 5, 'juin': 6,
        'juillet': 7, 'aout': 8, 'septembre': 9, 'octobre': 10, 'novembre': 11, 'decembre': 12,
        'jan': 1, 'fev': 2, 'mar': 3, 'avr': 4, 'jun': 6, 'jul': 7, 'aou': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
    }
    
    filename_lower = filename.lower()
    for month_name, month_num in months_fr.items():
        if month_name in filename_lower:
            return month_num
    
    import re
    month_pattern = re.search(r'[_\-\s](\d{1,2})[_\-\s\.]', filename)
    if month_pattern:
        month_num = int(month_pattern.group(1))
        if 1 <= month_num <= 12:
            return month_num
    
    return None

def combine_bam_files(files_data):
    combined_df = pd.DataFrame()
    for file_info in files_data:
        df = file_info['data'].copy()
        df['mois'] = file_info['month']
        combined_df = pd.concat([combined_df, df], ignore_index=True)
    return combined_df

def detect_missing_values(df, data_type):
    report = {
        'data_type': data_type,
        'total_rows': len(df),
        'columns': list(df.columns),
        'missing_details': []
    }
    
    for idx, row in df.iterrows():
        missing_cols = []
        for col in df.columns:
            value = row[col]
            if pd.isna(value) or (isinstance(value, str) and value.strip() == ''):
                missing_cols.append(col)
        
        if missing_cols:
            report['missing_details'].append({
                'row_number': idx + 2,
                'excel_row': idx + 2,
                'missing_columns': missing_cols,
                'data_preview': {col: row[col] for col in df.columns if col not in missing_cols}
            })
    
    report['total_missing_rows'] = len(report['missing_details'])
    report['percentage_complete'] = ((len(df) - len(report['missing_details'])) / len(df) * 100) if len(df) > 0 else 0
    
    return report

def convert_df_to_excel(df, filename):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Data')
    output.seek(0)
    return output

def normalize_bam_columns(df):
    """Normalise les noms de colonnes BAM - version ultra-robuste"""
    df_normalized = df.copy()
    
    # Nettoyer d'abord tous les noms de colonnes
    df_normalized.columns = df_normalized.columns.str.strip()
    
    # Nouvelle stratégie : chercher les colonnes par mot-clé
    new_columns = {}
    
    for col in df_normalized.columns:
        col_lower = col.lower()
        
        # Recherche par mot-clé
        if 'code' in col_lower and 'localit' in col_lower:
            new_columns[col] = 'Code_Localite'
        elif 'localit' in col_lower and 'code' not in col_lower:
            new_columns[col] = 'Localite'
        elif 'nombre' in col_lower and 'guichet' in col_lower:
            new_columns[col] = 'Nombre_Guichets'
        elif 'montant' in col_lower and 'd' in col_lower and 'p' in col_lower:
            # Montant des dépôts
            new_columns[col] = 'Montant_Depots'
        elif 'montant' in col_lower and 'cr' in col_lower:
            # Montant des crédits
            new_columns[col] = 'Montant_Credits'
        elif col_lower == 'mois':
            new_columns[col] = 'mois'
    
    # Renommer
    df_normalized.rename(columns=new_columns, inplace=True)
    
    return df_normalized

def clean_numeric_columns(df):
    """Nettoie les colonnes numériques - version ultra-robuste"""
    df_clean = df.copy()
    numeric_cols = ['Montant_Depots', 'Montant_Credits', 'Nombre_Guichets']
    
    for col in numeric_cols:
        if col in df_clean.columns:
            # Convertir en string
            df_clean[col] = df_clean[col].astype(str)
            
            # Enlever TOUS les types d'espaces (normaux, insécables, tabs, etc.)
            df_clean[col] = df_clean[col].str.replace(r'\s+', '', regex=True)
            
            # Enlever les caractères non-numériques sauf point et virgule
            df_clean[col] = df_clean[col].str.replace(r'[^\d.,]', '', regex=True)
            
            # Remplacer virgule par point pour décimales
            df_clean[col] = df_clean[col].str.replace(',', '.')
            
            # Convertir en numérique (NaN si impossible)
            df_clean[col] = pd.to_numeric(df_clean[col], errors='coerce')
            
            # Remplacer NaN par 0
            df_clean[col] = df_clean[col].fillna(0)
    
    return df_clean

def add_direction_regionale(df):
    direction_mapping = {
        'CASABLANCA': 'CASA', 'RABAT': 'CASA', 'SALE': 'CASA', 'MOHAMMEDIA': 'CASA',
        'TEMARA': 'CASA', 'KENITRA': 'CASA', 'SETTAT': 'CASA', 'BERRECHID': 'CASA',
        'FES': 'OUEST', 'MEKNES': 'OUEST', 'TAZA': 'OUEST', 'OUJDA': 'OUEST',
        'BERKANE': 'OUEST', 'NADOR': 'OUEST', 'AL HOCEIMA': 'OUEST',
        'MARRAKECH': 'SUD', 'AGADIR': 'SUD', 'ESSAOUIRA': 'SUD', 'OUARZAZATE': 'SUD',
        'TANGER': 'NORD', 'TETOUAN': 'NORD',
    }
    
    df_with_region = df.copy()
    df_with_region['DirectionRegionale'] = df_with_region['Localite'].apply(
        lambda x: direction_mapping.get(x.upper() if isinstance(x, str) else '', 'OUEST')
    )
    
    return df_with_region

# VISUALISATIONS
def create_powerbi_visualizations(df):
    """Visualisations des données BAM"""
    
    try:
        # IMPORTANT : Normaliser et nettoyer les données
        df = normalize_bam_columns(df)
        df = clean_numeric_columns(df)
        
        # DIAGNOSTIC : Expander pour vérifier les données
        with st.expander("🔍 Diagnostic des Données", expanded=False):
            st.write("**Colonnes du DataFrame :**")
            st.write(df.columns.tolist())
            
            st.write("**Types des colonnes numériques :**")
            if 'Montant_Depots' in df.columns:
                st.write(f"- Montant_Depots : {df['Montant_Depots'].dtype}")
            if 'Montant_Credits' in df.columns:
                st.write(f"- Montant_Credits : {df['Montant_Credits'].dtype}")
            if 'Nombre_Guichets' in df.columns:
                st.write(f"- Nombre_Guichets : {df['Nombre_Guichets'].dtype}")
            
            st.write("**Aperçu des données :**")
            st.dataframe(df.head(10))
            
            st.write("**Totaux calculés :**")
            if 'Montant_Depots' in df.columns:
                st.write(f"- Total Dépôts : {df['Montant_Depots'].sum()/1e6:.2f} Md")
            if 'Montant_Credits' in df.columns:
                st.write(f"- Total Crédits : {df['Montant_Credits'].sum()/1e6:.2f} Md")
            if 'Nombre_Guichets' in df.columns:
                st.write(f"- Total Guichets : {df['Nombre_Guichets'].sum():,.0f}".replace(',', ' '))
        
        # Vérifier que les colonnes nécessaires existent
        required_cols = ['Montant_Depots', 'Montant_Credits', 'Nombre_Guichets', 'Localite']
        missing_cols = [col for col in required_cols if col not in df.columns]
        
        if missing_cols:
            st.error(f"Colonnes manquantes : {missing_cols}")
            st.write("Colonnes disponibles :", df.columns.tolist())
            return
        
        st.header("Vue Générale")
        
        # Calculer les totaux
        # BAM en milliers de DH → /1e6 pour afficher en Md DH
        total_depots = df['Montant_Depots'].sum() / 1e6
        total_credits = df['Montant_Credits'].sum() / 1e6
        total_guichets = df['Nombre_Guichets'].sum() / 1000
        
        # GRANDES MÉTRIQUES
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown(f"""
            <div class="metric-card">
                <p class="metric-value">{total_depots:.0f}Md</p>
                <p class="metric-label">Total Dépôts</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class="metric-card">
                <p class="metric-value">{total_credits:.0f}Md</p>
                <p class="metric-label">Total Crédits</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div class="metric-card">
                <p class="metric-value">{total_guichets:.0f}K</p>
                <p class="metric-label">Total Guichets</p>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        # FILTRES
        st.subheader("Filtres")
        
        df_filtered = df.copy()
        
        if 'DirectionRegionale' in df.columns:
            all_regions = ['(Tous)'] + sorted(df['DirectionRegionale'].unique().tolist())
            selected_regions = st.multiselect(
                "Direction Régionale",
                options=all_regions,
                default=['(Tous)']
            )
            
            if '(Tous)' not in selected_regions and selected_regions:
                df_filtered = df[df['DirectionRegionale'].isin(selected_regions)]
        
        st.markdown("---")
        
        # TABLEAU PAR MOIS (seulement si la colonne mois existe)
        if 'mois' in df_filtered.columns:
            st.subheader("Données par Mois")
            
            monthly_summary = df_filtered.groupby('mois').agg({
                'Montant_Credits': 'sum',
                'Montant_Depots': 'sum',
                'Nombre_Guichets': 'sum'
            }).reset_index()
            
            monthly_summary.columns = ['mois', 'Total Credits', 'Depots', 'Total Guichets']
            
            month_names = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
                          'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
            
            monthly_summary['mois'] = monthly_summary['mois'].apply(
                lambda x: month_names[int(x)-1] if 1 <= int(x) <= 12 else str(x)
            )
            
            # Formater les nombres
            monthly_summary['Total Credits'] = monthly_summary['Total Credits'].apply(
                lambda x: f"{x/1e6:.2f} Md"
            )
            monthly_summary['Depots'] = monthly_summary['Depots'].apply(
                lambda x: f"{x/1e6:.2f} Md"
            )
            monthly_summary['Total Guichets'] = monthly_summary['Total Guichets'].apply(
                lambda x: f"{int(x):,}".replace(',', ' ')
            )
            
            st.dataframe(monthly_summary, use_container_width=True, hide_index=True)
            
            st.markdown("---")
        
        # GRAPHIQUES PRINCIPAUX
        col1, col2 = st.columns(2)
        
        with col1:
            # Graphique ligne par mois
            if 'mois' in df_filtered.columns:
                st.subheader("Dépôts et Crédits par mois")
                
                monthly_data = df_filtered.groupby('mois').agg({
                    'Montant_Depots': 'sum',
                    'Montant_Credits': 'sum'
                }).reset_index()
                
                monthly_data['Depots_Md']  = monthly_data['Montant_Depots']  / 1e6
                monthly_data['Credits_Md'] = monthly_data['Montant_Credits'] / 1e6
                monthly_data['mois_nom'] = monthly_data['mois'].apply(lambda x: str(int(x)))
                
                fig_line = go.Figure()
                
                fig_line.add_trace(go.Scatter(
                    x=monthly_data['mois_nom'],
                    y=monthly_data['Depots_Md'],
                    mode='lines+markers',
                    name='Dépôts (Md)',
                    line=dict(color='#1a4d3e', width=3),
                    marker=dict(size=10),
                    hovertemplate='%{x}<br>Dépôts: %{y:.2f} Md<extra></extra>'
                ))
                
                fig_line.add_trace(go.Scatter(
                    x=monthly_data['mois_nom'],
                    y=monthly_data['Credits_Md'],
                    mode='lines+markers',
                    name='Crédits (Md)',
                    line=dict(color='#ff6b35', width=3),
                    marker=dict(size=10),
                    hovertemplate='%{x}<br>Crédits: %{y:.2f} Md<extra></extra>'
                ))
                
                fig_line.update_layout(
                    xaxis_title='',
                    yaxis_title='Montant (Md MAD)',
                    height=350,
                    hovermode='x unified',
                    showlegend=True,
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
                )
                
                st.plotly_chart(fig_line, use_container_width=True)
            
            # Graphique barres Dépôts par région
            if 'DirectionRegionale' in df_filtered.columns:
                st.subheader("Dépôts par Direction Régionale")
                
                regional_depots = df_filtered.groupby('DirectionRegionale')['Montant_Depots'].sum().reset_index()
                regional_depots['Depots_Md'] = regional_depots['Montant_Depots'] / 1e6
                regional_depots = regional_depots.sort_values('Depots_Md', ascending=False)
                
                fig_bar = px.bar(
                    regional_depots,
                    x='DirectionRegionale',
                    y='Depots_Md',
                    color='Depots_Md',
                    color_continuous_scale='Blues'
                )
                fig_bar.update_traces(hovertemplate='%{x}<br>%{y:.2f} Md<extra></extra>')
                fig_bar.update_layout(
                    xaxis_title='',
                    yaxis_title='Total Dépôts (Md MAD)',
                    height=350,
                    showlegend=False
                )
                
                st.plotly_chart(fig_bar, use_container_width=True)
        
        with col2:
            # Camembert Crédits par région
            if 'DirectionRegionale' in df_filtered.columns:
                st.subheader("Crédits par Direction Régionale")
                
                regional_credits = df_filtered.groupby('DirectionRegionale')['Montant_Credits'].sum().reset_index()
                regional_credits['Credits_Md'] = regional_credits['Montant_Credits'] / 1e6
                
                fig_pie = px.pie(
                    regional_credits,
                    values='Credits_Md',
                    names='DirectionRegionale',
                    color_discrete_sequence=px.colors.sequential.RdBu
                )
                
                fig_pie.update_layout(height=350)
                fig_pie.update_traces(textposition='inside', textinfo='percent+label')
                
                st.plotly_chart(fig_pie, use_container_width=True)
            
            # Graphique barres Guichets par région
            if 'DirectionRegionale' in df_filtered.columns:
                st.subheader("Guichets par Direction Régionale")
                
                regional_guichets = df_filtered.groupby('DirectionRegionale')['Nombre_Guichets'].sum().reset_index()
                regional_guichets = regional_guichets.sort_values('Nombre_Guichets', ascending=False)
                
                fig_guichets = px.bar(
                    regional_guichets,
                    x='DirectionRegionale',
                    y='Nombre_Guichets',
                    color='Nombre_Guichets',
                    color_continuous_scale='Oranges'
                )
                
                fig_guichets.update_layout(
                    xaxis_title='',
                    yaxis_title='Total Guichets',
                    height=350,
                    showlegend=False
                )
                
                st.plotly_chart(fig_guichets, use_container_width=True)
        
        # PAR DIRECTION RÉGIONALE
        if 'DirectionRegionale' in df_filtered.columns and 'mois' in df_filtered.columns:
            st.markdown("---")
            st.header("Par Direction Régionale")
            
            col1, col2 = st.columns(2)
            with col1:
                selected_months = st.multiselect(
                    "Sélectionner les mois",
                    options=sorted(df_filtered['mois'].unique()),
                    default=sorted(df_filtered['mois'].unique()),
                    format_func=lambda x: ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
                                           'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre'][int(x)-1]
                )
            
            df_month_filtered = df_filtered[df_filtered['mois'].isin(selected_months)]
            
            st.subheader("Tableau Récapitulatif")
            
            detailed_summary = df_month_filtered.groupby(['DirectionRegionale', 'mois']).agg({
                'Montant_Depots': 'sum',
                'Montant_Credits': 'sum',
                'Nombre_Guichets': 'sum'
            }).reset_index()
            
            detailed_summary.columns = ['DirectionRegionale', 'mois', 'TDepots', 'Total Guichets', 'Total Credits']
            
            month_abbr = ['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Jun',
                         'Jul', 'Aoû', 'Sep', 'Oct', 'Nov', 'Déc']
            
            detailed_summary['mois'] = detailed_summary['mois'].apply(
                lambda x: month_abbr[int(x)-1] if 1 <= int(x) <= 12 else str(x)
            )
            
            st.dataframe(detailed_summary, use_container_width=True, hide_index=True)
            
            st.subheader("Comparaisons")
            col1, col2, col3 = st.columns(3)
            
            regional_summary = df_month_filtered.groupby('DirectionRegionale').agg({
                'Montant_Depots': 'sum',
                'Montant_Credits': 'sum',
                'Nombre_Guichets': 'sum'
            }).reset_index()
            
            regional_summary['Depots_Md']  = regional_summary['Montant_Depots']  / 1e6
            regional_summary['Credits_Md'] = regional_summary['Montant_Credits'] / 1e6

            with col1:
                fig1 = px.bar(regional_summary, x='DirectionRegionale', y='Depots_Md',
                             title='Dépôts (Md MAD)', color_discrete_sequence=['#1a4d3e'])
                fig1.update_traces(hovertemplate='%{x}<br>%{y:.2f} Md<extra></extra>')
                fig1.update_layout(height=300, showlegend=False, xaxis_title='', yaxis_title='Md MAD')
                st.plotly_chart(fig1, use_container_width=True)
            
            with col2:
                fig2 = px.bar(regional_summary, x='DirectionRegionale', y='Credits_Md',
                             title='Crédits (Md MAD)', color_discrete_sequence=['#ff6b35'])
                fig2.update_traces(hovertemplate='%{x}<br>%{y:.2f} Md<extra></extra>')
                fig2.update_layout(height=300, showlegend=False, xaxis_title='', yaxis_title='Md MAD')
                st.plotly_chart(fig2, use_container_width=True)
            
            with col3:
                fig3 = px.bar(regional_summary, x='DirectionRegionale', y='Nombre_Guichets',
                             title='Guichets', color_discrete_sequence=['#17a2b8'])
                fig3.update_layout(height=300, showlegend=False, xaxis_title='', yaxis_title='')
                st.plotly_chart(fig3, use_container_width=True)
        
        # PAR LOCALITÉ
        st.markdown("---")
        st.header("Par Localité")
        
        col1, col2 = st.columns([3, 1])
        with col1:
            st.subheader("Dépôts par Localité")
            
            top_localites = df_filtered.groupby('Localite')['Montant_Depots'].sum().reset_index()
            top_localites['Depots_Md'] = top_localites['Montant_Depots'] / 1e6
            top_localites = top_localites.nlargest(15, 'Depots_Md').sort_values('Depots_Md')
            
            fig_loc = px.bar(
                top_localites,
                y='Localite',
                x='Depots_Md',
                orientation='h',
                color='Depots_Md',
                color_continuous_scale='Viridis'
            )
            fig_loc.update_traces(hovertemplate='%{y}<br>%{x:.2f} Md<extra></extra>')
            fig_loc.update_layout(
                height=600,
                showlegend=False,
                xaxis_title='Total Dépôts (Md MAD)',
                yaxis_title=''
            )
            
            st.plotly_chart(fig_loc, use_container_width=True)
        
        with col2:
            st.write("")
            st.write("")
            st.write("")
            
            st.markdown(f"""
            <div class="metric-card" style="margin-bottom: 1rem;">
                <p class="metric-value">{total_depots:.0f}Md</p>
                <p class="metric-label">Total Dépôts</p>
            </div>
            """, unsafe_allow_html=True)
            
            st.markdown(f"""
            <div class="metric-card" style="margin-bottom: 1rem;">
                <p class="metric-value">{total_credits:.0f}Md</p>
                <p class="metric-label">Total Crédits</p>
            </div>
            """, unsafe_allow_html=True)
            
            st.markdown(f"""
            <div class="metric-card">
                <p class="metric-value">{total_guichets:.0f}K</p>
                <p class="metric-label">Total Guichets</p>
            </div>
            """, unsafe_allow_html=True)
        
        # TABLEAU DÉTAILLÉ
        if 'DirectionRegionale' in df_filtered.columns:
            st.markdown("---")
            st.subheader("Analyse Détaillée")
            
            detailed_table = df_filtered.groupby(['DirectionRegionale', 'Localite']).agg({
                'Montant_Depots': 'sum',
                'Montant_Credits': 'sum',
                'Nombre_Guichets': 'sum'
            }).reset_index()
            
            detailed_table.columns = ['DirectionRegionale', 'Localite', 'Total Dépôts', 'Total Crédits', 'Total Guichets']
            
            total_row = pd.DataFrame([{
                'DirectionRegionale': 'Total',
                'Localite': '',
                'Total Dépôts': detailed_table['Total Dépôts'].sum(),
                'Total Crédits': detailed_table['Total Crédits'].sum(),
                'Total Guichets': detailed_table['Total Guichets'].sum()
            }])
            detailed_table = pd.concat([detailed_table, total_row], ignore_index=True)
            detailed_table['Total Dépôts']  = detailed_table['Total Dépôts'].apply(lambda x: f"{x/1e6:.2f} Md")
            detailed_table['Total Crédits'] = detailed_table['Total Crédits'].apply(lambda x: f"{x/1e6:.2f} Md")
            
            st.dataframe(detailed_table, use_container_width=True, hide_index=True)
    
    except Exception as e:
        st.error(f"Erreur lors de la création des visualisations : {str(e)}")
        st.write("**Colonnes disponibles dans le DataFrame :**")
        st.write(df.columns.tolist())
        st.write("**Aperçu des données :**")
        st.dataframe(df.head())
        st.exception(e)

# ============================================================================
# FONCTIONS POUR VISUALISATIONS SAHAM BANK
# ============================================================================

def normalize_referentiel_columns(df):
    """Normalise les colonnes du référentiel agences"""
    df_normalized = df.copy()
    
    # Nettoyer les noms de colonnes
    df_normalized.columns = df_normalized.columns.str.strip()
    
    # Mapper les variations de noms
    for col in df_normalized.columns:
        col_lower = col.lower()
        
        if 'code' in col_lower and 'agence' in col_lower:
            df_normalized.rename(columns={col: 'Code_Agence'}, inplace=True)
        elif 'code' in col_lower and 'localit' in col_lower:
            df_normalized.rename(columns={col: 'Code_Localite'}, inplace=True)
        elif 'localit' in col_lower and 'code' not in col_lower:
            df_normalized.rename(columns={col: 'Localite'}, inplace=True)
    
    return df_normalized

def normalize_financial_columns(df):
    """Normalise les colonnes des données financières"""
    df_normalized = df.copy()
    
    # Nettoyer les noms de colonnes
    df_normalized.columns = df_normalized.columns.str.strip()
    
    # Mapper les variations de noms
    for col in df_normalized.columns:
        col_lower = col.lower()
        
        if 'p' in col_lower and 'riod' in col_lower:
            df_normalized.rename(columns={col: 'Periode'}, inplace=True)
        elif 'code' in col_lower and 'agence' in col_lower:
            df_normalized.rename(columns={col: 'Code_Agence'}, inplace=True)
        elif 'd' in col_lower and 'p' in col_lower and 't' in col_lower:
            df_normalized.rename(columns={col: 'Depots'}, inplace=True)
        elif 'cr' in col_lower and 'dit' in col_lower:
            df_normalized.rename(columns={col: 'Credits'}, inplace=True)
    
    return df_normalized

def clean_numeric_saham(df, columns):
    """Nettoie les colonnes numériques Saham"""
    df_clean = df.copy()
    
    for col in columns:
        if col in df_clean.columns:
            # Convertir en string
            df_clean[col] = df_clean[col].astype(str)
            # Enlever tous les espaces
            df_clean[col] = df_clean[col].str.replace(r'\s+', '', regex=True)
            # Enlever caractères non-numériques sauf point et virgule
            df_clean[col] = df_clean[col].str.replace(r'[^\d.,]', '', regex=True)
            # Remplacer virgule par point
            df_clean[col] = df_clean[col].str.replace(',', '.')
            # Convertir en numérique
            df_clean[col] = pd.to_numeric(df_clean[col], errors='coerce')
            # Remplacer NaN par 0
            df_clean[col] = df_clean[col].fillna(0)
    
    return df_clean

# ============================================================================
# NOUVELLES FONCTIONS SAHAM - TRAITEMENT PDM PAR LOCALITE (3 FICHIERS)
# ============================================================================


# ═══════════════════════════════════════════════════════════════════════════════
# FONCTIONS TRAITEMENT SAHAM PDM — formats réels des 3 fichiers
# ═══════════════════════════════════════════════════════════════════════════════

def lire_fichier_principal_saham(raw_str):
    """
    Lit le fichier principal Saham Bank (TXT/CSV) au format :
      Ligne 1  : "Rapport 1"              → ignorée
      Ligne 2  : vide                     → ignorée
      Ligne 3  : en-tête avec Débit/Crédit → ignorée (nb champs différent des données)
      Données  : "2 015";"1";"00001";"121302";"-24 276,29";"80 319,32"
        • séparateur : ;   • guillemets : "
        • milliers   : espace  → "-24 276,29" = -24276.29
        • décimale   : virgule

    Colonnes produites (par position) :
        Annee | Col_2 | Code_Agence | Chapitre | Debit | [cols…] | Credit
    """
    import io

    lines = raw_str.splitlines()

    # 1. Trouver la 1ère ligne de DONNÉES (1er champ = chiffre)
    data_start = None
    for i, line in enumerate(lines):
        stripped = line.strip()
        if not stripped:
            continue
        first_field = stripped.split(';')[0].strip('"').replace(' ', '')
        if first_field and (first_field[0].isdigit() or
                            (first_field[0] == '-' and len(first_field) > 1 and first_field[1].isdigit())):
            data_start = i
            break

    if data_start is None:
        raise ValueError("Impossible de détecter les lignes de données dans le fichier TXT.")

    # 2. Lire SANS header (header=None) — évite le décalage dû à l'en-tête à 7 champs
    data_str = '\n'.join(lines[data_start:])
    df = pd.read_csv(
        io.StringIO(data_str),
        sep=';',
        quotechar='"',
        header=None,
        dtype=str,
        low_memory=False
    )
    df = df.dropna(how='all').reset_index(drop=True)

    # 3. Nommer les colonnes par position
    nb_cols = len(df.columns)
    if nb_cols < 5:
        raise ValueError(f"Fichier trop peu de colonnes ({nb_cols}). Minimum attendu : 5.")

    new_names = []
    for i in range(nb_cols):
        if   i == 0:          new_names.append('Annee')
        elif i == 1:          new_names.append('Col_2')
        elif i == 2:          new_names.append('Code_Agence')
        elif i == 3:          new_names.append('Chapitre')
        elif i == 4:          new_names.append('Debit')
        elif i == nb_cols-1:  new_names.append('Credit')
        else:                 new_names.append(f'Col_{i+1}')
    df.columns = new_names

    # 4. Nettoyer les montants (espaces milliers + virgule décimale)
    def clean_montant(val):
        if pd.isna(val) or str(val).strip() in ('', 'nan', 'None'):
            return 0.0
        s = (str(val).strip().strip('"')
             .replace(' ', '').replace('\xa0', '').replace('\u202f', '').replace('\u00a0', '')
             .replace(',', '.'))
        s = ''.join(c for c in s if c.isdigit() or c in '.-')
        try:
            return float(s) if s not in ('', '-', '.') else 0.0
        except ValueError:
            return 0.0

    for col in ['Debit', 'Credit']:
        if col in df.columns:
            df[col] = df[col].apply(clean_montant)

    # 5. Nettoyer Code_Agence et Chapitre
    for col in ['Code_Agence', 'Chapitre']:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip().str.strip('"').str.strip()

    # 6. Supprimer les lignes sans code agence valide
    if 'Code_Agence' in df.columns:
        df = df[df['Code_Agence'].str.len() > 0]
        df = df[~df['Code_Agence'].isin(['nan', 'None', ''])]

    return df.reset_index(drop=True)


def normalize_code_agence(val):
    """
    Harmonise les codes agence :
    "00001" → "1"  |  "1" → "1"  |  1 → "1"
    Supprime les zéros de tête pour matcher les 2 fichiers.
    """
    try:
        return str(int(str(val).strip().strip('"')))
    except (ValueError, TypeError):
        return str(val).strip().strip('"')


def normalize_mapping_chapitre(df):
    """
    Fichier 2 – Mapping Chapitre
    Colonnes attendues (positions B, C, D dans Excel → index 0,1,2 après lecture) :
        Chapitre | Solde (SD / SC / SD-SC) | Credits/Depots (Crédits / Dépôts)

    Stratégie :
    1. Si colonnes nommées correctement → mapping par mot-clé
    2. Si colonnes Unnamed → détection header dans les premières lignes, sinon position
    """
    df = df.copy()
    df = df.dropna(how='all').reset_index(drop=True)

    # ── Cas Unnamed ──────────────────────────────────────────────────────────
    all_unnamed = all(str(c).startswith('Unnamed') for c in df.columns)
    if all_unnamed:
        found = None
        for i in range(min(5, len(df))):
            row_vals = [str(v).lower().strip() for v in df.iloc[i].values]
            if any('chap' in v for v in row_vals):
                found = i
                break
        if found is not None:
            df.columns = [str(v).strip() for v in df.iloc[found].values]
            df = df.iloc[found + 1:].reset_index(drop=True)
        else:
            # nommer par position
            nb = len(df.columns)
            names = ['Chapitre', 'Solde', 'Credits_Depots'] + [f'Col_{i}' for i in range(3, nb)]
            df.columns = names[:nb]

    # ── Mapping par mot-clé ───────────────────────────────────────────────────
    df.columns = df.columns.astype(str).str.strip()
    rename_map = {}
    for col in df.columns:
        if col in ('Chapitre', 'Solde', 'Credits_Depots'):
            continue
        cl = col.lower().strip()
        if 'chapitre' in cl or 'chap' in cl:
            rename_map[col] = 'Chapitre'
        elif ('solde' in cl or cl in ['sd', 'sc', 'sd-sc', 'type solde', 'type_solde']
              or (cl.startswith('sd') or cl.startswith('sc'))):
            rename_map[col] = 'Solde'
        elif ('crédit' in cl or 'credit' in cl or 'dépôt' in cl or 'depot' in cl
              or 'credits' in cl or 'depots' in cl or 'produit' in cl or 'operation' in cl):
            rename_map[col] = 'Credits_Depots'
    df.rename(columns=rename_map, inplace=True)

    # ── Nettoyage valeurs ────────────────────────────────────────────────────
    for col in ['Chapitre', 'Solde', 'Credits_Depots']:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip().str.strip('"')

    if 'Chapitre' in df.columns:
        df = df[df['Chapitre'].notna()]
        df = df[~df['Chapitre'].isin(['', 'nan', 'None'])]
        df = df.reset_index(drop=True)

    return df


def normalize_mapping_localites(df):
    """
    Fichier 3 – Mapping Localités-Agences
    Colonnes : Code Agence | Code Localité | Localité
    """
    df = df.copy()
    df = df.dropna(how='all').reset_index(drop=True)

    all_unnamed = all(str(c).startswith('Unnamed') for c in df.columns)
    if all_unnamed:
        found = None
        for i in range(min(5, len(df))):
            row_vals = [str(v).lower().strip() for v in df.iloc[i].values]
            if any('agence' in v or 'localit' in v or 'code' in v for v in row_vals):
                found = i
                break
        if found is not None:
            df.columns = [str(v).strip() for v in df.iloc[found].values]
            df = df.iloc[found + 1:].reset_index(drop=True)
        else:
            nb = len(df.columns)
            names = ['Code_Agence', 'Code_Localite', 'Localite'] + [f'Col_{i}' for i in range(3, nb)]
            df.columns = names[:nb]

    df.columns = df.columns.astype(str).str.strip()
    rename_map = {}
    for col in df.columns:
        if col in ('Code_Agence', 'Code_Localite', 'Localite'):
            continue
        cl = col.lower().strip()
        if 'code' in cl and ('agence' in cl or cl == 'code agence'):
            rename_map[col] = 'Code_Agence'
        elif 'code' in cl and ('localit' in cl or cl == 'code localité'):
            rename_map[col] = 'Code_Localite'
        elif 'localit' in cl and 'code' not in cl:
            rename_map[col] = 'Localite'
    df.rename(columns=rename_map, inplace=True)

    # Normaliser Code_Agence (supprimer zéros de tête)
    if 'Code_Agence' in df.columns:
        df['Code_Agence'] = df['Code_Agence'].apply(normalize_code_agence)
        df = df[~df['Code_Agence'].isin(['', 'nan', 'None'])]
        df = df.reset_index(drop=True)

    if 'Localite' in df.columns:
        df['Localite'] = df['Localite'].astype(str).str.strip().str.upper()

    return df


def calculer_montant_saham(debit, credit, solde):
    """
    Calcule le montant Saham selon le type de solde :
      SD      → montant = abs(Débit)
      SC      → montant = abs(Crédit)
      SD-SC   → montant = abs(Débit) - abs(Crédit)  (net, peut être négatif)
      autres  → max(abs(Débit), abs(Crédit))
    """
    d = abs(float(debit) if not pd.isna(debit) else 0)
    c = abs(float(credit) if not pd.isna(credit) else 0)
    s = str(solde).upper().strip()
    if s == 'SD':
        return d
    elif s == 'SC':
        return c
    elif s in ('SD-SC', 'SD/SC', 'SD - SC'):
        return d - c   # net signé
    else:
        return max(d, c)


def process_saham_pdm(df_main, df_mapping_chapitre, df_mapping_localites, df_bam=None):
    """
    Calcule les PDM Saham par localité.

    Fichier principal (TXT) colonnes par position :
        0=Annee | 1=Mois | 2=Code_Agence | 3=Chapitre | 4=Debit | 5(last)=Credit

    Fichier mapping chapitre :
        Chapitre | Solde (SD/SC/SD-SC) | Credits_Depots (Crédits/Dépôts)

    Fichier mapping localités :
        Code_Agence | Code_Localite | Localite
    """

    # ── 1. Vérifier et utiliser les colonnes du fichier principal ─────────────
    cols = list(df_main.columns)
    if len(cols) < 5:
        raise ValueError(f"Fichier principal : au moins 5 colonnes requises, trouvé {len(cols)}")

    if 'Code_Agence' in cols and 'Chapitre' in cols and 'Debit' in cols and 'Credit' in cols:
        df = df_main[['Code_Agence', 'Chapitre', 'Debit', 'Credit']].copy()
        df['Debit']  = pd.to_numeric(df['Debit'],  errors='coerce').fillna(0)
        df['Credit'] = pd.to_numeric(df['Credit'], errors='coerce').fillna(0)
    else:
        # fallback position
        df = df_main.iloc[:, [2, 3, 4, -1]].copy()
        df.columns = ['Code_Agence', 'Chapitre', 'Debit', 'Credit']
        df['Debit']  = pd.to_numeric(df['Debit'],  errors='coerce').fillna(0)
        df['Credit'] = pd.to_numeric(df['Credit'], errors='coerce').fillna(0)

    # ── 2. Normaliser Code_Agence (supprimer zéros de tête) ──────────────────
    df['Code_Agence'] = df['Code_Agence'].apply(normalize_code_agence)
    df['Chapitre']    = df['Chapitre'].astype(str).str.strip().str.strip('"')
    df = df[~df['Code_Agence'].isin(['', 'nan', 'None'])]
    df = df[~df['Chapitre'].isin(['', 'nan', 'None'])]

    # ── 3. Normaliser mapping chapitre ────────────────────────────────────────
    df_mc = normalize_mapping_chapitre(df_mapping_chapitre.copy())
    missing_mc = [c for c in ['Chapitre', 'Solde', 'Credits_Depots'] if c not in df_mc.columns]
    if missing_mc:
        raise ValueError(
            f"Colonnes manquantes dans le mapping chapitre : {missing_mc}\n"
            f"Colonnes disponibles : {list(df_mc.columns)}\n"
            f"Vérifiez que votre fichier contient : Chapitre | Solde | Credits/Depots"
        )
    df_mc['Chapitre'] = df_mc['Chapitre'].astype(str).str.strip()

    # ── 4. Jointure fichier principal × mapping chapitre ─────────────────────
    df = df.merge(df_mc[['Chapitre', 'Solde', 'Credits_Depots']], on='Chapitre', how='left')

    # Remplir les chapitres non mappés
    df['Solde']          = df['Solde'].fillna('SD-SC')
    df['Credits_Depots'] = df['Credits_Depots'].fillna('Inconnu')

    # ── 5. Calculer le montant selon SD/SC ────────────────────────────────────
    df['Montant'] = df.apply(
        lambda r: calculer_montant_saham(r['Debit'], r['Credit'], r['Solde']),
        axis=1
    )
    df['Montant_Abs'] = df['Montant'].abs()  # pour sommation PDM

    # Classifier Credits / Dépôts
    def is_credit(val):
        v = str(val).lower()
        return 'cr' in v or 'crédit' in v or 'credit' in v
    def is_depot(val):
        v = str(val).lower()
        return 'dép' in v or 'dep' in v or 'dôt' in v

    df['Is_Credit'] = df['Credits_Depots'].apply(is_credit)
    df['Is_Depot']  = df['Credits_Depots'].apply(is_depot)

    # ── 6. Normaliser mapping localités ──────────────────────────────────────
    df_ml = normalize_mapping_localites(df_mapping_localites.copy())
    missing_ml = [c for c in ['Code_Agence', 'Localite'] if c not in df_ml.columns]
    if missing_ml:
        raise ValueError(
            f"Colonnes manquantes dans le mapping localités : {missing_ml}\n"
            f"Colonnes disponibles : {list(df_ml.columns)}"
        )

    # ── 7. Jointure avec mapping localités ───────────────────────────────────
    df = df.merge(df_ml[['Code_Agence', 'Localite']], on='Code_Agence', how='left')
    df_unknown = df[df['Localite'].isna()]
    nb_unknown = len(df_unknown['Code_Agence'].unique())
    df = df[df['Localite'].notna()].copy()

    # ── 8. Agrégation par Localité ────────────────────────────────────────────
    df_credits = (df[df['Is_Credit']]
                  .groupby('Localite')['Montant_Abs'].sum()
                  .reset_index()
                  .rename(columns={'Montant_Abs': 'Credits_Saham'}))

    df_depots = (df[df['Is_Depot']]
                 .groupby('Localite')['Montant_Abs'].sum()
                 .reset_index()
                 .rename(columns={'Montant_Abs': 'Depots_Saham'}))

    df_result = df_credits.merge(df_depots, on='Localite', how='outer').fillna(0)
    df_result = df_result.sort_values('Localite').reset_index(drop=True)

    # Sauvegarder stats de jointure
    df_result.attrs['nb_agences_non_mappes'] = nb_unknown

    # ── 9. Calcul PDM vs BAM ──────────────────────────────────────────────────
    if df_bam is not None:
        bam_norm = normalize_bam_columns(df_bam.copy())
        bam_norm = clean_numeric_columns(bam_norm)

        if 'Localite' in bam_norm.columns:
            agg_cols = {}
            if 'Montant_Credits' in bam_norm.columns:
                agg_cols['Montant_Credits'] = 'sum'
            if 'Montant_Depots' in bam_norm.columns:
                agg_cols['Montant_Depots'] = 'sum'

            if agg_cols:
                bam_agg = bam_norm.groupby('Localite').agg(agg_cols).reset_index()
                # Normaliser la casse pour merge
                bam_agg['_merge_key']    = bam_agg['Localite'].str.upper().str.strip()
                df_result['_merge_key']  = df_result['Localite'].str.upper().str.strip()

                df_result = df_result.merge(
                    bam_agg[['_merge_key'] + list(agg_cols.keys())],
                    on='_merge_key', how='left'
                ).drop('_merge_key', axis=1)

                df_result[list(agg_cols.keys())] = df_result[list(agg_cols.keys())].fillna(0)

                # ⚠️ Les montants BAM sont en milliers de DH, ceux de Saham sont en DH
                # → BAM × 1000 pour homogénéiser les unités avant calcul PDM
                if 'Montant_Credits' in df_result.columns:
                    df_result['PDM_Credits'] = np.where(
                        df_result['Montant_Credits'] > 0,
                        df_result['Credits_Saham'] / (df_result['Montant_Credits'] * 1000) * 100, 0
                    )
                if 'Montant_Depots' in df_result.columns:
                    df_result['PDM_Depots'] = np.where(
                        df_result['Montant_Depots'] > 0,
                        df_result['Depots_Saham'] / (df_result['Montant_Depots'] * 1000) * 100, 0
                    )

    return df_result




def viz_pdm_saham_localites(df_pdm):
    has_pdm_credits   = 'PDM_Credits'     in df_pdm.columns
    has_pdm_depots    = 'PDM_Depots'      in df_pdm.columns
    has_bam_credits   = 'Montant_Credits' in df_pdm.columns
    has_bam_depots    = 'Montant_Depots'  in df_pdm.columns
    has_credits_saham = 'Credits_Saham'   in df_pdm.columns
    has_depots_saham  = 'Depots_Saham'    in df_pdm.columns

    nb_localites  = df_pdm['Localite'].nunique() if 'Localite' in df_pdm.columns else 0
    total_credits = df_pdm['Credits_Saham'].sum() if has_credits_saham else 0
    total_depots  = df_pdm['Depots_Saham'].sum()  if has_depots_saham  else 0
    # ⚠️ BAM en milliers de DH → × 1000 pour comparer avec Saham (en DH)
    total_bam_c   = df_pdm['Montant_Credits'].sum() * 1000 if has_bam_credits else 0
    total_bam_d   = df_pdm['Montant_Depots'].sum()  * 1000 if has_bam_depots  else 0
    pdm_cred_g    = (total_credits / total_bam_c * 100 if total_bam_c > 0 else 0)
    pdm_dep_g     = (total_depots  / total_bam_d * 100 if total_bam_d  > 0 else 0)
    pdm_glob      = ((total_credits + total_depots) / (total_bam_c + total_bam_d) * 100
                     if (total_bam_c + total_bam_d) > 0 else 0)

    # Période
    periode_saham = getattr(st.session_state, 'saham_periode_saham', '')
    periode_bam   = getattr(st.session_state, 'saham_periode_bam',   '')

    # ═══════════════════════════════════════════════════════════════════
    # BANDEAU PÉRIODE
    # ═══════════════════════════════════════════════════════════════════
    if periode_saham or periode_bam:
        st.markdown(f"""
        <div style="background:linear-gradient(135deg,#0d3d26,#1a5c3a);border:1px solid #4dff9e;
                    border-radius:10px;padding:1rem 1.5rem;margin-bottom:1rem;display:flex;
                    gap:2rem;align-items:center;">
            <span style="color:#4dff9e;font-size:1.2em;">📅</span>
            <span style="color:#e8e8e8;">
                <strong style="color:#4dff9e;">Saham :</strong> {periode_saham or '—'}
                &nbsp;&nbsp;|&nbsp;&nbsp;
                <strong style="color:#4dff9e;">BAM :</strong> {periode_bam or '—'}
            </span>
        </div>
        """, unsafe_allow_html=True)

    # ═══════════════════════════════════════════════════════════════════
    # KPI CARDS — vue directeur  (clic = détail calcul)
    # ═══════════════════════════════════════════════════════════════════

    # Init toggle states
    for _k in ['kpi_loc','kpi_cred','kpi_dep','kpi_pdmc','kpi_pdmd','kpi_pdmg']:
        if _k not in st.session_state:
            st.session_state[_k] = False

    # CSS commun pour les boutons KPI
    st.markdown("""
    <style>
    div[data-testid="column"] button[kind="secondary"] {
        background: transparent !important;
        border: none !important;
        padding: 0 !important;
        width: 100% !important;
    }
    </style>
    """, unsafe_allow_html=True)

    def kpi_card_html(val, label, color, active=False):
        border = f"2px solid {color}" if active else "1px solid #2a3050"
        bg     = "#1e2640" if active else "#1a1e2e"
        hint   = f'<div style="color:{color};font-size:0.65em;margin-top:2px;opacity:0.7;">🔍 cliquer pour détail</div>' if not active else f'<div style="color:{color};font-size:0.65em;margin-top:2px;opacity:0.7;">✖ cliquer pour fermer</div>'
        return f"""
        <div style="background:{bg};border:{border};border-radius:10px;
                    padding:0.9rem 0.5rem;text-align:center;border-top:3px solid {color};
                    cursor:pointer;transition:all 0.2s;">
            <div style="font-size:1.5em;font-weight:700;color:{color};font-family:monospace;">{val}</div>
            <div style="color:#8892a4;font-size:0.78em;margin-top:3px;">{label}</div>
            {hint}
        </div>"""

    def detail_box(content, color):
        st.markdown(f"""
        <div style="background:#f8f9fa;border:2px solid {color};border-left:5px solid {color};
                    border-radius:10px;padding:1.2rem 1.5rem;margin:0.5rem 0 1rem 0;
                    font-size:0.88em;line-height:1.8;color:#1a1a1a !important;">
            <div style="color:#1a1a1a !important;">{content}</div>
        </div>
        """, unsafe_allow_html=True)

    k1, k2, k3, k4, k5, k6 = st.columns(6)

    with k1:
        st.markdown(kpi_card_html(f"{nb_localites}", "Localités", "#4dff9e", st.session_state.kpi_loc), unsafe_allow_html=True)
        if st.button("▲", key="btn_kpi_loc", use_container_width=True, help="Voir le détail"):
            st.session_state.kpi_loc = not st.session_state.kpi_loc
            st.rerun()

    with k2:
        st.markdown(kpi_card_html(f"{total_credits/1e9:.2f} Md", "Crédits Saham", "#ff6b35", st.session_state.kpi_cred), unsafe_allow_html=True)
        if st.button("▲", key="btn_kpi_cred", use_container_width=True, help="Voir le détail"):
            st.session_state.kpi_cred = not st.session_state.kpi_cred
            st.rerun()

    with k3:
        st.markdown(kpi_card_html(f"{total_depots/1e9:.2f} Md", "Dépôts Saham", "#1a9e6e", st.session_state.kpi_dep), unsafe_allow_html=True)
        if st.button("▲", key="btn_kpi_dep", use_container_width=True, help="Voir le détail"):
            st.session_state.kpi_dep = not st.session_state.kpi_dep
            st.rerun()

    with k4:
        st.markdown(kpi_card_html(f"{pdm_cred_g:.3f}%", "PDM Crédits", "#ff6b35", st.session_state.kpi_pdmc), unsafe_allow_html=True)
        if st.button("▲", key="btn_kpi_pdmc", use_container_width=True, help="Voir le détail"):
            st.session_state.kpi_pdmc = not st.session_state.kpi_pdmc
            st.rerun()

    with k5:
        st.markdown(kpi_card_html(f"{pdm_dep_g:.3f}%", "PDM Dépôts", "#1a9e6e", st.session_state.kpi_pdmd), unsafe_allow_html=True)
        if st.button("▲", key="btn_kpi_pdmd", use_container_width=True, help="Voir le détail"):
            st.session_state.kpi_pdmd = not st.session_state.kpi_pdmd
            st.rerun()

    with k6:
        st.markdown(kpi_card_html(f"{pdm_glob:.3f}%", "PDM Globale", "#ffd700", st.session_state.kpi_pdmg), unsafe_allow_html=True)
        if st.button("▲", key="btn_kpi_pdmg", use_container_width=True, help="Voir le détail"):
            st.session_state.kpi_pdmg = not st.session_state.kpi_pdmg
            st.rerun()

    # ── Panneaux de détail (sous les cards) ─────────────────────────────────
    bam_c_raw = df_pdm['Montant_Credits'].sum() if has_bam_credits else 0
    bam_d_raw = df_pdm['Montant_Depots'].sum()  if has_bam_depots  else 0

    if st.session_state.kpi_loc:
        detail_box(f"""
        <span style="color:#4dff9e;font-weight:bold;">📍 Nombre de Localités</span><br><br>
        <b>Définition :</b> Nombre de localités distinctes présentes dans le fichier Saham
        <em>après</em> jointure avec le mapping agences→localités et avec les données BAM.<br><br>
        <b>Calcul :</b><br>
        &nbsp;&nbsp;→ Fichier Saham filtré sur la période sélectionnée<br>
        &nbsp;&nbsp;→ Chaque agence est rattachée à sa localité via le mapping<br>
        &nbsp;&nbsp;→ Seules les localités également présentes dans BAM sont conservées<br><br>
        <b>Résultat :</b> <span style="color:#4dff9e;font-size:1.2em;"><b>{nb_localites} localités</b></span>
        """, "#4dff9e")

    if st.session_state.kpi_cred:
        nb_agences_cred = df_pdm[df_pdm['Credits_Saham'] > 0]['Localite'].nunique() if has_credits_saham else 0
        detail_box(f"""
        <span style="color:#ff6b35;font-weight:bold;">💳 Crédits Saham — Détail du calcul</span><br><br>
        <b>Formule :</b><br>
        &nbsp;&nbsp;Σ Montant_Abs de toutes les lignes classées <em>Crédit</em> dans le mapping Chapitre<br><br>
        <b>Étapes :</b><br>
        &nbsp;&nbsp;1. Pour chaque ligne du fichier TXT :<br>
        &nbsp;&nbsp;&nbsp;&nbsp;• Si Type_Solde = <b>SD</b> → Montant = abs(Débit)<br>
        &nbsp;&nbsp;&nbsp;&nbsp;• Si Type_Solde = <b>SC</b> → Montant = abs(Crédit)<br>
        &nbsp;&nbsp;&nbsp;&nbsp;• Si Type_Solde = <b>SD-SC</b> → Montant = Débit − Crédit<br>
        &nbsp;&nbsp;2. Filtrer les lignes où <b>Type_Produit = Crédit</b><br>
        &nbsp;&nbsp;3. Grouper par Localité et sommer les montants<br>
        &nbsp;&nbsp;4. Sommer toutes les localités → Total global<br><br>
        <b>Localités avec crédits :</b> {nb_agences_cred}<br>
        <b>Total brut (DH) :</b> {total_credits:,.0f} DH<br>
        <b>Résultat affiché :</b> <span style="color:#ff6b35;font-size:1.2em;"><b>{total_credits/1e9:.4f} Md DH</b></span>
        &nbsp;(÷ 1 000 000 000)
        """, "#ff6b35")

    if st.session_state.kpi_dep:
        nb_agences_dep = df_pdm[df_pdm['Depots_Saham'] > 0]['Localite'].nunique() if has_depots_saham else 0
        detail_box(f"""
        <span style="color:#1a9e6e;font-weight:bold;">🏦 Dépôts Saham — Détail du calcul</span><br><br>
        <b>Formule :</b><br>
        &nbsp;&nbsp;Σ Montant_Abs de toutes les lignes classées <em>Dépôt</em> dans le mapping Chapitre<br><br>
        <b>Étapes :</b><br>
        &nbsp;&nbsp;1. Pour chaque ligne du fichier TXT :<br>
        &nbsp;&nbsp;&nbsp;&nbsp;• Si Type_Solde = <b>SD</b> → Montant = abs(Débit)<br>
        &nbsp;&nbsp;&nbsp;&nbsp;• Si Type_Solde = <b>SC</b> → Montant = abs(Crédit)<br>
        &nbsp;&nbsp;&nbsp;&nbsp;• Si Type_Solde = <b>SD-SC</b> → Montant = Débit − Crédit<br>
        &nbsp;&nbsp;2. Filtrer les lignes où <b>Type_Produit = Dépôt</b><br>
        &nbsp;&nbsp;3. Grouper par Localité et sommer les montants<br>
        &nbsp;&nbsp;4. Sommer toutes les localités → Total global<br><br>
        <b>Localités avec dépôts :</b> {nb_agences_dep}<br>
        <b>Total brut (DH) :</b> {total_depots:,.0f} DH<br>
        <b>Résultat affiché :</b> <span style="color:#1a9e6e;font-size:1.2em;"><b>{total_depots/1e9:.4f} Md DH</b></span>
        &nbsp;(÷ 1 000 000 000)
        """, "#1a9e6e")

    if st.session_state.kpi_pdmc:
        detail_box(f"""
        <span style="color:#ff6b35;font-weight:bold;">📊 PDM Crédits — Détail du calcul</span><br><br>
        <b>Formule :</b><br>
        <div style="background:#1a1e2e;padding:0.6rem 1rem;border-radius:6px;margin:0.5rem 0;font-family:monospace;">
        PDM Crédits = Crédits_Saham (DH) ÷ (Montant_Credits_BAM × 1 000) × 100
        </div>
        <b>⚠️ Le × 1 000</b> est nécessaire car BAM publie en <b>milliers de DH</b>,
        Saham est en <b>DH bruts</b>.<br><br>
        <b>Valeurs utilisées :</b><br>
        &nbsp;&nbsp;• Crédits Saham total : <b>{total_credits:,.0f} DH</b><br>
        &nbsp;&nbsp;• Marché BAM Crédits : <b>{bam_c_raw:,.0f}</b> (milliers DH)
        = <b>{total_bam_c:,.0f} DH</b> après conversion<br><br>
        <b>Calcul :</b><br>
        &nbsp;&nbsp;{total_credits:,.0f} ÷ {total_bam_c:,.0f} × 100<br><br>
        <b>Résultat :</b> <span style="color:#ff6b35;font-size:1.2em;"><b>{pdm_cred_g:.4f}%</b></span>
        """, "#ff6b35")

    if st.session_state.kpi_pdmd:
        detail_box(f"""
        <span style="color:#1a9e6e;font-weight:bold;">📊 PDM Dépôts — Détail du calcul</span><br><br>
        <b>Formule :</b><br>
        <div style="background:#1a1e2e;padding:0.6rem 1rem;border-radius:6px;margin:0.5rem 0;font-family:monospace;">
        PDM Dépôts = Dépôts_Saham (DH) ÷ (Montant_Depots_BAM × 1 000) × 100
        </div>
        <b>⚠️ Le × 1 000</b> est nécessaire car BAM publie en <b>milliers de DH</b>,
        Saham est en <b>DH bruts</b>.<br><br>
        <b>Valeurs utilisées :</b><br>
        &nbsp;&nbsp;• Dépôts Saham total : <b>{total_depots:,.0f} DH</b><br>
        &nbsp;&nbsp;• Marché BAM Dépôts : <b>{bam_d_raw:,.0f}</b> (milliers DH)
        = <b>{total_bam_d:,.0f} DH</b> après conversion<br><br>
        <b>Calcul :</b><br>
        &nbsp;&nbsp;{total_depots:,.0f} ÷ {total_bam_d:,.0f} × 100<br><br>
        <b>Résultat :</b> <span style="color:#1a9e6e;font-size:1.2em;"><b>{pdm_dep_g:.4f}%</b></span>
        """, "#1a9e6e")

    if st.session_state.kpi_pdmg:
        detail_box(f"""
        <span style="color:#ffd700;font-weight:bold;">🌐 PDM Globale — Détail du calcul</span><br><br>
        <b>Formule :</b><br>
        <div style="background:#1a1e2e;padding:0.6rem 1rem;border-radius:6px;margin:0.5rem 0;font-family:monospace;">
        PDM Globale = (Crédits_Saham + Dépôts_Saham) ÷ (BAM_Crédits + BAM_Dépôts) × 100
        </div>
        <b>Raisonnement :</b> On additionne l'encours total Saham (crédits + dépôts)
        et on le rapporte au marché total BAM (crédits + dépôts), pour obtenir
        une part de marché <em>globale</em> toutes catégories confondues.<br><br>
        <b>Valeurs utilisées :</b><br>
        &nbsp;&nbsp;• Crédits Saham : <b>{total_credits:,.0f} DH</b><br>
        &nbsp;&nbsp;• Dépôts Saham : <b>{total_depots:,.0f} DH</b><br>
        &nbsp;&nbsp;• Total Saham : <b>{(total_credits+total_depots):,.0f} DH</b><br>
        &nbsp;&nbsp;• BAM Crédits converti : <b>{total_bam_c:,.0f} DH</b><br>
        &nbsp;&nbsp;• BAM Dépôts converti : <b>{total_bam_d:,.0f} DH</b><br>
        &nbsp;&nbsp;• Total marché BAM : <b>{(total_bam_c+total_bam_d):,.0f} DH</b><br><br>
        <b>Calcul :</b><br>
        &nbsp;&nbsp;{(total_credits+total_depots):,.0f} ÷ {(total_bam_c+total_bam_d):,.0f} × 100<br><br>
        <b>Résultat :</b> <span style="color:#ffd700;font-size:1.2em;"><b>{pdm_glob:.4f}%</b></span>
        """, "#ffd700")

    st.markdown("<br>", unsafe_allow_html=True)

    if not has_pdm_credits and not has_pdm_depots:
        st.warning("⚠️ Données BAM non disponibles — PDM non calculées. Importez les données BAM puis relancez.")
        return

    # ═══════════════════════════════════════════════════════════════════
    # ONGLETS DIRECTEUR
    # ═══════════════════════════════════════════════════════════════════
    tab_vue, tab_dep, tab_cred, tab_comp, tab_heatmap, tab_table = st.tabs([
        "🏆 Vue Directeur",
        "🏦 PDM Dépôts",
        "💳 PDM Crédits",
        "⚖️ Saham vs BAM",
        "🔥 Carte de Chaleur",
        "📋 Tableau Complet"
    ])

    # ─── TAB VUE DIRECTEUR ───────────────────────────────────────────
    with tab_vue:
        st.markdown("### 🏆 Vue d'ensemble — PDM par Localité")

        if has_pdm_credits and has_pdm_depots:
            # Classement combiné PDM (moyenne crédits + dépôts)
            df_dir = df_pdm.copy()
            df_dir['PDM_Moyenne'] = (df_dir['PDM_Credits'] + df_dir['PDM_Depots']) / 2
            df_dir = df_dir.sort_values('PDM_Moyenne', ascending=False).reset_index(drop=True)
            df_dir['Rang'] = df_dir.index + 1

            col_nb, col_seuil = st.columns(2)
            nb_dir = col_nb.slider("Nombre de localités", 5, min(50, len(df_dir)), min(20, len(df_dir)), key="nb_dir")
            seuil_dir = col_seuil.number_input("PDM moyenne min (%)", 0.0, 100.0, 0.0, 0.1, key="seuil_dir")

            df_top = df_dir[df_dir['PDM_Moyenne'] >= seuil_dir].head(nb_dir)

            # Graphe double barres PDM Crédits + Dépôts côte à côte
            fig_dir = go.Figure()
            fig_dir.add_trace(go.Bar(
                name='PDM Crédits (%)',
                x=df_top['Localite'], y=df_top['PDM_Credits'],
                marker_color='#ff6b35',
                text=df_top['PDM_Credits'].apply(lambda x: f"{x:.2f}%"),
                textposition='outside',
                hovertemplate='<b>%{x}</b><br>PDM Crédits : %{y:.3f}%<extra></extra>'
            ))
            fig_dir.add_trace(go.Bar(
                name='PDM Dépôts (%)',
                x=df_top['Localite'], y=df_top['PDM_Depots'],
                marker_color='#1a9e6e',
                text=df_top['PDM_Depots'].apply(lambda x: f"{x:.2f}%"),
                textposition='outside',
                hovertemplate='<b>%{x}</b><br>PDM Dépôts : %{y:.3f}%<extra></extra>'
            ))
            fig_dir.update_layout(
                barmode='group',
                title=f"Top {nb_dir} Localités — PDM Crédits & Dépôts",
                height=500,
                xaxis_tickangle=-40,
                xaxis_title='Localité', yaxis_title='PDM (%)',
                legend=dict(orientation='h', y=1.08, x=0.5, xanchor='center'),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                font=dict(color='#e8e8e8')
            )
            st.plotly_chart(fig_dir, use_container_width=True)

            # Top 5 & Bottom 5
            st.markdown("#### 🥇 Top 5 meilleures localités")
            col_t1, col_t2 = st.columns(2)
            with col_t1:
                st.markdown("**PDM Crédits**")
                top5_c = df_pdm.nlargest(5, 'PDM_Credits')[['Localite','PDM_Credits']].reset_index(drop=True)
                top5_c.index += 1
                top5_c.columns = ['Localité', 'PDM Crédits (%)']
                top5_c['PDM Crédits (%)'] = top5_c['PDM Crédits (%)'].apply(lambda x: f"{x:.3f}%")
                st.dataframe(top5_c, use_container_width=True)
            with col_t2:
                st.markdown("**PDM Dépôts**")
                top5_d = df_pdm.nlargest(5, 'PDM_Depots')[['Localite','PDM_Depots']].reset_index(drop=True)
                top5_d.index += 1
                top5_d.columns = ['Localité', 'PDM Dépôts (%)']
                top5_d['PDM Dépôts (%)'] = top5_d['PDM Dépôts (%)'].apply(lambda x: f"{x:.3f}%")
                st.dataframe(top5_d, use_container_width=True)

    # ─── TAB PDM DÉPÔTS ──────────────────────────────────────────────
    with tab_dep:
        if not has_pdm_depots:
            st.info("PDM Dépôts non calculée.")
        else:
            c1, c2 = st.columns(2)
            nb_d    = c1.slider("Nb localités", 5, min(60, len(df_pdm)), 20, key="nb_d_pdm")
            seuil_d = c2.number_input("PDM min (%)", 0.0, 100.0, 0.0, 0.01, key="seuil_d")
            dfd = df_pdm[df_pdm['PDM_Depots'] >= seuil_d].nlargest(nb_d, 'PDM_Depots')
            if dfd.empty:
                st.warning("Aucune localité avec ce seuil.")
            else:
                dfd_s = dfd.sort_values('PDM_Depots', ascending=True)
                fig_d = go.Figure(go.Bar(
                    y=dfd_s['Localite'], x=dfd_s['PDM_Depots'], orientation='h',
                    marker=dict(color=dfd_s['PDM_Depots'], colorscale='Greens', showscale=True),
                    text=dfd_s['PDM_Depots'].apply(lambda x: f"{x:.3f}%"), textposition='outside',
                    hovertemplate='<b>%{y}</b><br>PDM Dép : %{x:.3f}%<extra></extra>'
                ))
                fig_d.update_layout(
                    title=f"Top {nb_d} — PDM Dépôts par Localité",
                    height=max(500, nb_d * 32), xaxis_title='PDM (%)', yaxis_title='',
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    font=dict(color='#e8e8e8')
                )
                st.plotly_chart(fig_d, use_container_width=True)

    # ─── TAB PDM CRÉDITS ─────────────────────────────────────────────
    with tab_cred:
        if not has_pdm_credits:
            st.info("PDM Crédits non calculée.")
        else:
            c1, c2 = st.columns(2)
            nb_c    = c1.slider("Nb localités", 5, min(60, len(df_pdm)), 20, key="nb_c_pdm")
            seuil_c = c2.number_input("PDM min (%)", 0.0, 100.0, 0.0, 0.01, key="seuil_c")
            dfc = df_pdm[df_pdm['PDM_Credits'] >= seuil_c].nlargest(nb_c, 'PDM_Credits')
            if dfc.empty:
                st.warning("Aucune localité avec ce seuil.")
            else:
                dfc_s = dfc.sort_values('PDM_Credits', ascending=True)
                fig_c = go.Figure(go.Bar(
                    y=dfc_s['Localite'], x=dfc_s['PDM_Credits'], orientation='h',
                    marker=dict(color=dfc_s['PDM_Credits'], colorscale='Oranges', showscale=True),
                    text=dfc_s['PDM_Credits'].apply(lambda x: f"{x:.3f}%"), textposition='outside',
                    hovertemplate='<b>%{y}</b><br>PDM Cred : %{x:.3f}%<extra></extra>'
                ))
                fig_c.update_layout(
                    title=f"Top {nb_c} — PDM Crédits par Localité",
                    height=max(500, nb_c * 32), xaxis_title='PDM (%)', yaxis_title='',
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    font=dict(color='#e8e8e8')
                )
                st.plotly_chart(fig_c, use_container_width=True)

    # ─── TAB COMPARAISON SAHAM vs BAM ────────────────────────────────
    with tab_comp:
        if not has_bam_credits and not has_bam_depots:
            st.info("Données BAM non disponibles.")
        else:
            type_c = st.radio("Type", ["Crédits", "Dépôts", "Scatter PDM"],
                              horizontal=True, key="type_comp_pdm")
            nb_cp  = st.slider("Nb localités", 5, min(40, len(df_pdm)), 15, key="nb_cp_pdm")

            if type_c == "Scatter PDM" and has_pdm_credits and has_pdm_depots:
                # Quadrants
                avg_c = df_pdm['PDM_Credits'].mean()
                avg_d = df_pdm['PDM_Depots'].mean()
                fig_sc = go.Figure()
                fig_sc.add_shape(type='line', x0=avg_c, x1=avg_c,
                                 y0=df_pdm['PDM_Depots'].min(), y1=df_pdm['PDM_Depots'].max(),
                                 line=dict(color='rgba(255,255,255,0.3)', dash='dash'))
                fig_sc.add_shape(type='line', y0=avg_d, y1=avg_d,
                                 x0=df_pdm['PDM_Credits'].min(), x1=df_pdm['PDM_Credits'].max(),
                                 line=dict(color='rgba(255,255,255,0.3)', dash='dash'))
                fig_sc.add_trace(go.Scatter(
                    x=df_pdm['PDM_Credits'], y=df_pdm['PDM_Depots'],
                    mode='markers+text', text=df_pdm['Localite'],
                    textposition='top center',
                    marker=dict(size=10, color=df_pdm['PDM_Credits'],
                                colorscale='RdYlGn', showscale=True,
                                colorbar=dict(title='PDM Crédits %')),
                    hovertemplate='<b>%{text}</b><br>PDM Cred : %{x:.3f}%<br>PDM Dép : %{y:.3f}%<extra></extra>'
                ))
                fig_sc.update_layout(
                    title='Positionnement PDM Crédits vs Dépôts par Localité',
                    xaxis_title='PDM Crédits (%)', yaxis_title='PDM Dépôts (%)',
                    height=600, plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    font=dict(color='#e8e8e8')
                )
                st.plotly_chart(fig_sc, use_container_width=True)
                st.caption("Quadrant haut-droite = localités leaders sur les 2 dimensions")
            else:
                col_s = None
                if type_c == "Crédits" and has_bam_credits and has_credits_saham:
                    col_s, col_b, cs, label = 'Credits_Saham', 'Montant_Credits', '#ff6b35', 'Crédits'
                elif type_c == "Dépôts" and has_bam_depots and has_depots_saham:
                    col_s, col_b, cs, label = 'Depots_Saham', 'Montant_Depots', '#1a9e6e', 'Dépôts'
                else:
                    st.info("Données insuffisantes.")
                if col_s:
                    dfcp = df_pdm.nlargest(nb_cp, col_s).sort_values(col_s)
                    fig_cp = go.Figure()
                    fig_cp.add_trace(go.Bar(
                        # BAM en milliers de DH → /1e6 pour afficher en Md DH
                        name='Marché BAM', y=dfcp['Localite'], x=dfcp[col_b]/1e6,
                        orientation='h', marker_color='#b0c4de',
                        hovertemplate='<b>%{y}</b><br>BAM : %{x:.4f} Md<extra></extra>'
                    ))
                    fig_cp.add_trace(go.Bar(
                        # Saham en DH → /1e9 pour afficher en Md DH
                        name='Saham', y=dfcp['Localite'], x=dfcp[col_s]/1e9,
                        orientation='h', marker_color=cs,
                        hovertemplate='<b>%{y}</b><br>Saham : %{x:.4f} Md<extra></extra>'
                    ))
                    fig_cp.update_layout(
                        barmode='overlay', height=max(500, nb_cp * 32),
                        title=f'{label} Saham vs Marché BAM par Localité',
                        xaxis_title='Montant (Md DH)', yaxis_title='',
                        legend=dict(orientation='h', y=1.08),
                        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                        font=dict(color='#e8e8e8')
                    )
                    st.plotly_chart(fig_cp, use_container_width=True)

    # ─── TAB CARTE DE CHALEUR ─────────────────────────────────────────
    with tab_heatmap:
        st.markdown("### 🔥 Carte de Chaleur — PDM par Localité")
        if has_pdm_credits and has_pdm_depots and len(df_pdm) >= 3:
            nb_hm = st.slider("Nb localités à afficher", 10, min(80, len(df_pdm)), min(40, len(df_pdm)), key="nb_hm")
            df_hm = df_pdm.nlargest(nb_hm, 'PDM_Credits')[['Localite','PDM_Credits','PDM_Depots']].copy()
            df_hm = df_hm.sort_values('PDM_Credits', ascending=False)

            z_vals   = [df_hm['PDM_Credits'].tolist(), df_hm['PDM_Depots'].tolist()]
            y_labels = ['PDM Crédits (%)', 'PDM Dépôts (%)']
            x_labels = df_hm['Localite'].tolist()

            hover_texts = []
            for row_label in y_labels:
                row_hover = []
                col_key = 'PDM_Credits' if 'Crédits' in row_label else 'PDM_Depots'
                for _, row in df_hm.iterrows():
                    row_hover.append(f"{row['Localite']}<br>{row_label}: {row[col_key]:.3f}%")
                hover_texts.append(row_hover)

            fig_hm = go.Figure(go.Heatmap(
                z=z_vals, x=x_labels, y=y_labels,
                colorscale='RdYlGn',
                text=[[f"{v:.2f}%" for v in row] for row in z_vals],
                texttemplate="%{text}",
                hovertext=hover_texts,
                hoverinfo='text',
                colorbar=dict(title='PDM (%)')
            ))
            fig_hm.update_layout(
                title=f"Carte de Chaleur PDM — Top {nb_hm} Localités",
                height=300,
                xaxis_tickangle=-45,
                xaxis_title='Localité',
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                font=dict(color='#e8e8e8')
            )
            st.plotly_chart(fig_hm, use_container_width=True)
            st.caption("🟢 Vert = forte PDM | 🔴 Rouge = faible PDM")
        else:
            st.info("Il faut les PDM Crédits et Dépôts (au moins 3 localités) pour afficher la carte de chaleur.")

    # ─── TAB TABLEAU COMPLET ─────────────────────────────────────────
    with tab_table:
        st.subheader("Tableau Complet PDM par Localité")
        col_srch, col_sort = st.columns(2)
        srch     = col_srch.text_input("🔍 Rechercher une localité", key="srch_pdm")
        sort_opt = col_sort.selectbox("Trier par",
            [o for o in ["PDM Crédits","PDM Dépôts","Crédits Saham","Dépôts Saham"] if
             (o == "PDM Crédits"   and has_pdm_credits)  or
             (o == "PDM Dépôts"    and has_pdm_depots)   or
             (o == "Crédits Saham" and has_credits_saham) or
             (o == "Dépôts Saham"  and has_depots_saham)],
            key="sort_pdm")
        dft = df_pdm.copy()
        if srch:
            dft = dft[dft['Localite'].str.contains(srch, case=False, na=False)]
        smap = {"PDM Crédits":'PDM_Credits',"PDM Dépôts":'PDM_Depots',
                "Crédits Saham":'Credits_Saham',"Dépôts Saham":'Depots_Saham'}
        sc = smap.get(sort_opt, 'Credits_Saham')
        if sc in dft.columns:
            dft = dft.sort_values(sc, ascending=False)

        disp = pd.DataFrame({'Localité': dft['Localite'].values})
        if has_credits_saham:
            disp['Cred Saham (Md)']  = (dft['Credits_Saham']/1e9).apply(lambda x:f"{x:.4f}")
        if has_bam_credits:
            # BAM en milliers de DH → /1e6 pour afficher en Md DH
            disp['Cred BAM (Md)']    = (dft['Montant_Credits']/1e6).apply(lambda x:f"{x:.4f}")
        if has_pdm_credits:
            disp['PDM Crédits (%)']  = dft['PDM_Credits'].apply(lambda x:f"{x:.3f}%")
        if has_depots_saham:
            disp['Dep Saham (Md)']   = (dft['Depots_Saham']/1e9).apply(lambda x:f"{x:.4f}")
        if has_bam_depots:
            # BAM en milliers de DH → /1e6 pour afficher en Md DH
            disp['Dep BAM (Md)']     = (dft['Montant_Depots']/1e6).apply(lambda x:f"{x:.4f}")
        if has_pdm_depots:
            disp['PDM Dépôts (%)']   = dft['PDM_Depots'].apply(lambda x:f"{x:.3f}%")
        st.dataframe(disp, use_container_width=True, hide_index=True, height=480)

        st.divider()
        out_xl = BytesIO()
        with pd.ExcelWriter(out_xl, engine='openpyxl') as w:
            df_pdm.to_excel(w, index=False, sheet_name='PDM_Localites')
        out_xl.seek(0)
        st.download_button("\U0001f4e5 Télécharger PDM (Excel)", data=out_xl,
                           file_name=f"Saham_PDM_{datetime.now().strftime('%Y%m%d')}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



def create_saham_visualizations():
    """Page principale Visualisations Saham Bank – PDM par Localité"""

    st.header("📊 Visualisations Saham Bank – Parts de Marché par Localité")

    if st.session_state.saham_pdm_localites is not None:
        viz_pdm_saham_localites(st.session_state.saham_pdm_localites)
    else:
        st.markdown("""
        <div class="warning-box">
            <strong>⏳ Aucune PDM calculée pour l'instant.</strong><br>
            Importez vos données et lancez le calcul depuis
            <strong>Import &amp; Traitement → 🏦 Import Saham Bank</strong>.
        </div>
        """, unsafe_allow_html=True)

# ============================================================================
# MODULE BAM
# ============================================================================

# Module BAM
def import_bam_multi_annees():
    """
    Interface d'import BAM avec gestion multi-années (2016-2025)
    Permet d'importer des données mensuelles pour chaque année et de tout combiner
    """
    
    st.subheader("📊 Import BAM Multi-Années (2016-2025)")
    
    # =========================================================================
    # CHOIX DU MODE D'IMPORT
    # =========================================================================
    
    mode_import = st.radio(
        "Choisissez le mode d'import",
        ["📂 Import mensuel (fichiers séparés)", "⚡ Import fichier combiné (rapide)"],
        horizontal=True,
        help="Import mensuel : uploader mois par mois | Import combiné : uploader un seul fichier Excel déjà combiné"
    )
    
    st.divider()
    
    # =========================================================================
    # MODE 1 : IMPORT FICHIER COMBINÉ (RAPIDE)
    # =========================================================================
    
    if mode_import == "⚡ Import fichier combiné (rapide)":
        st.markdown("""
        <div class="info-box">
            <h4>⚡ Import Rapide</h4>
            <p>Uploadez directement un fichier Excel qui contient déjà toutes les données BAM combinées.</p>
            <p><strong>Format attendu :</strong></p>
            <ul>
                <li>Colonnes : <code>Annee</code>, <code>mois</code>, <code>Code</code>, <code>Localite</code>, <code>Nombre_Guichets</code>, <code>Montant_Depots</code>, <code>Montant_Credits</code></li>
                <li>Exemple : résultat d'un export précédent de cette application</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
        
        fichier_combine = None
        if st.session_state.combined_data_bam is None:
            fichier_combine = st.file_uploader(
                "📂 Sélectionnez votre fichier Excel combiné",
                type=['xlsx', 'xls'],
                key="fichier_bam_combine"
            )
        
        if fichier_combine:
            try:
                with st.spinner("⏳ Chargement du fichier..."):
                    df_combine = pd.read_excel(fichier_combine)
                    
                    # Normaliser et nettoyer
                    df_combine = normalize_bam_columns(df_combine)
                    df_combine = clean_numeric_columns(df_combine)
                    
                    # Vérifier les colonnes essentielles
                    required_cols = ['Annee', 'mois', 'Localite', 'Montant_Depots', 'Montant_Credits']
                    missing_cols = [col for col in required_cols if col not in df_combine.columns]
                    
                    if missing_cols:
                        st.error(f"❌ Colonnes manquantes : {missing_cols}")
                        st.info("Colonnes présentes : " + ", ".join(df_combine.columns.tolist()))
                        return
                    
                    # Ajouter Direction_Regionale si absente
                    if 'Direction_Regionale' not in df_combine.columns:
                        df_combine = add_direction_regionale(df_combine)
                    
                    # Stocker en session ET persister sur disque
                    st.session_state.combined_data_bam = df_combine
                    save_cache("combined_data_bam", df_combine)
                    st.rerun()
                    
            except Exception as e:
                st.error(f"❌ Erreur lors du chargement : {str(e)}")
                st.exception(e)

        # ── Affichage si données déjà chargées ───────────────────────────
        if st.session_state.combined_data_bam is not None:
            df_loaded = st.session_state.combined_data_bam
            st.success(f"✅ Base de données chargée — **{len(df_loaded):,} lignes**")
            col_info1, col_info2, col_info3 = st.columns(3)
            with col_info1:
                st.metric("Années", df_loaded['Annee'].nunique() if 'Annee' in df_loaded.columns else "—")
            with col_info2:
                nb_m = df_loaded.groupby('Annee')['mois'].nunique().sum() if 'Annee' in df_loaded.columns else "—"
                st.metric("Mois (total)", nb_m)
            with col_info3:
                st.metric("Localités", df_loaded['Localite'].nunique() if 'Localite' in df_loaded.columns else "—")
            st.markdown("<br>", unsafe_allow_html=True)
            col_btn1, col_btn2, col_btn3 = st.columns([1,2,1])
            with col_btn2:
                if st.button("🔄 Changer de base de données", use_container_width=True,
                             key="change_bam_db"):
                    st.session_state.combined_data_bam = None
                    clear_cache("combined_data_bam")
                    st.rerun()
            st.info("💡 Allez sur **'Visualisations BAM'** pour analyser les données.")
        
        return  # Sortir de la fonction ici pour le mode combiné
    
    # =========================================================================
    # MODE 2 : IMPORT MENSUEL (FICHIERS SÉPARÉS)
    # =========================================================================
    
    st.markdown("""
    <div class="info-box">
        <h4>📋 Instructions - Import Mensuel</h4>
        <ul>
            <li><strong>Étape 1 :</strong> Sélectionnez une ou plusieurs années (2016-2025)</li>
            <li><strong>Étape 2 :</strong> Pour chaque année, uploadez les fichiers mensuels disponibles</li>
            <li><strong>Étape 3 :</strong> Combinez les mois de chaque année</li>
            <li><strong>Étape 4 :</strong> Combinez toutes les années ensemble</li>
            <li><strong>Étape 5 :</strong> Téléchargez le fichier combiné final</li>
            <li>⚠️ Les mois manquants seront ignorés automatiquement</li>
        </ul>
    </div>
    """, unsafe_allow_html=True)
    
    st.divider()
    
    # =========================================================================
    # INITIALISATION SESSION STATE
    # =========================================================================
    
    if 'bam_years_data' not in st.session_state:
        st.session_state.bam_years_data = {}  # {année: {data, nb_mois, mois_list}}
    
    if 'bam_final_combined' not in st.session_state:
        st.session_state.bam_final_combined = None
    
    # =========================================================================
    # ÉTAPE 1 : SÉLECTION DES ANNÉES
    # =========================================================================
    
    st.markdown("### 1️⃣ Sélection des Années")
    
    annees_disponibles = list(range(2016, 2026))  # 2016 à 2025
    
    annees_selectionnees = st.multiselect(
        "Choisissez les années à importer",
        options=annees_disponibles,
        default=[2024, 2025] if not st.session_state.bam_years_data else list(st.session_state.bam_years_data.keys()),
        key="select_annees_bam",
        help="Vous pouvez sélectionner plusieurs années"
    )
    
    if not annees_selectionnees:
        st.warning("⚠️ Veuillez sélectionner au moins une année pour continuer")
        return
    
    st.info(f"✅ {len(annees_selectionnees)} année(s) sélectionnée(s): {', '.join(map(str, sorted(annees_selectionnees)))}")
    
    st.divider()
    
    # =========================================================================
    # ÉTAPE 2 : IMPORT MENSUEL PAR ANNÉE
    # =========================================================================
    
    st.markdown("### 2️⃣ Import des Données Mensuelles")
    st.write("Dépliez chaque année pour uploader les fichiers mensuels")
    
    for annee in sorted(annees_selectionnees):
        
        # Déterminer si l'expander doit être ouvert par défaut
        is_expanded = (annee == max(annees_selectionnees) and annee not in st.session_state.bam_years_data)
        
        with st.expander(f"📅 **Année {annee}**", expanded=is_expanded):
            
            # Afficher le statut si déjà combiné
            if annee in st.session_state.bam_years_data:
                year_info = st.session_state.bam_years_data[annee]
                
                st.success(f"✅ Année {annee} déjà combinée")
                
                col_info1, col_info2, col_info3 = st.columns(3)
                
                with col_info1:
                    st.metric("Mois importés", year_info['nb_mois'])
                
                with col_info2:
                    st.metric("Total lignes", f"{len(year_info['data']):,}")
                
                with col_info3:
                    mois_names = [['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Jui', 
                                  'Jul', 'Aoû', 'Sep', 'Oct', 'Nov', 'Déc'][m-1] 
                                 for m in sorted(year_info['mois_list'])]
                    st.write(f"**Mois:** {', '.join(mois_names)}")
                
                # Bouton pour recombiner
                if st.button(f"🔄 Recombiner {annee}", key=f"recombine_{annee}"):
                    del st.session_state.bam_years_data[annee]
                    st.rerun()
                
                st.divider()
            
            # Upload pour cette année
            st.write(f"**📂 Fichiers Excel BAM pour l'année {annee}**")
            
            uploaded_files_year = st.file_uploader(
                f"Sélectionnez un ou plusieurs fichiers mensuels pour {annee}",
                type=['xlsx', 'xls'],
                accept_multiple_files=True,
                key=f'upload_bam_{annee}',
                help="Sélectionnez plusieurs fichiers avec Ctrl+Clic"
            )
            
            if uploaded_files_year:
                
                st.write(f"**{len(uploaded_files_year)} fichier(s) uploadé(s)**")
                
                fichiers_pour_annee = []
                
                for idx, uploaded_file in enumerate(uploaded_files_year):
                    try:
                        # Lire le fichier
                        df = pd.read_excel(uploaded_file)
                        
                        # Détecter le mois automatiquement
                        detected_month = get_month_from_filename(uploaded_file.name)
                        
                        # Affichage en colonnes
                        col_file, col_month, col_lines = st.columns([3, 2, 1])
                        
                        with col_file:
                            st.write(f"📄 {uploaded_file.name}")
                        
                        with col_month:
                            mois_selectionne = st.selectbox(
                                "Mois",
                                options=list(range(1, 13)),
                                index=(detected_month - 1) if detected_month else 0,
                                format_func=lambda x: ['Janvier', 'Février', 'Mars', 'Avril', 
                                                       'Mai', 'Juin', 'Juillet', 'Août', 
                                                       'Septembre', 'Octobre', 'Novembre', 'Décembre'][x-1],
                                key=f"mois_{annee}_{idx}_{uploaded_file.name}"
                            )
                        
                        with col_lines:
                            st.metric("Lignes", len(df))
                        
                        # Stocker les infos
                        fichiers_pour_annee.append({
                            'filename': uploaded_file.name,
                            'data': df,
                            'month': mois_selectionne,
                            'year': annee
                        })
                    
                    except Exception as e:
                        st.error(f"❌ Erreur lors de la lecture de {uploaded_file.name}: {str(e)}")
                
                # Bouton pour combiner les mois de cette année
                if fichiers_pour_annee:
                    st.divider()
                    
                    col_btn_left, col_btn_center, col_btn_right = st.columns([1, 2, 1])
                    
                    with col_btn_center:
                        if st.button(
                            f"✨ Combiner les {len(fichiers_pour_annee)} mois de {annee}",
                            type="primary",
                            use_container_width=True,
                            key=f"btn_combine_{annee}"
                        ):
                            with st.spinner(f"⏳ Combinaison des mois de {annee} en cours..."):
                                try:
                                    # Combiner les fichiers
                                    df_combined = combine_bam_files(fichiers_pour_annee)
                                    
                                    # Normaliser les colonnes
                                    df_combined = normalize_bam_columns(df_combined)
                                    
                                    # Nettoyer les colonnes numériques
                                    df_combined = clean_numeric_columns(df_combined)
                                    
                                    # Ajouter Direction Régionale
                                    df_combined = add_direction_regionale(df_combined)
                                    
                                    # Ajouter colonne année
                                    df_combined['Annee'] = annee
                                    
                                    # Stocker dans session state
                                    st.session_state.bam_years_data[annee] = {
                                        'data': df_combined,
                                        'nb_mois': len(fichiers_pour_annee),
                                        'mois_list': [f['month'] for f in fichiers_pour_annee]
                                    }
                                    
                                    st.success(f"✅ {len(fichiers_pour_annee)} mois combinés pour l'année {annee} !")
                                    st.rerun()
                                
                                except Exception as e:
                                    st.error(f"❌ Erreur lors de la combinaison de {annee}: {str(e)}")
                                    st.exception(e)
    
    st.divider()
    
    # =========================================================================
    # ÉTAPE 3 : COMBINAISON FINALE DE TOUTES LES ANNÉES
    # =========================================================================
    
    st.markdown("### 3️⃣ Combinaison Finale de Toutes les Années")
    
    nb_annees_combinees = len(st.session_state.bam_years_data)
    
    if nb_annees_combinees == 0:
        st.info("ℹ️ Aucune année n'a encore été combinée. Combinez d'abord les mois de chaque année ci-dessus.")
    
    else:
        st.write(f"**{nb_annees_combinees} année(s) prête(s) à être combinée(s):**")
        
        # Tableau récapitulatif
        recap_data = []
        for annee in sorted(st.session_state.bam_years_data.keys()):
            info = st.session_state.bam_years_data[annee]
            mois_names = [['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Jui', 
                          'Jul', 'Aoû', 'Sep', 'Oct', 'Nov', 'Déc'][m-1] 
                         for m in sorted(info['mois_list'])]
            
            recap_data.append({
                'Année': annee,
                'Mois importés': info['nb_mois'],
                'Total lignes': f"{len(info['data']):,}",
                'Mois': ', '.join(mois_names)
            })
        
        df_recap = pd.DataFrame(recap_data)
        st.dataframe(df_recap, use_container_width=True, hide_index=True)
        
        st.divider()
        
        col_final_left, col_final_center, col_final_right = st.columns([1, 2, 1])
        
        with col_final_center:
            if st.button(
                f"🔗 Combiner les {nb_annees_combinees} Année(s)",
                type="primary",
                use_container_width=True,
                key="btn_combine_all_years"
            ):
                with st.spinner("⏳ Combinaison de toutes les années en cours..."):
                    try:
                        # Combiner toutes les années
                        all_dfs = []
                        
                        for annee in sorted(st.session_state.bam_years_data.keys()):
                            df_year = st.session_state.bam_years_data[annee]['data'].copy()
                            all_dfs.append(df_year)
                        
                        # Concaténer toutes les dataframes
                        df_final = pd.concat(all_dfs, ignore_index=True)
                        
                        # Stocker le résultat final
                        st.session_state.bam_final_combined = df_final
                        st.session_state.combined_data_bam = df_final  # Pour compatibilité
                        st.session_state.processing_done = True
                        
                        # Calculer les totaux globaux
                        st.session_state.total_depots_bam = df_final['Montant_Depots'].sum()
                        st.session_state.total_credits_bam = df_final['Montant_Credits'].sum()
                        
                        st.success(f"✅ Combinaison réussie ! {nb_annees_combinees} année(s) combinée(s)")
                        st.rerun()
                    
                    except Exception as e:
                        st.error(f"❌ Erreur lors de la combinaison finale: {str(e)}")
                        st.exception(e)
    
    # =========================================================================
    # ÉTAPE 4 : RÉSULTATS FINAUX ET TÉLÉCHARGEMENT
    # =========================================================================
    
    if st.session_state.bam_final_combined is not None:
        st.divider()
        
        st.markdown("### 4️⃣ Résultats Finaux")
        
        df_final = st.session_state.bam_final_combined
        
        # Métriques globales
        col_m1, col_m2 = st.columns(2)
        
        with col_m1:
            st.metric("📊 Total Lignes", f"{len(df_final):,}")
        
        with col_m2:
            nb_annees = df_final['Annee'].nunique() if 'Annee' in df_final.columns else 0
            st.metric("📅 Années", nb_annees)
        
        st.divider()
        
        # Aperçu des données
        st.write("**📋 Aperçu des données combinées (10 premières lignes)**")
        st.dataframe(df_final.head(10), use_container_width=True)
        
        st.divider()
        
        # Bouton de téléchargement
        st.markdown("### 5️⃣ Téléchargement")
        
        col_dl_left, col_dl_center, col_dl_right = st.columns([1, 2, 1])
        
        with col_dl_center:
            # Préparer le fichier Excel
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_final.to_excel(writer, index=False, sheet_name='BAM_Combined')
            
            excel_data = output.getvalue()
            
            # Nom du fichier avec timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"BAM_Combined_{timestamp}.xlsx"
            
            # Bouton de téléchargement
            st.download_button(
                label="📥 Télécharger le Fichier Combiné (Excel)",
                data=excel_data,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                help=f"Fichier: {filename}"
            )
        
        st.success("✅ Données prêtes pour l'analyse ! Vous pouvez maintenant accéder aux visualisations.")
        
        # Info sur l'utilisation
        st.info("💡 **Prochaines étapes:** Allez dans 'Visualisations BAM' ou importez les données Saham Bank pour calculer les PDM.")


def visualisations_bam_avancees():
    """
    Visualisations BAM - DERNIER MOIS UNIQUEMENT + GRAPHIQUES COMPLETS
    """
    
    st.header("📊 Visualisations BAM")
    
    if st.session_state.combined_data_bam is None:
        st.warning("⚠️ Veuillez d'abord importer des données BAM")
        return
    
    df_all = st.session_state.combined_data_bam.copy()
    
    if 'Annee' not in df_all.columns or 'mois' not in df_all.columns:
        st.error("❌ Colonnes 'Annee' et 'mois' nécessaires")
        return
    
    # =========================================================================
    # FONCTION POUR OBTENIR LE DERNIER MOIS DE CHAQUE ANNÉE
    # =========================================================================
    
    def get_dernier_mois_par_annee(df):
        """Retourne uniquement le dernier mois de chaque année"""
        result = []
        for annee in df['Annee'].unique():
            df_y = df[df['Annee'] == annee]
            mois_max = df_y['mois'].max()
            df_dernier = df_y[df_y['mois'] == mois_max].copy()
            result.append(df_dernier)
        return pd.concat(result, ignore_index=True)
    
    df_derniers_mois = get_dernier_mois_par_annee(df_all)
    
    st.info("⚠️ **MÉTHODOLOGIE** : Toutes les analyses utilisent UNIQUEMENT le dernier mois de chaque année (JAMAIS de somme)")
    
    st.divider()
    
    # =========================================================================
    # ONGLETS
    # =========================================================================
    
    tab1, tab2, tab3, tab4 = st.tabs([
        "📊 Répartition par Année",
        "🔍 Exploration Détaillée", 
        "📈 Évolution par Année",
        "🏙️ Top Localités"
    ])
    
    # =========================================================================
    # TAB 1 : RÉPARTITION PAR ANNÉE
    # =========================================================================
    
    with tab1:
        st.subheader("Répartition par Année")
        
        st.info("💡 Chaque année est évaluée par son DERNIER MOIS uniquement")
        
        # Tableau
        data_tableau = []
        
        for annee in sorted(df_derniers_mois['Annee'].unique()):
            df_y = df_derniers_mois[df_derniers_mois['Annee'] == annee]
            
            mois_num = df_y['mois'].iloc[0] if len(df_y) > 0 else 0
            mois_noms = ['Jan','Fév','Mar','Avr','Mai','Jun','Jul','Aoû','Sep','Oct','Nov','Déc']
            nom_mois = mois_noms[int(mois_num)-1] if 1 <= mois_num <= 12 else str(mois_num)
            
            nb_lignes = len(df_y)
            total_depots = df_y['Montant_Depots'].sum()
            total_credits = df_y['Montant_Credits'].sum()
            
            data_tableau.append({
                'Année': annee,
                'Dernier Mois': nom_mois,
                'Nb Lignes': nb_lignes,
                'Total Dépôts': f"{total_depots/1e6:.2f} Md",
                'Total Crédits': f"{total_credits/1e6:.2f} Md"
            })
        
        df_tableau = pd.DataFrame(data_tableau)
        st.dataframe(df_tableau, use_container_width=True, hide_index=True)
        
        st.divider()
        
        # NOUVEAU : Graphiques côte à côte
        col_g1, col_g2 = st.columns(2)
        
        df_graph = df_derniers_mois.groupby('Annee').agg({
            'Montant_Depots': 'sum',
            'Montant_Credits': 'sum'
        }).reset_index().sort_values('Annee')
        
        # Graphique Dépôts
        with col_g1:
            st.write("**💰 Évolution Dépôts (Dernier Mois)**")
            
            fig_d = go.Figure()
            fig_d.add_trace(go.Scatter(
                x=df_graph['Annee'],
                y=df_graph['Montant_Depots']/1e6,
                mode='lines+markers',
                fill='tozeroy',
                line=dict(width=3, color='#1f77b4'),
                marker=dict(size=10),
                name='Dépôts'
            ))
            
            fig_d.update_layout(
                height=400,
                yaxis_title="Dépôts (Md MAD)",
                xaxis_title="Année",
                showlegend=False
            )
            
            st.plotly_chart(fig_d, use_container_width=True)
        
        # Graphique Crédits
        with col_g2:
            st.write("**💳 Évolution Crédits (Dernier Mois)**")
            
            fig_c = go.Figure()
            fig_c.add_trace(go.Scatter(
                x=df_graph['Annee'],
                y=df_graph['Montant_Credits']/1e6,
                mode='lines+markers',
                fill='tozeroy',
                line=dict(width=3, color='#ff7f0e'),
                marker=dict(size=10),
                name='Crédits'
            ))
            
            fig_c.update_layout(
                height=400,
                yaxis_title="Crédits (Md MAD)",
                xaxis_title="Année",
                showlegend=False
            )
            
            st.plotly_chart(fig_c, use_container_width=True)
    
    # =========================================================================
    # TAB 2 : EXPLORATION DÉTAILLÉE
    # =========================================================================
    
    with tab2:
        st.subheader("Exploration Détaillée")
        
        col_sel1, col_sel2 = st.columns(2)
        
        # Sélection année
        with col_sel1:
            annees_dispo = sorted(df_all['Annee'].unique())
            annee_select = st.selectbox(
                "📅 Sélectionnez une année",
                options=annees_dispo,
                key="annee_explore"
            )
        
        df_annee = df_all[df_all['Annee'] == annee_select]
        
        # Sélection mois
        with col_sel2:
            mois_dispo = sorted(df_annee['mois'].unique())
            mois_noms = ['Jan','Fév','Mar','Avr','Mai','Jun','Jul','Aoû','Sep','Oct','Nov','Déc']
            
            mois_select = st.selectbox(
                "📆 Sélectionnez un mois",
                options=mois_dispo,
                format_func=lambda x: mois_noms[int(x)-1] if 1 <= x <= 12 else str(x),
                key="mois_explore"
            )
        
        df_mois = df_annee[df_annee['mois'] == mois_select]
        
        st.divider()
        
        # Métriques du mois sélectionné
        st.write(f"**📊 {mois_noms[int(mois_select)-1] if 1 <= mois_select <= 12 else mois_select} {annee_select}**")
        
        col_m1, col_m2, col_m3 = st.columns(3)
        
        total_depots_mois = df_mois['Montant_Depots'].sum()
        total_credits_mois = df_mois['Montant_Credits'].sum()
        nb_localites = df_mois['Localite'].nunique()
        
        with col_m1:
            st.metric("Dépôts", f"{total_depots_mois/1e6:.2f} Md")
        with col_m2:
            st.metric("Crédits", f"{total_credits_mois/1e6:.2f} Md")
        with col_m3:
            st.metric("Localités", nb_localites)
        
        st.divider()
        
        # NOUVEAU : Graphiques Dépôts et Crédits par Mois de l'année
        st.write(f"**📈 Évolution Mensuelle en {annee_select}**")
        
        # Grouper par mois
        df_mois_year = df_annee.groupby('mois').agg({
            'Montant_Depots': 'sum',
            'Montant_Credits': 'sum'
        }).reset_index().sort_values('mois')
        
        df_mois_year['Mois_Nom'] = df_mois_year['mois'].apply(
            lambda x: mois_noms[int(x)-1] if 1 <= x <= 12 else str(x)
        )
        
        col_graph1, col_graph2 = st.columns(2)
        
        # Graphique Dépôts par Mois
        with col_graph1:
            st.write("**💰 Dépôts par Mois**")
            
            fig_dm = go.Figure()
            fig_dm.add_trace(go.Scatter(
                x=df_mois_year['Mois_Nom'],
                y=df_mois_year['Montant_Depots']/1e6,  # milliers÷1e6=Md
                mode='lines+markers',
                line=dict(width=3, color='#1f77b4'),
                marker=dict(size=10),
                fill='tozeroy',
                fillcolor='rgba(31, 119, 180, 0.2)'
            ))
            
            fig_dm.update_layout(
                height=400,
                yaxis_title="Dépôts (Md MAD)",
                xaxis_title="Mois",
                showlegend=False
            )
            
            st.plotly_chart(fig_dm, use_container_width=True)
        
        # Graphique Crédits par Mois
        with col_graph2:
            st.write("**💳 Crédits par Mois**")
            
            fig_cm = go.Figure()
            fig_cm.add_trace(go.Scatter(
                x=df_mois_year['Mois_Nom'],
                y=df_mois_year['Montant_Credits']/1e6,
                mode='lines+markers',
                line=dict(width=3, color='#ff7f0e'),
                marker=dict(size=10),
                fill='tozeroy',
                fillcolor='rgba(255, 127, 14, 0.2)'
            ))
            
            fig_cm.update_layout(
                height=400,
                yaxis_title="Crédits (Md MAD)",
                xaxis_title="Mois",
                showlegend=False
            )
            
            st.plotly_chart(fig_cm, use_container_width=True)
        
        st.divider()
        
        # Filtre par région
        if 'Direction_Regionale' in df_mois.columns:
            st.write("**🗺️ Exploration par Région**")
            
            regions_dispo = sorted(df_mois['Direction_Regionale'].unique())
            
            region_select = st.selectbox(
                "Sélectionnez une région",
                options=['Toutes'] + regions_dispo,
                key="region_explore"
            )
            
            if region_select != 'Toutes':
                df_region = df_mois[df_mois['Direction_Regionale'] == region_select]
                
                st.write(f"**📍 Région : {region_select}**")
                
                col_r1, col_r2, col_r3 = st.columns(3)
                
                depots_region = df_region['Montant_Depots'].sum()
                credits_region = df_region['Montant_Credits'].sum()
                nb_villes_region = df_region['Localite'].nunique()
                
                with col_r1:
                    st.metric("Dépôts Région", f"{depots_region/1e6:.2f} Md")
                with col_r2:
                    st.metric("Crédits Région", f"{credits_region/1e6:.2f} Md")
                with col_r3:
                    st.metric("Localités", nb_villes_region)
                
                st.divider()
                
                st.write(f"**🏙️ Localités de {region_select}**")
                
                df_loc_region = df_region.groupby('Localite').agg({
                    'Montant_Depots': 'sum',
                    'Montant_Credits': 'sum'
                }).reset_index().sort_values('Montant_Depots', ascending=False)
                
                # NOUVEAU : Graphiques Dépôts ET Crédits par localité
                col_loc1, col_loc2 = st.columns(2)
                
                df_loc_region['Depots_Md']  = df_loc_region['Montant_Depots']  / 1e6
                df_loc_region['Credits_Md'] = df_loc_region['Montant_Credits'] / 1e6

                with col_loc1:
                    st.write("**Dépôts par Localité (Md MAD)**")
                    fig_ld = px.bar(
                        df_loc_region.sort_values('Depots_Md', ascending=True),
                        x='Depots_Md',
                        y='Localite',
                        orientation='h',
                        color='Depots_Md',
                        color_continuous_scale='Blues'
                    )
                    fig_ld.update_traces(hovertemplate='%{y}<br>%{x:.2f} Md<extra></extra>')
                    fig_ld.update_layout(height=max(400, len(df_loc_region)*30), showlegend=False, xaxis_title='Md MAD')
                    st.plotly_chart(fig_ld, use_container_width=True)
                
                with col_loc2:
                    st.write("**Crédits par Localité (Md MAD)**")
                    fig_lc = px.bar(
                        df_loc_region.sort_values('Credits_Md', ascending=True),
                        x='Credits_Md',
                        y='Localite',
                        orientation='h',
                        color='Credits_Md',
                        color_continuous_scale='Oranges'
                    )
                    fig_lc.update_traces(hovertemplate='%{y}<br>%{x:.2f} Md<extra></extra>')
                    fig_lc.update_layout(height=max(400, len(df_loc_region)*30), showlegend=False, xaxis_title='Md MAD')
                    st.plotly_chart(fig_lc, use_container_width=True)
                
                # Tableau
                df_loc_display = df_loc_region.copy()
                df_loc_display['Dépôts'] = df_loc_display['Montant_Depots'].apply(lambda x: f"{x/1e6:.2f} Md")
                df_loc_display['Crédits'] = df_loc_display['Montant_Credits'].apply(lambda x: f"{x/1e6:.2f} Md")
                df_loc_display = df_loc_display[['Localite', 'Dépôts', 'Crédits']]
                
                st.dataframe(df_loc_display, use_container_width=True, hide_index=True)
        else:
            st.warning("Colonne 'Direction_Regionale' non disponible")
    
    # =========================================================================
    # TAB 3 : ÉVOLUTION PAR ANNÉE
    # =========================================================================
    
    with tab3:
        st.subheader("Évolution par Année")
        
        st.info("💡 Chaque point représente le DERNIER MOIS de l'année")
        
        df_evol = df_derniers_mois.groupby('Annee').agg({
            'Montant_Depots': 'sum',
            'Montant_Credits': 'sum',
            'mois': 'first'
        }).reset_index().sort_values('Annee')
        
        df_evol['Var_Depots_%'] = df_evol['Montant_Depots'].pct_change() * 100
        df_evol['Var_Credits_%'] = df_evol['Montant_Credits'].pct_change() * 100
        
        # Graphiques évolution
        col_ev1, col_ev2 = st.columns(2)
        
        with col_ev1:
            st.write("**💰 Évolution Dépôts**")
            
            fig1 = go.Figure()
            fig1.add_trace(go.Scatter(
                x=df_evol['Annee'],
                y=df_evol['Montant_Depots']/1e6,
                mode='lines+markers+text',
                fill='tozeroy',
                line=dict(width=4, color='#1f77b4'),
                marker=dict(size=12),
                text=[
                    f"{v/1e6:.2f} Md" if pd.isna(r) else f"{v/1e6:.2f} Md\n({'▲' if r>=0 else '▼'}{abs(r):.1f}%)"
                    for v, r in zip(df_evol['Montant_Depots'], df_evol['Var_Depots_%'])
                ],
                textposition='top center',
                textfont=dict(size=10, color='#1f77b4'),
                hovertemplate='<b>%{x}</b><br>Dépôts: %{y:.2f} Md<extra></extra>'
            ))
            fig1.update_layout(height=430, yaxis_title="Md MAD", showlegend=False, margin=dict(t=40))
            st.plotly_chart(fig1, use_container_width=True)
        
        with col_ev2:
            st.write("**💳 Évolution Crédits**")
            
            fig2 = go.Figure()
            fig2.add_trace(go.Scatter(
                x=df_evol['Annee'],
                y=df_evol['Montant_Credits']/1e6,
                mode='lines+markers+text',
                fill='tozeroy',
                line=dict(width=4, color='#ff7f0e'),
                marker=dict(size=12),
                text=[
                    f"{v/1e6:.2f} Md" if pd.isna(r) else f"{v/1e6:.2f} Md\n({'▲' if r>=0 else '▼'}{abs(r):.1f}%)"
                    for v, r in zip(df_evol['Montant_Credits'], df_evol['Var_Credits_%'])
                ],
                textposition='top center',
                textfont=dict(size=10, color='#ff7f0e'),
                hovertemplate='<b>%{x}</b><br>Crédits: %{y:.2f} Md<extra></extra>'
            ))
            fig2.update_layout(height=430, yaxis_title="Md MAD", showlegend=False, margin=dict(t=40))
            st.plotly_chart(fig2, use_container_width=True)
        
        st.divider()
        
        # Taux de croissance
        st.write("**📈 Taux de Croissance Annuel**")
        
        df_var = df_evol[df_evol['Var_Depots_%'].notna()].copy()
        
        if len(df_var) > 0:
            fig3 = go.Figure()
            
            fig3.add_trace(go.Bar(
                x=df_var['Annee'],
                y=df_var['Var_Depots_%'],
                name='Dépôts',
                marker_color='#1f77b4',
                text=df_var['Var_Depots_%'].apply(lambda x: f"{x:+.1f}%"),
                textposition='outside', textfont=dict(size=12)
            ))
            
            fig3.add_trace(go.Bar(
                x=df_var['Annee'],
                y=df_var['Var_Credits_%'],
                name='Crédits',
                marker_color='#ff7f0e',
                text=df_var['Var_Credits_%'].apply(lambda x: f"{x:+.1f}%"),
                textposition='outside', textfont=dict(size=12)
            ))
            
            fig3.update_layout(barmode='group', height=430, yaxis_title="Variation (%)", margin=dict(t=50))
            st.plotly_chart(fig3, use_container_width=True)
        
        st.divider()
        
        # Comparaison 2 périodes
        st.write("**🔄 Comparaison entre 2 Périodes**")
        
        st.info("Sélectionnez 2 périodes (Année + Mois)")
        
        col_p1, col_p2 = st.columns(2)
        
        mois_noms = ['Jan','Fév','Mar','Avr','Mai','Jun','Jul','Aoû','Sep','Oct','Nov','Déc']
        
        with col_p1:
            st.write("**Période 1**")
            col_a1, col_m1 = st.columns(2)
            with col_a1:
                annee_p1 = st.selectbox("Année", options=sorted(df_all['Annee'].unique()), key="annee_p1")
            df_a1 = df_all[df_all['Annee'] == annee_p1]
            with col_m1:
                mois_p1 = st.selectbox(
                    "Mois",
                    options=sorted(df_a1['mois'].unique()),
                    format_func=lambda x: mois_noms[int(x)-1] if 1 <= x <= 12 else str(x),
                    key="mois_p1"
                )
        
        with col_p2:
            st.write("**Période 2**")
            col_a2, col_m2 = st.columns(2)
            with col_a2:
                annee_p2 = st.selectbox(
                    "Année",
                    options=sorted(df_all['Annee'].unique()),
                    index=min(len(sorted(df_all['Annee'].unique()))-1, 1),
                    key="annee_p2"
                )
            df_a2 = df_all[df_all['Annee'] == annee_p2]
            with col_m2:
                mois_p2 = st.selectbox(
                    "Mois",
                    options=sorted(df_a2['mois'].unique()),
                    format_func=lambda x: mois_noms[int(x)-1] if 1 <= x <= 12 else str(x),
                    key="mois_p2"
                )
        
        if st.button("Calculer le Taux de Croissance & TCAM", type="primary", use_container_width=True):
            
            df_per1 = df_all[(df_all['Annee'] == annee_p1) & (df_all['mois'] == mois_p1)]
            d1 = df_per1['Montant_Depots'].sum()
            c1 = df_per1['Montant_Credits'].sum()
            
            df_per2 = df_all[(df_all['Annee'] == annee_p2) & (df_all['mois'] == mois_p2)]
            d2 = df_per2['Montant_Depots'].sum()
            c2 = df_per2['Montant_Credits'].sum()
            
            taux_d = ((d2 - d1) / d1 * 100) if d1 > 0 else 0
            taux_c = ((c2 - c1) / c1 * 100) if c1 > 0 else 0

            # ── TCAM : n = (annee_p2 - annee_p1) = nombre d'années - 1 ───────
            n = annee_p2 - annee_p1   # n = nombre d'années - 1 (intervalles)
            if n > 0 and d1 > 0 and d2 > 0:
                tcam_d = ((d2 / d1) ** (1 / n) - 1) * 100
            elif n == 0:
                tcam_d = taux_d
            else:
                tcam_d = None

            if n > 0 and c1 > 0 and c2 > 0:
                tcam_c = ((c2 / c1) ** (1 / n) - 1) * 100
            elif n == 0:
                tcam_c = taux_c
            else:
                tcam_c = None

            nom_p1 = f"{mois_noms[int(mois_p1)-1] if 1 <= mois_p1 <= 12 else mois_p1} {annee_p1}"
            nom_p2 = f"{mois_noms[int(mois_p2)-1] if 1 <= mois_p2 <= 12 else mois_p2} {annee_p2}"

            st.divider()
            st.write(f"**Résultat : {nom_p1} → {nom_p2}**")

            col_r1, col_r2, col_r3, col_r4 = st.columns(4)
            with col_r1:
                st.metric("Dépôts P1", f"{d1/1e6:.2f} Md")
            with col_r2:
                st.metric("Dépôts P2", f"{d2/1e6:.2f} Md", f"{taux_d:+.2f}%")
            with col_r3:
                st.metric("Crédits P1", f"{c1/1e6:.2f} Md")
            with col_r4:
                st.metric("Crédits P2", f"{c2/1e6:.2f} Md", f"{taux_c:+.2f}%")

            # ── Bloc TCAM ────────────────────────────────────────────────────
            st.divider()
            if n <= 0:
                st.info("ℹ️ Les deux périodes sont dans la même année — le TCAM est égal au taux de variation simple.")
            else:
                d1_md  = f"{d1/1e6:.4f}"
                d2_md  = f"{d2/1e6:.4f}"
                c1_md  = f"{c1/1e6:.4f}"
                c2_md  = f"{c2/1e6:.4f}"
                exp_str = f"1/{n}"
                tcam_d_str = f"{tcam_d:+.2f}%" if tcam_d is not None else "N/A"
                tcam_c_str = f"{tcam_c:+.2f}%" if tcam_c is not None else "N/A"
                col_dep, col_cred = st.columns(2)
                with col_dep:
                    st.markdown(f"""
                    <div style="background:#f0f8ff; border:1.5px solid #1f77b4;
                                border-radius:12px; padding:1.2rem 1.5rem; margin-bottom:0.5rem;">
                        <p style="color:#1f77b4; font-weight:bold; font-size:1em; margin:0 0 0.6rem;">
                            💰 TCAM Dépôts
                        </p>
                        <p style="color:#333; font-size:0.87em; margin:0 0 0.4rem;">
                            <b>Formule :</b><br>
                            TCAM = ( V₂ / V₁ )^(1/n) − 1<br>
                            &emsp;&emsp;&emsp;= ( {d2_md} / {d1_md} )^({exp_str}) − 1<br>
                            &emsp;&emsp;&emsp;<i>avec n = {annee_p2} − {annee_p1} = <b>{n}</b>
                            &nbsp;(nombre d'années − 1)</i>
                        </p>
                        <p style="color:{'#2ecc71' if (tcam_d or 0)>=0 else '#e74c3c'};
                                  font-size:1.8em; font-weight:bold; margin:0.4rem 0 0;">
                            {tcam_d_str} / an
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
                with col_cred:
                    st.markdown(f"""
                    <div style="background:#fff8f0; border:1.5px solid #ff7f0e;
                                border-radius:12px; padding:1.2rem 1.5rem; margin-bottom:0.5rem;">
                        <p style="color:#ff7f0e; font-weight:bold; font-size:1em; margin:0 0 0.6rem;">
                            💳 TCAM Crédits
                        </p>
                        <p style="color:#333; font-size:0.87em; margin:0 0 0.4rem;">
                            <b>Formule :</b><br>
                            TCAM = ( V₂ / V₁ )^(1/n) − 1<br>
                            &emsp;&emsp;&emsp;= ( {c2_md} / {c1_md} )^({exp_str}) − 1<br>
                            &emsp;&emsp;&emsp;<i>avec n = {annee_p2} − {annee_p1} = <b>{n}</b>
                            &nbsp;(nombre d'années − 1)</i>
                        </p>
                        <p style="color:{'#2ecc71' if (tcam_c or 0)>=0 else '#e74c3c'};
                                  font-size:1.8em; font-weight:bold; margin:0.4rem 0 0;">
                            {tcam_c_str} / an
                        </p>
                    </div>
                    """, unsafe_allow_html=True)

            st.divider()

            # ── Graphique comparaison avec taux affichés ──────────────────────
            fig_comp = go.Figure()
            fig_comp.add_trace(go.Bar(
                x=['Dépôts', 'Crédits'],
                y=[d1/1e6, c1/1e6],
                name=nom_p1,
                marker_color='#1f77b4',
                text=[f"{d1/1e6:.2f} Md", f"{c1/1e6:.2f} Md"],
                textposition='outside'
            ))
            fig_comp.add_trace(go.Bar(
                x=['Dépôts', 'Crédits'],
                y=[d2/1e6, c2/1e6],
                name=nom_p2,
                marker_color='#ff7f0e',
                text=[f"{d2/1e6:.2f} Md ({taux_d:+.1f}%)",
                      f"{c2/1e6:.2f} Md ({taux_c:+.1f}%)"],
                textposition='outside'
            ))
            # Annotations TCAM si disponibles
            if n > 0 and tcam_d is not None and tcam_c is not None:
                for xval, tcam_val, col_ann in [
                    ('Dépôts',  tcam_d, '#1f77b4'),
                    ('Crédits', tcam_c, '#ff7f0e')
                ]:
                    df_xval = df_comp_annot = [d1 if xval=='Dépôts' else c1,
                                               d2 if xval=='Dépôts' else c2]
                    y_ann = max(df_xval)/1e6 * 1.28
                    fig_comp.add_annotation(
                        x=xval, y=y_ann,
                        text=f"TCAM : {tcam_val:+.1f}%/an",
                        showarrow=False,
                        font=dict(size=12, color=col_ann, family='Arial Black'),
                        bgcolor='rgba(255,255,255,0.85)',
                        bordercolor=col_ann, borderwidth=1.5, borderpad=4
                    )
            fig_comp.update_layout(
                barmode='group', height=450,
                yaxis_title="Montant (Md MAD)",
                title=f"Comparaison {nom_p1} vs {nom_p2}",
                margin=dict(t=70)
            )
            st.plotly_chart(fig_comp, use_container_width=True)
            
            # =========================================================
            # ANALYSE DÉTAILLÉE PAR LOCALITÉ
            # =========================================================
            
            st.divider()
            
            st.write("**📊 Analyse Détaillée : Contribution des Localités à la Croissance**")
            
            # Grouper par localité pour chaque période
            df_loc_p1 = df_per1.groupby('Localite').agg({
                'Montant_Depots': 'sum',
                'Montant_Credits': 'sum'
            }).reset_index()
            
            df_loc_p2 = df_per2.groupby('Localite').agg({
                'Montant_Depots': 'sum',
                'Montant_Credits': 'sum'
            }).reset_index()
            
            # Fusionner les deux périodes
            df_compare_loc = df_loc_p1.merge(
                df_loc_p2,
                on='Localite',
                how='outer',
                suffixes=('_P1', '_P2')
            ).fillna(0)
            
            # Calculer les variations
            df_compare_loc['Variation_Depots'] = df_compare_loc['Montant_Depots_P2'] - df_compare_loc['Montant_Depots_P1']
            df_compare_loc['Variation_Credits'] = df_compare_loc['Montant_Credits_P2'] - df_compare_loc['Montant_Credits_P1']
            
            # Calculer la contribution en % de chaque localité
            variation_totale_depots = d2 - d1
            variation_totale_credits = c2 - c1
            
            if variation_totale_depots != 0:
                df_compare_loc['Contribution_Depots_%'] = (df_compare_loc['Variation_Depots'] / variation_totale_depots * 100)
            else:
                df_compare_loc['Contribution_Depots_%'] = 0
            
            if variation_totale_credits != 0:
                df_compare_loc['Contribution_Credits_%'] = (df_compare_loc['Variation_Credits'] / variation_totale_credits * 100)
            else:
                df_compare_loc['Contribution_Credits_%'] = 0
            
            # Trier par contribution aux dépôts
            df_compare_loc = df_compare_loc.sort_values('Contribution_Depots_%', ascending=False)
            
            # Afficher les top contributeurs
            st.write(f"**🔝 Top 10 Localités Contribuant à la Croissance**")
            
            df_top_contrib = df_compare_loc.head(10).copy()
            
            # Préparer l'affichage
            df_display = pd.DataFrame({
                'Localité': df_top_contrib['Localite'],
                'Variation Dépôts (Md)': df_top_contrib['Variation_Depots'].apply(lambda x: f"{x/1e6:.2f}"),
                'Contribution Dépôts (%)': df_top_contrib['Contribution_Depots_%'].apply(lambda x: f"{x:.2f}%"),
                'Variation Crédits (Md)': df_top_contrib['Variation_Credits'].apply(lambda x: f"{x/1e6:.2f}"),
                'Contribution Crédits (%)': df_top_contrib['Contribution_Credits_%'].apply(lambda x: f"{x:.2f}%")
            })
            
            st.dataframe(df_display, use_container_width=True, hide_index=True)
            
            st.divider()
            
            # Graphiques de contribution
            col_contrib1, col_contrib2 = st.columns(2)
            
            with col_contrib1:
                st.write("**💰 Contribution aux Dépôts**")
                
                fig_contrib_d = px.bar(
                    df_top_contrib,
                    x='Contribution_Depots_%',
                    y='Localite',
                    orientation='h',
                    color='Contribution_Depots_%',
                    color_continuous_scale='RdYlGn',
                    labels={'Contribution_Depots_%': 'Contribution (%)', 'Localite': 'Localité'}
                )
                
                fig_contrib_d.update_layout(height=400, showlegend=False)
                st.plotly_chart(fig_contrib_d, use_container_width=True)
            
            with col_contrib2:
                st.write("**💳 Contribution aux Crédits**")
                
                fig_contrib_c = px.bar(
                    df_top_contrib,
                    x='Contribution_Credits_%',
                    y='Localite',
                    orientation='h',
                    color='Contribution_Credits_%',
                    color_continuous_scale='RdYlGn',
                    labels={'Contribution_Credits_%': 'Contribution (%)', 'Localite': 'Localité'}
                )
                
                fig_contrib_c.update_layout(height=400, showlegend=False)
                st.plotly_chart(fig_contrib_c, use_container_width=True)
            
            st.divider()
            
            # Résumé de l'analyse
            st.write("**📋 Résumé de l'Analyse**")
            
            top_localite_depots = df_top_contrib.iloc[0]['Localite'] if len(df_top_contrib) > 0 else "N/A"
            top_contrib_depots = df_top_contrib.iloc[0]['Contribution_Depots_%'] if len(df_top_contrib) > 0 else 0
            
            top_localite_credits = df_compare_loc.sort_values('Contribution_Credits_%', ascending=False).iloc[0]['Localite'] if len(df_compare_loc) > 0 else "N/A"
            top_contrib_credits = df_compare_loc.sort_values('Contribution_Credits_%', ascending=False).iloc[0]['Contribution_Credits_%'] if len(df_compare_loc) > 0 else 0
            
            st.info(f"""
            **Dépôts :**
            - Taux de croissance : **{taux_d:+.2f}%**
            - TCAM : **{f'{tcam_d:+.2f}%/an' if (tcam_d is not None and n > 0) else 'N/A (même année)'}**
            - Principale contributrice : **{top_localite_depots}** ({top_contrib_depots:.2f}% de la variation)
            
            **Crédits :**
            - Taux de croissance : **{taux_c:+.2f}%**
            - TCAM : **{f'{tcam_c:+.2f}%/an' if (tcam_c is not None and n > 0) else 'N/A (même année)'}**
            - Principale contributrice : **{top_localite_credits}** ({top_contrib_credits:.2f}% de la variation)
            """)
    
    
    # =========================================================================
    # TAB 4 : TOP LOCALITÉS
    # =========================================================================
    
    with tab4:
        st.subheader("Top Localités")
        
        st.info("💡 Basé sur les DERNIERS MOIS des années sélectionnées")
        
        annees_dispo_top = sorted(df_derniers_mois['Annee'].unique())
        annees_select_top = st.multiselect(
            "Sélectionnez les années",
            options=annees_dispo_top,
            default=annees_dispo_top,
            key="annees_top"
        )
        
        if not annees_select_top:
            st.warning("Sélectionnez au moins une année")
            return
        
        df_top_filtered = df_derniers_mois[df_derniers_mois['Annee'].isin(annees_select_top)]
        
        top_n = st.slider("Nombre de localités dans le Top", 5, 50, 10, key="topn_slider")
        
        st.divider()
        
        df_loc = df_top_filtered.groupby('Localite').agg({
            'Montant_Depots': 'sum',
            'Montant_Credits': 'sum'
        }).reset_index().sort_values('Montant_Depots', ascending=False)
        
        df_top = df_loc.head(top_n)
        df_reste = df_loc.iloc[top_n:]
        
        col_t1, col_t2 = st.columns(2)
        
        with col_t1:
            st.write(f"**Top {top_n}**")
            d_top = df_top['Montant_Depots'].sum()
            c_top = df_top['Montant_Credits'].sum()
            st.metric("Dépôts", f"{d_top/1e6:.2f} Md")
            st.metric("Crédits", f"{c_top/1e6:.2f} Md")
        
        with col_t2:
            st.write(f"**Reste ({len(df_reste)} localités)**")
            d_reste = df_reste['Montant_Depots'].sum()
            c_reste = df_reste['Montant_Credits'].sum()
            st.metric("Dépôts", f"{d_reste/1e6:.2f} Md")
            st.metric("Crédits", f"{c_reste/1e6:.2f} Md")
        
        st.divider()
        
        # NOUVEAU : Graphiques Dépôts ET Crédits Top N
        st.write(f"**📊 Top {top_n} Localités**")
        
        col_top1, col_top2 = st.columns(2)
        
        with col_top1:
            st.write("**💰 Par Dépôts**")
            fig_td = px.bar(
                df_top.sort_values('Montant_Depots', ascending=True),
                x='Montant_Depots',
                y='Localite',
                orientation='h',
                color='Montant_Depots',
                color_continuous_scale='Blues'
            )
            fig_td.update_layout(height=max(400, top_n*25), showlegend=False)
            st.plotly_chart(fig_td, use_container_width=True)
        
        with col_top2:
            st.write("**💳 Par Crédits**")
            fig_tc = px.bar(
                df_top.sort_values('Montant_Credits', ascending=True),
                x='Montant_Credits',
                y='Localite',
                orientation='h',
                color='Montant_Credits',
                color_continuous_scale='Oranges'
            )
            fig_tc.update_layout(height=max(400, top_n*25), showlegend=False)
            st.plotly_chart(fig_tc, use_container_width=True)
        
        st.divider()
        
        # Graphique Top vs Reste
        st.write("**📊 Top vs Reste**")
        
        fig_vs = go.Figure()
        
        fig_vs.add_trace(go.Bar(
            x=['Dépôts', 'Crédits'],
            y=[d_top/1e6, c_top/1e6],
            name=f'Top {top_n}',
            marker_color='#1f77b4'
        ))
        
        fig_vs.add_trace(go.Bar(
            x=['Dépôts', 'Crédits'],
            y=[d_reste/1e6, c_reste/1e6],
            name=f'Reste ({len(df_reste)})',
            marker_color='#ff7f0e'
        ))
        
        fig_vs.update_layout(barmode='group', height=400, yaxis_title="Montant (Md MAD)")
        st.plotly_chart(fig_vs, use_container_width=True)



# ============================================================================
# MODULE BALANCE - TRAITEMENT FICHIERS TXT CHARGES/PRODUITS
# ============================================================================

import re
from datetime import datetime

def parse_balance_txt(txt_content, type_document="Produits"):
    """
    Parser CORRIGÉ pour fichiers Balance SGMB
    
    LOGIQUE CORRECTE :
    Les comptes détaillés (000 XXXXX) appartiennent au TOP_ACCOUNT qui vient APRÈS eux.
    
    Exemple :
    | 000 0570102  |INT DEBIT...|0,00|4 211,40|...    ← Compte
    |   7023       |- INT.S/... |0,00|4 211,40|...    ← TOP_ACCOUNT pour le compte au-dessus
    | 000 0570550  |INTERETS ...|0,00|15 683,77|...   ← Nouveau compte
    |   7101       |INTERETS ...|0,00|131 787,35|...  ← TOP_ACCOUNT pour les comptes au-dessus
    """
    
    # Nettoyer le contenu
    lines = txt_content.replace('\r', '').split('\n')
    
    data = []
    current_agence_id = None
    current_agence_desc = None
    
    # NOUVEAU : Buffer pour stocker les comptes en attente d'un TOP_ACCOUNT
    pending_comptes = []
    
    for line in lines:
        # ─────────────────────────────────────────────────────────────────
        # 1. Détecter l'agence
        # ─────────────────────────────────────────────────────────────────
        if 'AGENCE' in line and ':' in line:
            match = re.search(r'AGENCE\s*:\s*(\d+)\s+(.+)', line)
            if match:
                current_agence_id = match.group(1).strip()
                current_agence_desc = match.group(2).strip()
                # Réinitialiser le buffer pour une nouvelle agence
                pending_comptes = []
                continue
        
        # ─────────────────────────────────────────────────────────────────
        # 2. Ignorer les lignes non-données
        # ─────────────────────────────────────────────────────────────────
        if '*---' in line or '+---' in line:
            continue
        if 'No DU COMPTE' in line or 'INTITULE DU COMPTE' in line or 'SOLDES' in line:
            continue
        if not line.strip() or len(line.strip()) < 10:
            continue
        if 'TOTAL' in line and 'GENERAL' in line:
            continue
        
        # ─────────────────────────────────────────────────────────────────
        # 3. Parser les lignes avec des pipes |...|
        # ─────────────────────────────────────────────────────────────────
        if '|' in line:
            parts = line.split('|')
            
            if len(parts) < 5:
                continue
            
            try:
                compte = parts[1].strip()
                intitule = parts[2].strip()
                debit = parts[3].strip()
                credit = parts[4].strip() if len(parts) > 4 else "0,00"
            except:
                continue
            
            # ─────────────────────────────────────────────────────────────
            # 4. Identifier le type de ligne
            # ─────────────────────────────────────────────────────────────
            
            # Ligne TOP_ACCOUNT : juste des chiffres (ex: "7023" ou "  7023  ")
            if re.match(r'^\s*\d{4}\s*$', compte):
                top_account = compte.strip()
                top_account_desc = intitule
                
                # NOUVEAU : Assigner ce TOP_ACCOUNT à tous les comptes en attente
                for pending_compte in pending_comptes:
                    pending_compte['TOP_ACCOUNT'] = top_account
                    pending_compte['TOP_ACCOUNT_DESC'] = top_account_desc
                    data.append(pending_compte)
                
                # Vider le buffer
                pending_comptes = []
                continue
            
            # Ligne COMPTE détaillé : commence par "000 " (ex: "000 0570102")
            if re.match(r'^\s*000\s+\d+', compte):
                compte_num = re.sub(r'^\s*000\s+', '', compte).strip()
                try:
                    compte_num = str(int(compte_num))
                except ValueError:
                    pass
                
                # Convertir les montants
                try:
                    debit_clean = debit.replace(' ', '').replace(',', '.')
                    credit_clean = credit.replace(' ', '').replace(',', '.')
                    
                    debit_val = float(debit_clean) if debit_clean and debit_clean != '0.00' else 0
                    credit_val = float(credit_clean) if credit_clean and credit_clean != '0.00' else 0
                    
                    # Montant = Crédit - Débit (SC - SD)
                    # → Produits : Crédit > Débit  ⟹ montant POSITIF
                    # → Charges  : Débit  > Crédit ⟹ montant NÉGATIF
                    # Ce signe naturel permet d'écrire :
                    #   PNB = Produits + Charges  ≡  Produits - |Charges|
                    montant = credit_val - debit_val
                    
                except Exception as e:
                    montant = 0
                
                # NOUVEAU : Ajouter au buffer (sans TOP_ACCOUNT pour l'instant)
                pending_comptes.append({
                    'ID_AGENCI': current_agence_id,
                    'DESC_AGENCE': current_agence_desc,
                    'ID_ACCOUNT': compte_num,
                    'DESC_ACCOUNT': intitule,
                    'Montants': montant,
                    'TOP_ACCOUNT': None,  # Sera rempli plus tard
                    'TOP_ACCOUNT_DESC': None,  # Sera rempli plus tard
                    'Type': type_document
                })
    
    # NOUVEAU : À la fin, s'il reste des comptes sans TOP_ACCOUNT, les ajouter quand même
    for pending_compte in pending_comptes:
        if pending_compte['TOP_ACCOUNT'] is None:
            pending_compte['TOP_ACCOUNT'] = ''
            pending_compte['TOP_ACCOUNT_DESC'] = ''
        data.append(pending_compte)
    
    # Créer DataFrame
    df = pd.DataFrame(data)
    
    # Nettoyer et convertir
    if len(df) > 0:
        df['Montants'] = pd.to_numeric(df['Montants'], errors='coerce').fillna(0)
        df['ID_AGENCI'] = df['ID_AGENCI'].astype(str)
        df['ID_ACCOUNT'] = df['ID_ACCOUNT'].astype(str)
        df['TOP_ACCOUNT'] = df['TOP_ACCOUNT'].fillna('').astype(str)
    
    print(f"✅ Parser v3 : {len(df)} lignes extraites")
    print(f"   Agences uniques : {df['ID_AGENCI'].nunique() if len(df) > 0 else 0}")
    print(f"   Comptes uniques : {df['ID_ACCOUNT'].nunique() if len(df) > 0 else 0}")
    
    return df


def convert_balance_to_csv(df):
    """Export CSV avec virgules comme séparateur décimal et TOP_ACCOUNT_DESC forcé en texte."""
    df_export = df.copy()
    # Forcer TOP_ACCOUNT_DESC et TOP_ACCOUNT en texte (préfixe apostrophe pour éviter l'interprétation)
    for col in ['TOP_ACCOUNT_DESC', 'DESC_ACCOUNT', 'TOP_ACCOUNT']:
        if col in df_export.columns:
            df_export[col] = df_export[col].astype(str)
    # Montants : remplacer le point décimal par une virgule
    if 'Montants' in df_export.columns:
        df_export['Montants'] = df_export['Montants'].apply(
            lambda x: str(x).replace('.', ',') if pd.notna(x) else ''
        )
    return df_export.to_csv(index=False, sep=";", encoding="utf-8-sig").encode("utf-8-sig")


def convert_balance_to_excel(df, filename="Balance_Export.xlsx"):
    """
    Convertit le DataFrame en fichier Excel téléchargeable.
    Les colonnes descriptives sont forcées en type texte pour éviter
    que les valeurs commençant par '-' soient mal interprétées.
    """
    import openpyxl
    from openpyxl.styles import numbers as xl_numbers

    df_export = df.copy()
    # Forcer les colonnes descriptives en string explicitement
    text_cols = ['TOP_ACCOUNT_DESC', 'DESC_ACCOUNT', 'TOP_ACCOUNT', 'DESC_AGENCE']
    for col in text_cols:
        if col in df_export.columns:
            df_export[col] = df_export[col].astype(str)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='Balance')
        ws = writer.sheets['Balance']

        # Identifier les colonnes textuelles et leur appliquer le format @
        header = [cell.value for cell in ws[1]]
        for col_idx, col_name in enumerate(header, start=1):
            if col_name in text_cols:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.number_format = '@'   # format Texte dans Excel
                        if cell.value is not None:
                            cell.value = str(cell.value)

    return output.getvalue()


def comparer_deux_balances(df_balance1, df_balance2, nom_balance1="Balance 1", nom_balance2="Balance 2"):
    """
    Compare deux balances et retourne les différences
    
    Args:
        df_balance1: DataFrame de la première balance
        df_balance2: DataFrame de la deuxième balance  
        nom_balance1: Nom de la première balance
        nom_balance2: Nom de la deuxième balance
    
    Returns:
        DataFrame avec les différences
    """
    # Vérifier que les colonnes nécessaires existent
    required_cols = ['ID_AGENCI', 'ID_ACCOUNT', 'Montants']
    
    for col in required_cols:
        if col not in df_balance1.columns:
            st.error(f"❌ La colonne '{col}' est manquante dans {nom_balance1}")
            return None
        if col not in df_balance2.columns:
            st.error(f"❌ La colonne '{col}' est manquante dans {nom_balance2}")
            return None
    
    # Créer des clés uniques pour la jointure
    df_b1 = df_balance1.copy()
    df_b2 = df_balance2.copy()
    
    df_b1['Clé'] = df_b1['ID_AGENCI'].astype(str) + '_' + df_b1['ID_ACCOUNT'].astype(str)
    df_b2['Clé'] = df_b2['ID_AGENCI'].astype(str) + '_' + df_b2['ID_ACCOUNT'].astype(str)
    
    # Agréger par ID_AGENCI et ID_ACCOUNT (+ récupérer la description)
    df_b1_agg = df_b1.groupby(['Clé', 'ID_AGENCI', 'ID_ACCOUNT']).agg(
        Montant_B1=('Montants', 'sum'),
        DESC_ACCOUNT_B1=('DESC_ACCOUNT', 'first')
    ).reset_index()
    
    df_b2_agg = df_b2.groupby(['Clé', 'ID_AGENCI', 'ID_ACCOUNT']).agg(
        Montant_B2=('Montants', 'sum'),
        DESC_ACCOUNT_B2=('DESC_ACCOUNT', 'first')
    ).reset_index()
    
    # Fusion complète (outer join) — on inclut ID_AGENCI + ID_ACCOUNT des DEUX côtés
    # afin que les comptes présents uniquement dans Balance 2 soient correctement identifiés
    df_compare = pd.merge(
        df_b1_agg[['Clé', 'ID_AGENCI', 'ID_ACCOUNT', 'DESC_ACCOUNT_B1', 'Montant_B1']],
        df_b2_agg[['Clé', 'ID_AGENCI', 'ID_ACCOUNT', 'DESC_ACCOUNT_B2', 'Montant_B2']],
        on='Clé',
        how='outer',
        suffixes=('_B1', '_B2')
    )
    
    # Remplir les valeurs manquantes par 0 pour les montants
    df_compare['Montant_B1'] = df_compare['Montant_B1'].fillna(0)
    df_compare['Montant_B2'] = df_compare['Montant_B2'].fillna(0)
    
    # ID_AGENCI : coalesce (même agence des deux côtés, on prend la première valeur non-nulle)
    df_compare['ID_AGENCI'] = df_compare['ID_AGENCI_B1'].combine_first(df_compare['ID_AGENCI_B2'])
    df_compare.drop(columns=['ID_AGENCI_B1', 'ID_AGENCI_B2'], inplace=True)
    
    # DESC_ACCOUNT : coalesce (on prend la description de B1, sinon celle de B2)
    df_compare['DESC_ACCOUNT'] = df_compare['DESC_ACCOUNT_B1'].combine_first(df_compare['DESC_ACCOUNT_B2'])
    df_compare.drop(columns=['DESC_ACCOUNT_B1', 'DESC_ACCOUNT_B2'], inplace=True)
    
    # ID_ACCOUNT : on garde les deux colonnes séparées pour voir exactement
    # dans quelle balance chaque compte est présent
    # - ID_ACCOUNT_B1 = NaN  ➜ compte absent de Balance 1 (seulement dans B2)
    # - ID_ACCOUNT_B2 = NaN  ➜ compte absent de Balance 2 (seulement dans B1)
    
    # Calculer les différences
    df_compare['Différence'] = df_compare['Montant_B1'] - df_compare['Montant_B2']
    df_compare['Différence_%'] = ((df_compare['Montant_B1'] - df_compare['Montant_B2']) / 
                                   df_compare['Montant_B2'].replace(0, np.nan) * 100)
    
    # Statut de la différence
    def get_statut(row):
        if row['Montant_B1'] == 0 and row['Montant_B2'] == 0:
            return '='
        elif row['Différence'] > 0:
            return '▲ Augmentation'
        elif row['Différence'] < 0:
            return '▼ Diminution'
        else:
            return '= Identique'
    
    df_compare['Statut'] = df_compare.apply(get_statut, axis=1)
    
    # Présence dans les balances
    df_compare['Présence'] = 'Les deux'
    df_compare.loc[df_compare['Montant_B1'] == 0, 'Présence'] = f'Seulement {nom_balance2}'
    df_compare.loc[df_compare['Montant_B2'] == 0, 'Présence'] = f'Seulement {nom_balance1}'
    
    # Réorganiser les colonnes — deux colonnes ID_ACCOUNT distinctes + description
    df_compare = df_compare[[
        'ID_AGENCI', 'ID_ACCOUNT_B1', 'ID_ACCOUNT_B2', 'DESC_ACCOUNT',
        'Montant_B1', 'Montant_B2', 
        'Différence', 'Différence_%',
        'Statut', 'Présence'
    ]]
    
    # Renommer pour clarté
    df_compare.rename(columns={
        'ID_ACCOUNT_B1': f'ID_ACCOUNT ({nom_balance1})',
        'ID_ACCOUNT_B2': f'ID_ACCOUNT ({nom_balance2})',
        'DESC_ACCOUNT': 'Description Compte',
        'Montant_B1': f'Montant {nom_balance1}',
        'Montant_B2': f'Montant {nom_balance2}'
    }, inplace=True)
    
    return df_compare


def visualisations_balance(df_produits, df_charges):
    """
    Crée les visualisations pour le module Balance.
    """
    
    st.header("📊 Analyse Balance - Produits & Charges")
    
    # Combiner les deux DataFrames
    df_all = pd.concat([df_produits, df_charges], ignore_index=True)
    
    if len(df_all) == 0:
        st.warning("⚠️ Aucune donnée à visualiser")
        return
    
    # =========================================================================
    # MÉTRIQUES GLOBALES
    # =========================================================================
    
    total_produits = df_produits['Montants'].sum()
    total_charges = df_charges['Montants'].sum()
    resultat_net = total_produits + total_charges
    marge_nette = (resultat_net / total_produits * 100) if total_produits > 0 else 0
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("💰 Total Produits", f"{total_produits/1e6:.2f} Mrd DH")
    
    with col2:
        st.metric("💸 Total Charges", f"{total_charges/1e6:.2f} Mrd DH")
    
    with col3:
        delta_color = "normal" if resultat_net >= 0 else "inverse"
        st.metric("📈 PNB (P+C)", f"{resultat_net/1e6:.2f} Mrd DH", delta_color=delta_color)
    
    with col4:
        st.metric("📊 Ratio PNB/Produits", f"{marge_nette:.2f}%")
    
    st.divider()
    
    # =========================================================================
    # ONGLETS
    # =========================================================================
    
    tab1, tab2, tab3, tab4 = st.tabs([
        "📊 Vue Globale",
        "🏦 Analyse par Agence",
        "🎯 Analyse par Compte",
        "📋 Données Détaillées"
    ])
    
    # ─────────────────────────────────────────────────────────────────────────
    # TAB 1 : VUE GLOBALE
    # ─────────────────────────────────────────────────────────────────────────
    
    with tab1:
        st.subheader("Vue d'Ensemble")
        
        # Graphique Produits vs Charges
        col_g1, col_g2 = st.columns(2)
        
        with col_g1:
            st.write("**💰 Produits vs Charges**")
            
            fig_pvc = go.Figure()
            
            fig_pvc.add_trace(go.Bar(
                x=['Produits', 'Charges'],
                y=[total_produits/1e6, total_charges/1e6],
                marker_color=['#2ecc71', '#e74c3c'],
                text=[f"{total_produits/1e6:.2f} Mrd", f"{total_charges/1e6:.2f} Mrd"],
                textposition='outside'
            ))
            
            fig_pvc.update_layout(
                height=400,
                yaxis_title="Montant (Mrd DH)",
                showlegend=False
            )
            
            st.plotly_chart(fig_pvc, use_container_width=True)
        
        with col_g2:
            st.write("**📈 PNB (Produits + Charges)**")
            
            fig_res = go.Figure()
            
            fig_res.add_trace(go.Indicator(
                mode="gauge+number+delta",
                value=resultat_net/1e6,
                title={'text': "PNB (Mrd DH)"},
                delta={'reference': 0},
                gauge={
                    'axis': {'range': [None, total_produits/1e6]},
                    'bar': {'color': "darkgreen" if resultat_net > 0 else "darkred"},
                    'steps': [
                        {'range': [0, total_produits/1e6 * 0.3], 'color': "lightgray"},
                        {'range': [total_produits/1e6 * 0.3, total_produits/1e6 * 0.7], 'color': "gray"}
                    ],
                    'threshold': {
                        'line': {'color': "red", 'width': 4},
                        'thickness': 0.75,
                        'value': 0
                    }
                }
            ))
            
            fig_res.update_layout(height=400)
            
            st.plotly_chart(fig_res, use_container_width=True)
        
        st.divider()
        
        # Top 10 Comptes Produits et Charges
        st.write("**🔝 Top 10 Comptes**")
        
        col_t1, col_t2 = st.columns(2)
        
        with col_t1:
            st.write("**💰 Top 10 Produits**")
            
            top_produits = df_produits.nlargest(10, 'Montants')[['DESC_ACCOUNT', 'Montants']].copy()
            top_produits['Montants_Mrd'] = top_produits['Montants'] / 1e6
            
            fig_tp = px.bar(
                top_produits.sort_values('Montants_Mrd', ascending=True),
                x='Montants_Mrd',
                y='DESC_ACCOUNT',
                orientation='h',
                color='Montants_Mrd',
                color_continuous_scale='Greens'
            )
            fig_tp.update_traces(hovertemplate='%{y}<br>%{x:.2f} Mrd<extra></extra>')
            fig_tp.update_layout(height=400, showlegend=False, xaxis_title='Montant (Mrd DH)', yaxis_title='')
            
            st.plotly_chart(fig_tp, use_container_width=True)
        
        with col_t2:
            st.write("**💸 Top 10 Charges**")
            
            top_charges = df_charges.nlargest(10, 'Montants')[['DESC_ACCOUNT', 'Montants']].copy()
            top_charges['Montants_Mrd'] = top_charges['Montants'] / 1e6
            
            fig_tc = px.bar(
                top_charges.sort_values('Montants_Mrd', ascending=True),
                x='Montants_Mrd',
                y='DESC_ACCOUNT',
                orientation='h',
                color='Montants_Mrd',
                color_continuous_scale='Reds'
            )
            fig_tc.update_traces(hovertemplate='%{y}<br>%{x:.2f} Mrd<extra></extra>')
            fig_tc.update_layout(height=400, showlegend=False, xaxis_title='Montant (Mrd DH)', yaxis_title='')
            
            st.plotly_chart(fig_tc, use_container_width=True)
    
    # =========================================================================
    # TAB 2 : ANALYSE PAR AGENCE - VERSION AMÉLIORÉE
    # =========================================================================

    with tab2:
        st.subheader("Analyse par Agence")
    
        # Agrégation par agence
        agence_produits = df_produits.groupby(['ID_AGENCI', 'DESC_AGENCE'])['Montants'].sum().reset_index()
        agence_produits.columns = ['ID_AGENCI', 'DESC_AGENCE', 'Produits']
    
        agence_charges = df_charges.groupby(['ID_AGENCI', 'DESC_AGENCE'])['Montants'].sum().reset_index()
        agence_charges.columns = ['ID_AGENCI', 'DESC_AGENCE', 'Charges']
    
        agence_summary = agence_produits.merge(agence_charges, on=['ID_AGENCI', 'DESC_AGENCE'], how='outer').fillna(0)
        agence_summary['Resultat_Net'] = agence_summary['Produits'] + agence_summary['Charges']
        agence_summary['Marge_%'] = (agence_summary['Resultat_Net'] / agence_summary['Produits'] * 100).replace([float('inf'), -float('inf')], 0).fillna(0)
        agence_summary = agence_summary.sort_values('Resultat_Net', ascending=False)
    
    # ─────────────────────────────────────────────────────────────────────
    # MÉTRIQUES GLOBALES
    # ─────────────────────────────────────────────────────────────────────
    
    st.info(f"📊 **{len(agence_summary)} agences** au total")
    
    col_m1, col_m2, col_m3, col_m4 = st.columns(4)
    
    with col_m1:
        st.metric("Total Produits", f"{agence_summary['Produits'].sum()/1e6:.2f} Mrd")
    with col_m2:
        st.metric("Total Charges", f"{agence_summary['Charges'].sum()/1e6:.2f} Mrd")
    with col_m3:
        st.metric("PNB Total", f"{agence_summary['Resultat_Net'].sum()/1e6:.2f} Mrd")
    with col_m4:
        marge_globale = (agence_summary['Resultat_Net'].sum() / agence_summary['Produits'].sum() * 100)
        st.metric("Ratio PNB/Produits", f"{marge_globale:.2f}%")
    
    st.divider()
    
    # ─────────────────────────────────────────────────────────────────────
    # OPTIONS DE VISUALISATION
    # ─────────────────────────────────────────────────────────────────────
    
    col_opt1, col_opt2 = st.columns(2)
    
    with col_opt1:
        critere_tri = st.selectbox(
            "Trier par",
            ["PNB (P+C)", "Produits", "Charges", "Ratio %"],
            key="tri_agence"
        )
    
    with col_opt2:
        nb_agences = st.slider(
            "Nombre d'agences à afficher",
            min_value=10,
            max_value=min(100, len(agence_summary)),
            value=20,
            step=5,
            key="nb_agences"
        )
    
    # Mapper le critère de tri
    critere_map = {
        "PNB (P+C)": "Resultat_Net",
        "Produits": "Produits",
        "Charges": "Charges",
        "Ratio %": "Marge_%"
    }
    
    # Trier et sélectionner top N
    agence_top = agence_summary.sort_values(critere_map[critere_tri], ascending=False).head(nb_agences)
    
    st.divider()
    
    # ─────────────────────────────────────────────────────────────────────
    # GRAPHIQUE 1 : BARRES GROUPÉES (TOP N)
    # ─────────────────────────────────────────────────────────────────────
    
    st.write(f"**📊 Top {nb_agences} Agences par {critere_tri}**")
    
    fig_agence = go.Figure()
    
    # Trier pour l'affichage (ascending pour avoir le plus grand en haut)
    agence_plot = agence_top.sort_values(critere_map[critere_tri], ascending=True)
    
    fig_agence.add_trace(go.Bar(
        name='Produits',
        y=agence_plot['DESC_AGENCE'],
        x=agence_plot['Produits']/1e6,
        orientation='h',
        marker_color='#2ecc71',
        hovertemplate='<b>%{y}</b><br>Produits: %{x:.2f} Mrd<extra></extra>'
    ))
    
    fig_agence.add_trace(go.Bar(
        name='Charges',
        y=agence_plot['DESC_AGENCE'],
        x=agence_plot['Charges']/1e6,
        orientation='h',
        marker_color='#e74c3c',
        hovertemplate='<b>%{y}</b><br>Charges: %{x:.2f} Mrd<extra></extra>'
    ))
    
    fig_agence.add_trace(go.Scatter(
        name='PNB (P+C)',
        y=agence_plot['DESC_AGENCE'],
        x=agence_plot['Resultat_Net']/1e6,
        mode='markers',
        marker=dict(size=10, color='#3498db', symbol='diamond'),
        hovertemplate='<b>%{y}</b><br>PNB: %{x:.2f} Mrd<extra></extra>'
    ))
    
    fig_agence.update_layout(
        barmode='group',
        height=max(500, nb_agences * 25),
        xaxis_title='Montant (Mrd DH)',
        yaxis_title='',
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    
    st.plotly_chart(fig_agence, use_container_width=True)
    
    st.divider()
    
    # ─────────────────────────────────────────────────────────────────────
    # GRAPHIQUE 2 : TREEMAP (VISION GLOBALE)
    # ─────────────────────────────────────────────────────────────────────
    
    st.write("**🗺️ Vision Globale - Treemap des Produits (Top 30)**")
    
    agence_treemap = agence_summary.nlargest(30, 'Produits')
    
    fig_tree = px.treemap(
        agence_treemap,
        path=['DESC_AGENCE'],
        values='Produits',
        color='Marge_%',
        color_continuous_scale='RdYlGn',
        color_continuous_midpoint=0,
        hover_data={'Produits': ':,.0f', 'Charges': ':,.0f', 'Resultat_Net': ':,.0f', 'Marge_%': ':.2f'}
    )
    
    fig_tree.update_traces(
        textposition='middle center',
        textfont_size=12,
        hovertemplate='<b>%{label}</b><br>Produits: %{value:,.0f} DH<br>Ratio PNB: %{color:.2f}%<extra></extra>'
    )
    
    fig_tree.update_layout(height=600)
    
    st.plotly_chart(fig_tree, use_container_width=True)
    
    st.divider()
    
    # ─────────────────────────────────────────────────────────────────────
    # GRAPHIQUE 3 : MARGE PAR AGENCE
    # ─────────────────────────────────────────────────────────────────────
    
    st.write(f"**📈 Ratio PNB/Produits par Agence (Top {nb_agences})**")
    
    # Filtrer les agences avec produits > 0 pour éviter les marges infinies
    agence_marge = agence_top[agence_top['Produits'] > 0].sort_values('Marge_%', ascending=True)
    
    fig_marge = px.bar(
        agence_marge,
        y='DESC_AGENCE',
        x='Marge_%',
        orientation='h',
        color='Marge_%',
        color_continuous_scale='RdYlGn',
        color_continuous_midpoint=0,
        text='Marge_%'
    )
    
    fig_marge.update_traces(
        texttemplate='%{text:.1f}%',
        textposition='outside',
        hovertemplate='<b>%{y}</b><br>Ratio PNB: %{x:.2f}%<extra></extra>'
    )
    
    fig_marge.update_layout(
        height=max(500, len(agence_marge) * 25),
        xaxis_title='Ratio PNB/Produits (%)',
        yaxis_title='',
        showlegend=False
    )
    
    st.plotly_chart(fig_marge, use_container_width=True)
    
    st.divider()
    
    # ─────────────────────────────────────────────────────────────────────
    # TABLEAU RÉCAPITULATIF COMPLET
    # ─────────────────────────────────────────────────────────────────────
    
    st.write("**📋 Tableau Récapitulatif - Toutes les Agences**")
    
    # Filtres pour le tableau
    col_f1, col_f2 = st.columns(2)
    
    with col_f1:
        filtre_nom = st.text_input(
            "🔍 Rechercher une agence",
            placeholder="Tapez le nom de l'agence...",
            key="filtre_agence_nom"
        )
    
    with col_f2:
        filtre_min_produits = st.number_input(
            "Produits minimum (Millions DH)",
            min_value=0.0,
            value=0.0,
            step=1.0,
            key="filtre_min_prod"
        )
    
    # Appliquer les filtres
    agence_filtered = agence_summary.copy()
    
    if filtre_nom:
        agence_filtered = agence_filtered[
            agence_filtered['DESC_AGENCE'].str.contains(filtre_nom, case=False, na=False)
        ]
    
    if filtre_min_produits > 0:
        agence_filtered = agence_filtered[agence_filtered['Produits'] >= filtre_min_produits * 1e6]
    
    # Affichage
    st.write(f"**{len(agence_filtered)} agences** affichées (sur {len(agence_summary)} total)")
    
    agence_display = agence_filtered.copy()
    agence_display['Produits'] = agence_display['Produits'].apply(lambda x: f"{x/1e6:.2f} Mrd")
    agence_display['Charges'] = agence_display['Charges'].apply(lambda x: f"{x/1e6:.2f} Mrd")
    agence_display['Resultat_Net'] = agence_display['Resultat_Net'].apply(lambda x: f"{x/1e6:.2f} Mrd")
    agence_display['Marge_%'] = agence_display['Marge_%'].apply(lambda x: f"{x:.2f}%")
    agence_display = agence_display[['DESC_AGENCE', 'Produits', 'Charges', 'Resultat_Net', 'Marge_%']]
    agence_display.columns = ['Agence', 'Produits', 'Charges', 'PNB (P+C)', 'Ratio PNB (%)']
    
    st.dataframe(agence_display, use_container_width=True, hide_index=True, height=400)
    
    # Export
    st.divider()
    
    excel_agence = convert_balance_to_excel(agence_filtered)
    csv_agence   = convert_balance_to_csv(agence_filtered)
    _ca1, _ca2 = st.columns(2)
    with _ca1:
        st.download_button(label="📥 Agences (Excel)", data=excel_agence,
            file_name=f"Balance_Agences_{__import__('datetime').datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_age_xl", use_container_width=True)
    with _ca2:
        st.download_button(label="📥 Agences (CSV)", data=csv_agence,
            file_name=f"Balance_Agences_{__import__('datetime').datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv", key="dl_age_csv", use_container_width=True)

    # ─────────────────────────────────────────────────────────────────────────
    # TAB 3 : ANALYSE PAR COMPTE (TOP_ACCOUNT)
    # ─────────────────────────────────────────────────────────────────────────
    
    with tab3:
        st.subheader("Analyse par Catégorie de Compte (TOP_ACCOUNT)")
        
        # Filtre Type
        type_analyse = st.radio(
            "Type d'analyse",
            ["Produits", "Charges", "Les deux"],
            horizontal=True
        )
        
        if type_analyse == "Produits":
            df_analyse = df_produits
        elif type_analyse == "Charges":
            df_analyse = df_charges
        else:
            df_analyse = df_all
        
        # Agrégation par TOP_ACCOUNT
        top_summary = df_analyse.groupby(['TOP_ACCOUNT', 'TOP_ACCOUNT_DESC', 'Type'])['Montants'].sum().reset_index()
        top_summary = top_summary.sort_values('Montants', ascending=False)
        
        # Graphique
        st.write(f"**📊 Répartition par Catégorie de Compte**")
        
        fig_top = px.bar(
            top_summary.head(15).sort_values('Montants', ascending=True),
            x='Montants',
            y='TOP_ACCOUNT_DESC',
            orientation='h',
            color='Type',
            color_discrete_map={'Produits': '#2ecc71', 'Charges': '#e74c3c'}
        )
        fig_top.update_traces(hovertemplate='%{y}<br>%{x:,.0f} DH<extra></extra>')
        fig_top.update_layout(height=500, xaxis_title='Montant (DH)', yaxis_title='')
        
        st.plotly_chart(fig_top, use_container_width=True)
        
        st.divider()
        
        # Tableau
        st.write("**📋 Détail par TOP_ACCOUNT**")
        
        top_display = top_summary.copy()
        top_display['Montants'] = top_display['Montants'].apply(lambda x: f"{x/1e6:.2f} Mrd")
        top_display = top_display[['TOP_ACCOUNT', 'TOP_ACCOUNT_DESC', 'Type', 'Montants']]
        top_display.columns = ['Code', 'Description', 'Type', 'Montant']
        
        st.dataframe(top_display, use_container_width=True, hide_index=True)
    
    # ─────────────────────────────────────────────────────────────────────────
    # TAB 4 : DONNÉES DÉTAILLÉES
    # ─────────────────────────────────────────────────────────────────────────
    
    with tab4:
        st.subheader("Données Détaillées")
        
        # Filtres
        col_f1, col_f2, col_f3 = st.columns(3)
        
        with col_f1:
            type_filtre = st.selectbox(
                "Type",
                ["Tous", "Produits", "Charges"]
            )
        
        with col_f2:
            agences_list = ['Toutes'] + sorted(df_all['DESC_AGENCE'].unique().tolist())
            agence_filtre = st.selectbox(
                "Agence",
                agences_list
            )
        
        with col_f3:
            top_list = ['Tous'] + sorted(df_all['TOP_ACCOUNT'].dropna().unique().tolist())
            top_filtre = st.selectbox(
                "TOP_ACCOUNT",
                top_list
            )
        
        # Appliquer les filtres
        df_filtered = df_all.copy()
        
        if type_filtre != "Tous":
            df_filtered = df_filtered[df_filtered['Type'] == type_filtre]
        
        if agence_filtre != "Toutes":
            df_filtered = df_filtered[df_filtered['DESC_AGENCE'] == agence_filtre]
        
        if top_filtre != "Tous":
            df_filtered = df_filtered[df_filtered['TOP_ACCOUNT'] == top_filtre]
        
        # Afficher
        st.write(f"**{len(df_filtered):,} lignes**".replace(',', ' '))
        
        df_display = df_filtered.copy()
        df_display['Montants'] = df_display['Montants'].apply(lambda x: f"{x:,.2f} DH".replace(',', ' '))
        df_display = df_display[['ID_AGENCI', 'DESC_AGENCE', 'ID_ACCOUNT', 'DESC_ACCOUNT', 
                                  'Montants', 'TOP_ACCOUNT', 'TOP_ACCOUNT_DESC', 'Type']]
        df_display.columns = ['ID Agence', 'Agence', 'ID Compte', 'Description Compte', 
                               'Montant', 'TOP_ACCOUNT', 'Catégorie', 'Type']
        
        st.dataframe(df_display, use_container_width=True, hide_index=True)
        
        # Export
        st.divider()
        
        excel_data = convert_balance_to_excel(df_filtered)
        csv_data   = convert_balance_to_csv(df_filtered)
        _cf1, _cf2 = st.columns(2)
        with _cf1:
            st.download_button(label="📥 Données filtrées (Excel)", data=excel_data,
                file_name=f"Balance_Filtree_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_fil_xl", use_container_width=True)
        with _cf2:
            st.download_button(label="📥 Données filtrées (CSV)", data=csv_data,
                file_name=f"Balance_Filtree_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv", key="dl_fil_csv", use_container_width=True)


def parse_balance_pcec_immos(txt_content):
    """
    Parser Balance Générale PCEC — format spcpt062
    Garde uniquement les PCEC à exactement 4 chiffres.
    Stratégie rapide : lit toutes les lignes avec | en une passe,
    utilise pandas vectorisé pour le nettoyage.
    """
    import io

    rows = []
    current_pcec = None

    for line in txt_content.replace('\r', '').split('\n'):
        if '|' not in line or '*---' in line or 'SOLDES' in line:
            continue
        parts = line.split('|')
        if len(parts) < 4:
            continue

        pcec_raw = parts[1].strip()

        # Extraire les 2 dernières valeurs numériques significatives de la ligne
        def last_two_nums(parts_slice):
            vals = []
            for p in parts_slice:
                s = p.strip().replace(' ', '').replace(',', '.')
                s2 = s.replace('.', '').replace('-', '')
                if s2.isdigit() and len(s2) > 2:
                    try:
                        vals.append(float(s))
                    except ValueError:
                        pass
            return (vals[-2], vals[-1]) if len(vals) >= 2 else (0.0, vals[-1] if vals else 0.0)

        if re.match(r'^\d{4}$', pcec_raw):
            current_pcec = pcec_raw
            intitule = parts[2].strip() if len(parts) > 2 else ''
            deb, cred = last_two_nums(parts[3:])
            rows.append({'PCEC': pcec_raw, 'Intitule': intitule,
                         'sol_deb': deb, 'sol_cred': cred})

        elif pcec_raw == '' and current_pcec is not None:
            deb, cred = last_two_nums(parts[3:])
            if deb > 0 or cred > 0:
                rows[-1]['sol_deb']  += deb
                rows[-1]['sol_cred'] += cred

        elif re.match(r'^\d+$', pcec_raw):
            current_pcec = None  # code non-PCEC → reset

    if not rows:
        return pd.DataFrame(columns=['PCEC','Intitule','Solde_Debiteur','Solde_Crediteur','Montant'])

    df = pd.DataFrame(rows)
    # Consolider les PCEC dupliqués (cas multi-devises déjà sommé mais sécurité)
    df = df.groupby(['PCEC','Intitule'], as_index=False)[['sol_deb','sol_cred']].sum()
    df['Montant']         = (df['sol_cred'] - df['sol_deb']).round(2)
    df['Solde_Debiteur']  = df['sol_deb'].round(2)
    df['Solde_Crediteur'] = df['sol_cred'].round(2)
    return df[['PCEC','Intitule','Solde_Debiteur','Solde_Crediteur','Montant']]

def balance_module():
    """
    Module principal Balance - Import et Analyse Produits/Charges
    """
    
    st.markdown(f"""
    <div class="main-header">
        <h1>📊 Module Balance</h1>
        <p style='font-size: 1.2em; margin-top: 1rem;'>
            Traitement automatique des fichiers Balance (Produits & Charges)
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # =========================================================================
    # ONGLETS PRINCIPAUX
    # =========================================================================
    
    tab1, tab2, tab3, tab4 = st.tabs([
        "📂 Import Balance",
        "📊 Visualisations",
        "⚖️ Comparaison de Balances",
        "🏗️ Immos"
    ])
    
    # ─────────────────────────────────────────────────────────────────────────
    # TAB 1 : IMPORT
    # ─────────────────────────────────────────────────────────────────────────
    
    with tab1:
        st.subheader("📂 Import Fichiers Balance")
        
        st.markdown("""
        <div class="info-box">
            <h4>📋 Instructions</h4>
            <ul>
                <li>Uploadez deux fichiers TXT : un pour les <strong>Produits</strong> et un pour les <strong>Charges</strong></li>
                <li>Format attendu : Balance Générale SGMB (format standard)</li>
                <li>Les fichiers seront automatiquement parsés et convertis en Excel</li>
                <li>Les visualisations seront disponibles dans l'onglet suivant</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
        
        st.divider()
        
        col1, col2 = st.columns(2)
        
        # ─────────────────────────────────────────────────────────────────────
        # UPLOAD PRODUITS
        # ─────────────────────────────────────────────────────────────────────
        
        with col1:
            st.markdown("### 💰 Fichier Produits")
            
            fichier_produits = st.file_uploader(
                "Sélectionnez le fichier TXT des Produits",
                type=['txt'],
                key="fichier_produits"
            )
            
            if fichier_produits:
                try:
                    with st.spinner("⏳ Traitement du fichier Produits..."):
                        # Lire le contenu
                        content = fichier_produits.read().decode('utf-8', errors='ignore')
                        
                        # Parser
                        df_produits = parse_balance_txt(content, type_document="Produits")
                        
                        st.success(f"✅ **{len(df_produits):,} lignes** extraites".replace(',', ' '))
                        
                        # Stocker dans session_state
                        st.session_state.balance_produits = df_produits
                        
                        # Métriques
                        total_produits = df_produits['Montants'].sum()
                        nb_agences = df_produits['ID_AGENCI'].nunique()
                        nb_comptes = df_produits['ID_ACCOUNT'].nunique()
                        
                        col_m1, col_m2, col_m3 = st.columns(3)
                        with col_m1:
                            st.metric("Total", f"{total_produits/1e6:.2f} Mrd")
                        with col_m2:
                            st.metric("Agences", nb_agences)
                        with col_m3:
                            st.metric("Comptes", nb_comptes)
                        
                        # Aperçu
                        with st.expander("🔍 Aperçu des données", expanded=False):
                            st.dataframe(df_produits.head(20), use_container_width=True)
                        
                        # Export Excel + CSV
                        excel_produits = convert_balance_to_excel(df_produits, "Produits.xlsx")
                        csv_produits   = convert_balance_to_csv(df_produits)
                        _cp1, _cp2 = st.columns(2)
                        with _cp1:
                            st.download_button(label="📥 Produits (Excel)", data=excel_produits,
                                file_name=f"Balance_Produits_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_prod_xl", use_container_width=True)
                        with _cp2:
                            st.download_button(label="📥 Produits (CSV)", data=csv_produits,
                                file_name=f"Balance_Produits_{datetime.now().strftime('%Y%m%d')}.csv",
                                mime="text/csv", key="dl_prod_csv", use_container_width=True)
                        
                except Exception as e:
                    st.error(f"❌ Erreur lors du traitement : {str(e)}")
                    st.exception(e)
        
        # ─────────────────────────────────────────────────────────────────────
        # UPLOAD CHARGES
        # ─────────────────────────────────────────────────────────────────────
        
        with col2:
            st.markdown("### 💸 Fichier Charges")
            
            fichier_charges = st.file_uploader(
                "Sélectionnez le fichier TXT des Charges",
                type=['txt'],
                key="fichier_charges"
            )
            
            if fichier_charges:
                try:
                    with st.spinner("⏳ Traitement du fichier Charges..."):
                        # Lire le contenu
                        content = fichier_charges.read().decode('utf-8', errors='ignore')
                        
                        # Parser
                        df_charges = parse_balance_txt(content, type_document="Charges")
                        
                        st.success(f"✅ **{len(df_charges):,} lignes** extraites".replace(',', ' '))
                        
                        # Stocker dans session_state
                        st.session_state.balance_charges = df_charges
                        
                        # Métriques
                        total_charges = df_charges['Montants'].sum()
                        nb_agences = df_charges['ID_AGENCI'].nunique()
                        nb_comptes = df_charges['ID_ACCOUNT'].nunique()
                        
                        col_m1, col_m2, col_m3 = st.columns(3)
                        with col_m1:
                            st.metric("Total", f"{total_charges/1e6:.2f} Mrd")
                        with col_m2:
                            st.metric("Agences", nb_agences)
                        with col_m3:
                            st.metric("Comptes", nb_comptes)
                        
                        # Aperçu
                        with st.expander("🔍 Aperçu des données", expanded=False):
                            st.dataframe(df_charges.head(20), use_container_width=True)
                        
                        # Export Excel + CSV
                        excel_charges = convert_balance_to_excel(df_charges, "Charges.xlsx")
                        csv_charges   = convert_balance_to_csv(df_charges)
                        _cc1, _cc2 = st.columns(2)
                        with _cc1:
                            st.download_button(label="📥 Charges (Excel)", data=excel_charges,
                                file_name=f"Balance_Charges_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_charg_xl", use_container_width=True)
                        with _cc2:
                            st.download_button(label="📥 Charges (CSV)", data=csv_charges,
                                file_name=f"Balance_Charges_{datetime.now().strftime('%Y%m%d')}.csv",
                                mime="text/csv", key="dl_charg_csv", use_container_width=True)
                        
                except Exception as e:
                    st.error(f"❌ Erreur lors du traitement : {str(e)}")
                    st.exception(e)
    
    # ─────────────────────────────────────────────────────────────────────────
    # TAB 2 : VISUALISATIONS
    # ─────────────────────────────────────────────────────────────────────────
    
    with tab2:
        if 'balance_produits' not in st.session_state or 'balance_charges' not in st.session_state:
            st.warning("⚠️ Veuillez d'abord importer les fichiers Produits et Charges dans l'onglet **Import Balance**")
        else:
            df_produits = st.session_state.balance_produits
            df_charges = st.session_state.balance_charges
            
            visualisations_balance(df_produits, df_charges)
    
    # ─────────────────────────────────────────────────────────────────────────
    # TAB 3 : COMPARAISON DE BALANCES
    # ─────────────────────────────────────────────────────────────────────────
    
    with tab3:
        st.subheader("⚖️ Comparaison de Deux Balances")
        
        st.markdown("""
        <div class="info-box">
            <h4>📋 Instructions</h4>
            <ul>
                <li>Uploadez <strong>deux balances</strong> (Produits ou Charges) à comparer</li>
                <li>La comparaison se fera par <strong>ID_AGENCI</strong> et <strong>ID_ACCOUNT</strong></li>
                <li>Vous verrez les différences de montants entre les deux balances</li>
                <li>Format attendu : Fichiers TXT Balance SGMB</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
        
        st.divider()
        
        # Sélection du type de balance à comparer
        type_balance = st.radio(
            "Type de balance à comparer",
            ["Produits", "Charges"],
            horizontal=True
        )
        
        st.divider()
        
        col1, col2 = st.columns(2)
        
        # ─────────────────────────────────────────────────────────────────────
        # UPLOAD BALANCE 1
        # ─────────────────────────────────────────────────────────────────────
        
        with col1:
            st.markdown(f"### 1️⃣ Première Balance {type_balance}")
            
            fichier_balance1 = st.file_uploader(
                f"Sélectionnez le fichier TXT - Balance 1 ({type_balance})",
                type=['txt'],
                key="fichier_balance1_comp"
            )
            
            if fichier_balance1:
                try:
                    with st.spinner("⏳ Traitement Balance 1..."):
                        content1 = fichier_balance1.read().decode('utf-8', errors='ignore')
                        df_balance1 = parse_balance_txt(content1, type_document=type_balance)
                        
                        st.success(f"✅ **{len(df_balance1):,} lignes** extraites".replace(',', ' '))
                        
                        # Stocker dans session_state
                        st.session_state.balance1_comp = df_balance1
                        
                        # Métriques
                        total_b1 = df_balance1['Montants'].sum()
                        nb_agences_b1 = df_balance1['ID_AGENCI'].nunique()
                        nb_comptes_b1 = df_balance1['ID_ACCOUNT'].nunique()
                        
                        col_m1, col_m2, col_m3 = st.columns(3)
                        with col_m1:
                            st.metric("Total", f"{total_b1/1e6:.2f} Mrd")
                        with col_m2:
                            st.metric("Agences", nb_agences_b1)
                        with col_m3:
                            st.metric("Comptes", nb_comptes_b1)
                        
                        with st.expander("🔍 Aperçu Balance 1", expanded=False):
                            st.dataframe(df_balance1.head(10), use_container_width=True)
                        
                except Exception as e:
                    st.error(f"❌ Erreur lors du traitement : {str(e)}")
        
        # ─────────────────────────────────────────────────────────────────────
        # UPLOAD BALANCE 2
        # ─────────────────────────────────────────────────────────────────────
        
        with col2:
            st.markdown(f"### 2️⃣ Deuxième Balance {type_balance}")
            
            fichier_balance2 = st.file_uploader(
                f"Sélectionnez le fichier TXT - Balance 2 ({type_balance})",
                type=['txt'],
                key="fichier_balance2_comp"
            )
            
            if fichier_balance2:
                try:
                    with st.spinner("⏳ Traitement Balance 2..."):
                        content2 = fichier_balance2.read().decode('utf-8', errors='ignore')
                        df_balance2 = parse_balance_txt(content2, type_document=type_balance)
                        
                        st.success(f"✅ **{len(df_balance2):,} lignes** extraites".replace(',', ' '))
                        
                        # Stocker dans session_state
                        st.session_state.balance2_comp = df_balance2
                        
                        # Métriques
                        total_b2 = df_balance2['Montants'].sum()
                        nb_agences_b2 = df_balance2['ID_AGENCI'].nunique()
                        nb_comptes_b2 = df_balance2['ID_ACCOUNT'].nunique()
                        
                        col_m1, col_m2, col_m3 = st.columns(3)
                        with col_m1:
                            st.metric("Total", f"{total_b2/1e6:.2f} Mrd")
                        with col_m2:
                            st.metric("Agences", nb_agences_b2)
                        with col_m3:
                            st.metric("Comptes", nb_comptes_b2)
                        
                        with st.expander("🔍 Aperçu Balance 2", expanded=False):
                            st.dataframe(df_balance2.head(10), use_container_width=True)
                        
                except Exception as e:
                    st.error(f"❌ Erreur lors du traitement : {str(e)}")
        
        st.divider()
        
        # ─────────────────────────────────────────────────────────────────────
        # BOUTON DE COMPARAISON
        # ─────────────────────────────────────────────────────────────────────
        
        if 'balance1_comp' in st.session_state and 'balance2_comp' in st.session_state:
            
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                if st.button("🔄 Comparer les Deux Balances", type="primary", use_container_width=True):
                    with st.spinner("⏳ Comparaison en cours..."):
                        try:
                            df_comparison = comparer_deux_balances(
                                st.session_state.balance1_comp,
                                st.session_state.balance2_comp,
                                nom_balance1="Balance 1",
                                nom_balance2="Balance 2"
                            )
                            
                            if df_comparison is not None:
                                st.session_state.comparison_result = df_comparison
                                st.success("✅ Comparaison terminée !")
                                st.rerun()
                        
                        except Exception as e:
                            st.error(f"❌ Erreur lors de la comparaison : {str(e)}")
                            st.exception(e)
        
        else:
            st.info("👆 Veuillez charger les deux balances ci-dessus pour lancer la comparaison")
        
        # ─────────────────────────────────────────────────────────────────────
        # AFFICHAGE DES RÉSULTATS
        # ─────────────────────────────────────────────────────────────────────
        
        if 'comparison_result' in st.session_state:
            st.divider()
            st.subheader("📊 Résultats de la Comparaison")
            
            df_comp = st.session_state.comparison_result
            
            # ═════════════════════════════════════════════════════════════════
            # MÉTRIQUES GLOBALES
            # ═════════════════════════════════════════════════════════════════
            
            total_b1 = df_comp['Montant Balance 1'].sum()
            total_b2 = df_comp['Montant Balance 2'].sum()
            diff_totale = df_comp['Différence'].sum()
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric(
                    "Total Balance 1",
                    f"{total_b1/1e6:.2f} Mrd",
                )
            
            with col2:
                st.metric(
                    "Total Balance 2",
                    f"{total_b2/1e6:.2f} Mrd",
                )
            
            with col3:
                delta_pct = ((total_b1 - total_b2) / total_b2 * 100) if total_b2 != 0 else 0
                st.metric(
                    "Différence Totale",
                    f"{diff_totale/1e6:.2f} Mrd",
                    delta=f"{delta_pct:.2f}%"
                )
            
            with col4:
                nb_differences = len(df_comp[df_comp['Différence'] != 0])
                st.metric(
                    "Lignes avec différence",
                    f"{nb_differences:,}".replace(',', ' ')
                )
            
            st.divider()
            
            # ═════════════════════════════════════════════════════════════════
            # FILTRES
            # ═════════════════════════════════════════════════════════════════
            
            st.markdown("### 🔍 Filtres")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                filtre_statut = st.multiselect(
                    "Statut",
                    options=df_comp['Statut'].unique(),
                    default=df_comp['Statut'].unique()
                )
            
            with col2:
                filtre_presence = st.multiselect(
                    "Présence",
                    options=df_comp['Présence'].unique(),
                    default=df_comp['Présence'].unique()
                )
            
            with col3:
                afficher_zeros = st.checkbox("Afficher les différences nulles", value=False)
            
            # Appliquer les filtres
            df_filtered = df_comp[
                (df_comp['Statut'].isin(filtre_statut)) &
                (df_comp['Présence'].isin(filtre_presence))
            ]
            
            if not afficher_zeros:
                df_filtered = df_filtered[df_filtered['Différence'] != 0]
            
            # ═════════════════════════════════════════════════════════════════
            # TABLEAU DE RÉSULTATS
            # ═════════════════════════════════════════════════════════════════
            
            st.divider()
            st.markdown(f"### 📋 Détails ({len(df_filtered):,} lignes)".replace(',', ' '))
            
            # Formater les colonnes numériques
            df_display = df_filtered.copy()
            df_display['Montant Balance 1'] = df_display['Montant Balance 1'].apply(lambda x: f"{x:,.2f}".replace(',', ' '))
            df_display['Montant Balance 2'] = df_display['Montant Balance 2'].apply(lambda x: f"{x:,.2f}".replace(',', ' '))
            df_display['Différence'] = df_display['Différence'].apply(lambda x: f"{x:,.2f}".replace(',', ' '))
            df_display['Différence_%'] = df_display['Différence_%'].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else "N/A")
            
            st.dataframe(
                df_display,
                use_container_width=True,
                height=400
            )
            
            # ═════════════════════════════════════════════════════════════════
            # EXPORT EXCEL
            # ═════════════════════════════════════════════════════════════════
            
            st.divider()
            
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                # Créer Excel
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_filtered.to_excel(writer, index=False, sheet_name='Comparaison')
                excel_data = output.getvalue()
                
                st.download_button(
                    label="📥 Télécharger la Comparaison (Excel)",
                    data=excel_data,
                    file_name=f"Comparaison_Balances_{type_balance}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )


    # ─────────────────────────────────────────────────────────────────────────
    # TAB 4 : IMMOS
    # ─────────────────────────────────────────────────────────────────────────
    with tab4:
        st.subheader("🏗️ Rapprochement Immobilisations — Balance PCEC vs Immos")

        st.markdown("""
        <div class="info-box" style="color:#1a1a1a !important;">
            <h4>📋 Instructions — 3 fichiers requis</h4>
            <ul>
                <li><strong>Fichier 1 — Balance PCEC :</strong> contient les colonnes <code>PCEC</code> et <code>Montant</code> (Solde Créditeur − Solde Débiteur)</li>
                <li><strong>Fichier 2 — Clé de mapping :</strong> tableau orange uniquement — fait le lien <code>PCEC → PCI</code></li>
                <li><strong>Fichier 3 — Immos :</strong> contient <code>CPTE-IMMOS</code> (= PCI) et le montant de transaction</li>
            </ul>
            <p>Le résultat est un tableau à 4 colonnes : <strong>PCEC | Montant Balance | Montant Immos | Écart</strong></p>
        </div>
        """, unsafe_allow_html=True)

        st.divider()

        # ── Session state init ────────────────────────────────────────────────
        for _k in ['immos_balance_df', 'immos_mapping_df', 'immos_file3_df', 'immos_result_df',
                   'immos_f1_name', 'immos_f2_name', 'immos_f3_name']:
            if _k not in st.session_state:
                st.session_state[_k] = None

        # ── Upload 3 fichiers — persistants en session ────────────────────────
        col_f1, col_f2, col_f3 = st.columns(3)

        with col_f1:
            st.markdown("#### 📄 Fichier 1 — Balance PCEC")
            # Afficher statut si déjà chargé
            if st.session_state.immos_balance_df is not None:
                df_loaded = st.session_state.immos_balance_df
                st.success(f"✅ **{st.session_state.immos_f1_name}** — {len(df_loaded)} lignes en mémoire")
                with st.expander("Aperçu", expanded=False):
                    st.dataframe(df_loaded.head(5), use_container_width=True)
                if st.button("🗑️ Supprimer fichier 1", key="del_immos_f1"):
                    st.session_state.immos_balance_df = None
                    st.session_state.immos_f1_name    = None
                    st.session_state.immos_result_df  = None
                    st.rerun()
            else:
                f1 = st.file_uploader("Balance PCEC (TXT, Excel ou CSV)",
                                       type=['txt','xlsx','xls','csv'], key="immos_f1")
                if f1:
                    try:
                        if f1.name.lower().endswith('.txt'):
                            raw = f1.read().decode('utf-8', errors='ignore')
                            df_f1 = parse_balance_pcec_immos(raw)
                            if df_f1.empty:
                                st.error("❌ Aucune donnée PCEC détectée. Vérifiez le format du fichier.")
                            else:
                                st.session_state.immos_balance_df = df_f1
                                st.session_state.immos_f1_name    = f1.name
                                st.rerun()
                        elif f1.name.lower().endswith('.csv'):
                            df_f1 = pd.read_csv(f1, dtype=str)
                            st.session_state.immos_balance_df = df_f1
                            st.session_state.immos_f1_name    = f1.name
                            st.rerun()
                        else:
                            df_f1 = pd.read_excel(f1, dtype=str)
                            st.session_state.immos_balance_df = df_f1
                            st.session_state.immos_f1_name    = f1.name
                            st.rerun()
                    except Exception as e:
                        st.error(f"Erreur lecture fichier 1 : {e}")

        with col_f2:
            st.markdown("#### 🔑 Fichier 2 — Clé Mapping PCEC→PCI")
            if st.session_state.immos_mapping_df is not None:
                df_loaded = st.session_state.immos_mapping_df
                st.success(f"✅ **{st.session_state.immos_f2_name}** — {len(df_loaded)} lignes en mémoire")
                with st.expander("Aperçu", expanded=False):
                    st.dataframe(df_loaded.head(5), use_container_width=True)
                if st.button("🗑️ Supprimer fichier 2", key="del_immos_f2"):
                    st.session_state.immos_mapping_df = None
                    st.session_state.immos_f2_name    = None
                    st.session_state.immos_result_df  = None
                    st.rerun()
            else:
                f2 = st.file_uploader("Mapping (Excel)", type=['xlsx','xls'], key="immos_f2")
                if f2:
                    try:
                        xf = pd.ExcelFile(f2)
                        df_f2_raw = None
                        for sheet in xf.sheet_names:
                            _df = pd.read_excel(xf, sheet_name=sheet, header=None, dtype=str)
                            for i in range(min(30, len(_df))):
                                row_vals = [str(v).strip().upper() for v in _df.iloc[i].dropna().tolist()]
                                if any('PCEC' in v for v in row_vals) and any('PCI' in v for v in row_vals):
                                    df_f2_raw = pd.read_excel(xf, sheet_name=sheet, header=i, dtype=str)
                                    df_f2_raw.columns = [str(c).strip() for c in df_f2_raw.columns]
                                    df_f2_raw = df_f2_raw.dropna(how='all')
                                    break
                            if df_f2_raw is not None:
                                break
                        if df_f2_raw is None:
                            df_f2_raw = pd.read_excel(f2, dtype=str)
                            df_f2_raw.columns = [str(c).strip() for c in df_f2_raw.columns]
                        st.session_state.immos_mapping_df = df_f2_raw
                        st.session_state.immos_f2_name    = f2.name
                        st.rerun()
                    except Exception as e:
                        st.error(f"Erreur lecture fichier 2 : {e}")

        with col_f3:
            st.markdown("#### 🏗️ Fichier 3 — Immos (CPTE-IMMOS)")
            if st.session_state.immos_file3_df is not None:
                df_loaded = st.session_state.immos_file3_df
                st.success(f"✅ **{st.session_state.immos_f3_name}** — {len(df_loaded)} lignes en mémoire")
                with st.expander("Aperçu", expanded=False):
                    st.dataframe(df_loaded.head(5), use_container_width=True)
                if st.button("🗑️ Supprimer fichier 3", key="del_immos_f3"):
                    st.session_state.immos_file3_df  = None
                    st.session_state.immos_f3_name   = None
                    st.session_state.immos_result_df = None
                    st.rerun()
            else:
                f3 = st.file_uploader("Immos (Excel ou CSV)", type=['xlsx','xls','csv'], key="immos_f3")
                if f3:
                    try:
                        if f3.name.endswith('.csv'):
                            df_f3 = pd.read_csv(f3, dtype=str)
                        else:
                            df_f3 = pd.read_excel(f3, dtype=str)
                        st.session_state.immos_file3_df = df_f3
                        st.session_state.immos_f3_name  = f3.name
                        st.rerun()
                    except Exception as e:
                        st.error(f"Erreur lecture fichier 3 : {e}")

        st.divider()

        # ── Session state : sélections colonnes persistantes ─────────────────
        for _k in ['immos_col_pci_f3',   'immos_col_pcec_f2',  'immos_col_pci_f2',
                   'immos_col_pcec_f1',  'immos_mode_mont',
                   'immos_col_mont_f1',  'immos_col_cred_f1',  'immos_col_debit_f1',
                   # Tableau 1 — Bilan Immos
                   'immos_t1_pci_col',   'immos_t1_mont_col',
                   # Tableau 2 — Dotations
                   'immos_t2_pci_col',   'immos_t2_mont_col',
                   # Tableau 3 — Libre
                   'immos_t3_pci_col',   'immos_t3_mont_col',  'immos_t3_label',
                   # Résultats calculés
                   'immos_res_t1', 'immos_res_t2', 'immos_res_t3']:
            if _k not in st.session_state:
                st.session_state[_k] = None

        df_bal = st.session_state.immos_balance_df
        df_map = st.session_state.immos_mapping_df
        df_im3 = st.session_state.immos_file3_df

        all_loaded = (df_bal is not None and df_map is not None and df_im3 is not None)

        if not all_loaded:
            st.info("⬆️ Chargez les 3 fichiers ci-dessus pour configurer le rapprochement.")
        else:
            cols_f1 = list(df_bal.columns)
            cols_f2 = list(df_map.columns)
            cols_f3 = list(df_im3.columns)

            # ── Helper : index sécurisé ───────────────────────────────────────
            def safe_idx(lst, val, fallback=0):
                try:
                    return lst.index(val) if val in lst else fallback
                except Exception:
                    return fallback

            # ── Détection automatique defaults ────────────────────────────────
            def auto(lst, *keywords):
                for kw in keywords:
                    hit = next((c for c in lst if kw.upper() in str(c).upper()), None)
                    if hit: return hit
                return lst[0]

            st.markdown("### ⚙️ Configuration des colonnes")
            st.caption("Les sélections sont sauvegardées — rechargez les fichiers sans reconfigurer.")

            with st.expander("🔧 Paramètres de configuration", expanded=(st.session_state.immos_col_pci_f3 is None)):

                st.markdown("#### Fichier 1 — Balance PCEC")
                c1a, c1b = st.columns(2)
                with c1a:
                    col_pcec_f1 = st.selectbox("Colonne PCEC",
                        cols_f1,
                        index=safe_idx(cols_f1, st.session_state.immos_col_pcec_f1,
                                       safe_idx(cols_f1, auto(cols_f1,'PCEC','COMPTE','CPT'))),
                        key="cfg_pcec_f1")
                with c1b:
                    is_txt = 'Montant' in cols_f1 and 'Solde_Debiteur' in cols_f1
                    if is_txt:
                        st.info("✅ TXT parsé — montant calculé automatiquement")
                        mode_mont    = "direct"
                        col_mont_f1  = 'Montant'
                        col_cred_f1  = col_debit_f1 = None
                    else:
                        mode_mont = st.radio("Mode montant",
                            ["Créditeur − Débiteur", "Colonne directe"],
                            index=0 if st.session_state.immos_mode_mont != "direct" else 1,
                            key="cfg_mode_mont", horizontal=True)
                        if mode_mont == "Créditeur − Débiteur":
                            cc1, cc2 = st.columns(2)
                            col_cred_f1  = cc1.selectbox("Solde Créditeur", cols_f1,
                                index=safe_idx(cols_f1, st.session_state.immos_col_cred_f1,
                                               safe_idx(cols_f1, auto(cols_f1,'CREDIT','CRÉDIT'))),
                                key="cfg_cred_f1")
                            col_debit_f1 = cc2.selectbox("Solde Débiteur",  cols_f1,
                                index=safe_idx(cols_f1, st.session_state.immos_col_debit_f1,
                                               safe_idx(cols_f1, auto(cols_f1,'DEBIT','DÉBIT'))),
                                key="cfg_debit_f1")
                            col_mont_f1 = None
                        else:
                            col_mont_f1  = st.selectbox("Colonne Montant", cols_f1,
                                index=safe_idx(cols_f1, st.session_state.immos_col_mont_f1,
                                               safe_idx(cols_f1, auto(cols_f1,'MONTANT','SOLDE'))),
                                key="cfg_mont_f1")
                            col_cred_f1 = col_debit_f1 = None

                st.markdown("#### Fichier 2 — Mapping")
                c2a, c2b = st.columns(2)
                col_pcec_f2 = c2a.selectbox("Colonne PCEC", cols_f2,
                    index=safe_idx(cols_f2, st.session_state.immos_col_pcec_f2,
                                   safe_idx(cols_f2, auto(cols_f2,'PCEC'))),
                    key="cfg_pcec_f2")
                col_pci_f2  = c2b.selectbox("Colonne PCI",  cols_f2,
                    index=safe_idx(cols_f2, st.session_state.immos_col_pci_f2,
                                   safe_idx(cols_f2, auto(cols_f2,'PCI'))),
                    key="cfg_pci_f2")

                st.markdown("#### Fichier 3 — 3 Tableaux")
                st.caption("Colonne PCI commune aux 3 tableaux :")
                col_pci_f3 = st.selectbox("Colonne PCI (fichier 3)",  cols_f3,
                    index=safe_idx(cols_f3, st.session_state.immos_col_pci_f3,
                                   safe_idx(cols_f3, auto(cols_f3,'IMMO','PCI','CPTE','COMPTE'))),
                    key="cfg_pci_f3")

                st.markdown("**Tableau 1 — Bilan Immos** (ex: CPTE_Immo_Compte_Comptable + Coût)")
                tb1a, tb1b = st.columns(2)
                t1_pci_col  = tb1a.selectbox("Colonne PCI tableau 1",   cols_f3,
                    index=safe_idx(cols_f3, st.session_state.immos_t1_pci_col,
                                   safe_idx(cols_f3, auto(cols_f3,'IMMO_COMPTE','IMMO','CPTE_I'))),
                    key="cfg_t1_pci")
                t1_mont_col = tb1b.selectbox("Colonne montant tableau 1", cols_f3,
                    index=safe_idx(cols_f3, st.session_state.immos_t1_mont_col,
                                   safe_idx(cols_f3, auto(cols_f3,'COUT','COÛT','COST'))),
                    key="cfg_t1_mont")

                st.markdown("**Tableau 2 — Dotations** (ex: CPTE_Dotation_Compte_Comptable + Dotation exercice)")
                tb2a, tb2b = st.columns(2)
                t2_pci_col  = tb2a.selectbox("Colonne PCI tableau 2",   cols_f3,
                    index=safe_idx(cols_f3, st.session_state.immos_t2_pci_col,
                                   safe_idx(cols_f3, auto(cols_f3,'DOTATION_COMPTE','DOT_CPT','DOTATION'))),
                    key="cfg_t2_pci")
                t2_mont_col = tb2b.selectbox("Colonne montant tableau 2", cols_f3,
                    index=safe_idx(cols_f3, st.session_state.immos_t2_mont_col,
                                   safe_idx(cols_f3, auto(cols_f3,'DOTATION_EX','DOT_EX','EXERCICE'))),
                    key="cfg_t2_mont")

                st.markdown("**Tableau 3 — Libre**")
                tb3a, tb3b, tb3c = st.columns(3)
                t3_pci_col  = tb3a.selectbox("Colonne PCI tableau 3",   cols_f3,
                    index=safe_idx(cols_f3, st.session_state.immos_t3_pci_col, 0),
                    key="cfg_t3_pci")
                t3_mont_col = tb3b.selectbox("Colonne montant tableau 3", cols_f3,
                    index=safe_idx(cols_f3, st.session_state.immos_t3_mont_col,
                                   min(1, len(cols_f3)-1)),
                    key="cfg_t3_mont")
                t3_label    = tb3c.text_input("Nom du tableau 3",
                    value=st.session_state.immos_t3_label or "Tableau 3",
                    key="cfg_t3_label")

                st.divider()
                col_sv1, col_sv2, col_sv3 = st.columns([1,2,1])
                with col_sv2:
                    if st.button("💾 Sauvegarder la configuration", type="primary",
                                  use_container_width=True, key="btn_save_cfg"):
                        st.session_state.immos_col_pcec_f1  = col_pcec_f1
                        st.session_state.immos_mode_mont    = mode_mont
                        st.session_state.immos_col_mont_f1  = col_mont_f1
                        st.session_state.immos_col_cred_f1  = col_cred_f1
                        st.session_state.immos_col_debit_f1 = col_debit_f1
                        st.session_state.immos_col_pcec_f2  = col_pcec_f2
                        st.session_state.immos_col_pci_f2   = col_pci_f2
                        st.session_state.immos_col_pci_f3   = col_pci_f3
                        st.session_state.immos_t1_pci_col   = t1_pci_col
                        st.session_state.immos_t1_mont_col  = t1_mont_col
                        st.session_state.immos_t2_pci_col   = t2_pci_col
                        st.session_state.immos_t2_mont_col  = t2_mont_col
                        st.session_state.immos_t3_pci_col   = t3_pci_col
                        st.session_state.immos_t3_mont_col  = t3_mont_col
                        st.session_state.immos_t3_label     = t3_label
                        # Reset résultats pour forcer recalcul
                        st.session_state.immos_res_t1 = None
                        st.session_state.immos_res_t2 = None
                        st.session_state.immos_res_t3 = None
                        st.success("✅ Configuration sauvegardée !")
                        st.rerun()

            # ── Utiliser config sauvegardée si disponible ─────────────────────
            cfg_ok = st.session_state.immos_col_pci_f3 is not None

            if cfg_ok:
                col_pcec_f1  = st.session_state.immos_col_pcec_f1
                mode_mont    = st.session_state.immos_mode_mont
                col_mont_f1  = st.session_state.immos_col_mont_f1
                col_cred_f1  = st.session_state.immos_col_cred_f1
                col_debit_f1 = st.session_state.immos_col_debit_f1
                col_pcec_f2  = st.session_state.immos_col_pcec_f2
                col_pci_f2   = st.session_state.immos_col_pci_f2
                t1_pci_col   = st.session_state.immos_t1_pci_col
                t1_mont_col  = st.session_state.immos_t1_mont_col
                t2_pci_col   = st.session_state.immos_t2_pci_col
                t2_mont_col  = st.session_state.immos_t2_mont_col
                t3_pci_col   = st.session_state.immos_t3_pci_col
                t3_mont_col  = st.session_state.immos_t3_mont_col
                t3_label     = st.session_state.immos_t3_label or "Tableau 3"

                st.divider()
                col_btn1, col_btn2, col_btn3 = st.columns([1,2,1])
                with col_btn2:
                    calc_immos = st.button("🔗 Calculer les 3 Rapprochements",
                                           type="primary", use_container_width=True,
                                           key="btn_calc_immos")

                if calc_immos:
                    with st.spinner("⏳ Rapprochement en cours..."):
                        try:
                            # ── Helpers vectorisés ────────────────────────────
                            def clean_num(series):
                                s = (series.astype(str)
                                     .str.strip().str.strip('"')
                                     .str.replace(r'[\s\xa0\u202f\u00a0]', '', regex=True)
                                     .str.replace(',', '.', regex=False))
                                return pd.to_numeric(s, errors='coerce').fillna(0.0)

                            def clean_code(series):
                                return (series.astype(str)
                                        .str.strip().str.strip('"')
                                        .str.replace(r'[\s\xa0\u202f\u00a0]', '', regex=True)
                                        .str.replace(r'\.0$', '', regex=True))

                            # ── Balance PCEC → dict {pcec: montant} ───────────
                            df_b = df_bal[[col_pcec_f1]].copy()
                            df_b['PCEC'] = clean_code(df_b[col_pcec_f1])
                            if mode_mont == 'Créditeur − Débiteur':
                                df_b['Mont_Bal'] = (clean_num(df_bal[col_cred_f1])
                                                    - clean_num(df_bal[col_debit_f1]))
                            else:
                                df_b['Mont_Bal'] = clean_num(df_bal[col_mont_f1])
                            df_b = (df_b[['PCEC','Mont_Bal']]
                                    .query('PCEC.str.len() > 0')
                                    .groupby('PCEC', as_index=False)['Mont_Bal'].sum())

                            # ── Mapping PCI → PCEC (1 PCI peut avoir N PCEC) ──
                            df_m = df_map[[col_pcec_f2, col_pci_f2]].copy()
                            df_m.columns = ['PCEC','PCI']
                            df_m['PCEC'] = clean_code(df_m['PCEC'])
                            df_m['PCI']  = clean_code(df_m['PCI'])
                            df_m = (df_m.dropna(subset=['PCEC','PCI'])
                                    .query('PCEC.str.len()>0 and PCI.str.len()>0'))
                            # NE PAS dédupliquer : un PCI peut mapper vers plusieurs PCEC

                            def build_table(pci_col, mont_col, label_mont):
                                """
                                Construit un tableau de rapprochement pour une paire (pci_col, mont_col).
                                Logique :
                                  - Grouper fichier3 par PCI → somme(mont_col)
                                  - Joindre mapping pour obtenir PCEC(s) associé(s)
                                  - Un PCI peut avoir plusieurs PCEC → on éclate les lignes
                                  - Pour chaque (PCI, PCEC) : montant immos = somme du PCI entier
                                    (car la relation est PCI→PCEC, pas l'inverse)
                                  - Joindre balance par PCEC pour obtenir montant balance
                                  - Garder seulement les PCI présents dans le fichier 3
                                """
                                df_i = df_im3[[pci_col, mont_col]].copy()
                                df_i.columns = ['PCI','Mont_Immos']
                                df_i['PCI']        = clean_code(df_i['PCI'])
                                df_i['Mont_Immos'] = clean_num(df_i['Mont_Immos'])
                                df_i = (df_i.dropna(subset=['PCI'])
                                        .query('PCI.str.len()>0')
                                        .groupby('PCI', as_index=False)['Mont_Immos'].sum())

                                # Joindre mapping (inner = garder seulement PCI connus)
                                df_r = pd.merge(df_i, df_m, on='PCI', how='inner')
                                # Joindre balance par PCEC (left = garder tous les PCI du fichier 3)
                                df_r = pd.merge(df_r, df_b, on='PCEC', how='left')
                                df_r['Mont_Bal']   = df_r['Mont_Bal'].fillna(0.0)
                                df_r['Ecart']      = df_r['Mont_Immos'] - df_r['Mont_Bal']
                                df_r = df_r[['PCI','PCEC','Mont_Immos','Mont_Bal','Ecart']]
                                df_r.columns = ['PCI','PCEC',
                                                label_mont,
                                                'Montant Balance (Compta)',
                                                'Écart']
                                return df_r

                            t1 = build_table(t1_pci_col, t1_mont_col, 'Montant Immos (Gestion)')
                            t2 = build_table(t2_pci_col, t2_mont_col, f'Montant {t2_mont_col}')
                            t3 = build_table(t3_pci_col, t3_mont_col, f'Montant {t3_mont_col}')

                            st.session_state.immos_res_t1 = t1
                            st.session_state.immos_res_t2 = t2
                            st.session_state.immos_res_t3 = t3
                            st.success(f"✅ Terminé — T1: {len(t1)} lignes | T2: {len(t2)} lignes | T3: {len(t3)} lignes")
                            st.rerun()

                        except Exception as e:
                            st.error(f"❌ Erreur : {e}")
                            st.exception(e)

                # ── Affichage des 3 tableaux ──────────────────────────────────
                def show_table(df_res, title, icon, key_prefix):
                    if df_res is None or df_res.empty:
                        return
                    st.markdown(f"### {icon} {title}")

                    mont_col  = df_res.columns[2]   # montant immos/dotation/libre
                    bal_col   = df_res.columns[3]   # montant balance
                    ecar_col  = df_res.columns[4]   # écart

                    # KPIs
                    k1,k2,k3,k4 = st.columns(4)
                    k1.metric("Lignes", len(df_res))
                    k2.metric(f"Total {mont_col[:18]}",
                               f"{df_res[mont_col].sum():,.0f}".replace(',', ' '))
                    k3.metric("Total Balance",
                               f"{df_res[bal_col].sum():,.0f}".replace(',', ' '))
                    k4.metric("Écart total",
                               f"{df_res[ecar_col].sum():,.0f}".replace(',', ' '),
                               delta_color="inverse" if df_res[ecar_col].sum() != 0 else "off",
                               delta=f"{df_res[ecar_col].sum():,.0f}".replace(',', ' '))

                    # Filtres
                    fc1, fc2 = st.columns(2)
                    search = fc1.text_input("🔍 Rechercher PCI / PCEC",
                                             key=f"{key_prefix}_search")
                    ecarts_only = fc2.checkbox("Écarts ≠ 0 uniquement",
                                               key=f"{key_prefix}_ecarts")

                    df_d = df_res.copy()
                    if search:
                        mask = (df_d['PCI'].str.contains(search, case=False, na=False) |
                                df_d['PCEC'].str.contains(search, case=False, na=False))
                        df_d = df_d[mask]
                    if ecarts_only:
                        df_d = df_d[df_d[ecar_col] != 0]

                    # Formater numériques
                    df_fmt = df_d.copy()
                    for c in [mont_col, bal_col, ecar_col]:
                        df_fmt[c] = df_fmt[c].apply(lambda x: f"{x:,.2f}".replace(',', ' '))

                    st.dataframe(df_fmt, use_container_width=True, hide_index=True, height=380)

                    # Export
                    ce1,ce2,ce3 = st.columns([1,2,1])
                    with ce2:
                        out = BytesIO()
                        with pd.ExcelWriter(out, engine='openpyxl') as w:
                            df_res.to_excel(w, index=False, sheet_name=title[:31])
                        out.seek(0)
                        st.download_button(
                            f"📥 Télécharger {title} (Excel)", data=out,
                            file_name=f"{key_prefix}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True, key=f"{key_prefix}_dl")
                    st.divider()

                if (st.session_state.immos_res_t1 is not None or
                    st.session_state.immos_res_t2 is not None or
                    st.session_state.immos_res_t3 is not None):

                    show_table(st.session_state.immos_res_t1,
                               "Tableau 1 — Bilan Immos", "🏗️", "t1_immos")
                    show_table(st.session_state.immos_res_t2,
                               "Tableau 2 — Dotations",   "📉", "t2_dotations")
                    show_table(st.session_state.immos_res_t3,
                               f"Tableau 3 — {st.session_state.immos_t3_label or 'Libre'}",
                               "📋", "t3_libre")

            else:
                st.info("💡 Configurez et sauvegardez les colonnes ci-dessus pour lancer le rapprochement.")


def bam_module():
    st.markdown(f"""
    <div class="main-header">
        <h1>SAHAM BANK - Module BAM</h1>
        <p style="font-size: 1.2em;">Bienvenue, <strong>{st.session_state.username}</strong></p>
        <p style="font-size: 0.9em; opacity: 0.9;">Analyse des Données Bancaires</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Navigation
    page = st.sidebar.radio(
        "Navigation",
        ["Import & Traitement", "Visualisations BAM", "Visualisations Saham Bank", "À propos"]
    )
    
    st.sidebar.divider()
    if st.session_state.uploaded_files_bam:
        st.sidebar.metric("Fichiers", len(st.session_state.uploaded_files_bam))
    if st.session_state.combined_data_bam is not None:
        st.sidebar.metric("Lignes", len(st.session_state.combined_data_bam))
    
    st.sidebar.divider()
    if st.sidebar.button("Retour au Menu Principal"):
        st.session_state.selected_module = None
        st.rerun()
    
# NOUVELLE PAGE IMPORT & TRAITEMENT AVEC 2 ONGLETS
# À intégrer dans app_saham_final.py

    if page == "Import & Traitement":
        st.header("Import & Traitement des Données")
        
        # ONGLETS POUR SÉPARER BAM ET SAHAM BANK
        tab1, tab2 = st.tabs(["📊 Import BAM", "🏦 Import Saham Bank"])
        
        # =====================================================================
        # =====================================================================
        # ONGLET 1 : IMPORT BAM MULTI-ANNÉES
        # =====================================================================
        with tab1:
            import_bam_multi_annees()
        
        # ONGLET 2 : IMPORT SAHAM BANK
        # =====================================================================
        with tab2:
            st.subheader("🏦 Import Saham Bank")

            # ═════════════════════════════════════════════════════════════════
            # SECTION A — FICHIERS DE RÉFÉRENCE (une seule fois)
            # ═════════════════════════════════════════════════════════════════
            st.markdown("""
            <div style="background:linear-gradient(135deg,#1a4d3e,#2d5f4e);border-radius:12px;
                        padding:1.2rem 1.5rem;margin-bottom:1.2rem;">
                <h4 style="color:#fff;margin:0 0 0.3rem 0;">🗂️ Étape 1 — Fichiers de Référence</h4>
                <p style="color:#c8e6c9;margin:0;font-size:0.9em;">
                    Ces deux fichiers sont chargés <strong>une seule fois</strong> et restent en mémoire.
                    Modifiez-les uniquement si votre référentiel change.
                </p>
            </div>
            """, unsafe_allow_html=True)

            ref_col1, ref_col2 = st.columns(2)

            # ── Mapping Chapitre ──────────────────────────────────────────────
            with ref_col1:
                mc_loaded = st.session_state.saham_mapping_chapitre is not None
                badge_mc  = "✅ Chargé" if mc_loaded else "❌ Non chargé"
                color_mc  = "#28a745" if mc_loaded else "#dc3545"
                st.markdown(f"""
                <div style="background:#f8f9fa;border-left:5px solid {color_mc};
                            border-radius:8px;padding:1rem;margin-bottom:0.8rem;">
                    <strong>📋 Mapping Chapitre</strong>
                    <span style="float:right;background:{color_mc};color:white;
                                 padding:2px 10px;border-radius:12px;font-size:0.8em;">{badge_mc}</span>
                    <br><small style="color:#6c757d;">Colonnes : Chapitre · Type_Solde (SD/SC) · Type_Produit (Crédit/Dépôt)</small>
                </div>
                """, unsafe_allow_html=True)

                if mc_loaded:
                    df_mc_cur = st.session_state.saham_mapping_chapitre
                    st.success(f"📋 {len(df_mc_cur):,} chapitres en mémoire")
                    with st.expander("👁️ Aperçu Mapping Chapitre"):
                        st.dataframe(df_mc_cur.head(8), use_container_width=True)
                    col_mc1, col_mc2 = st.columns(2)
                    with col_mc1:
                        if st.button("🔄 Remplacer", key="btn_replace_mc", use_container_width=True):
                            st.session_state['show_upload_mc'] = True
                            st.rerun()
                    with col_mc2:
                        if st.button("🗑️ Supprimer", key="btn_del_mc", use_container_width=True):
                            st.session_state.saham_mapping_chapitre = None
                            clear_cache("saham_mapping_chapitre")
                            st.session_state.pop('show_upload_mc', None)
                            st.rerun()
                    if st.session_state.get('show_upload_mc'):
                        f_mc_new = st.file_uploader("Nouveau fichier Mapping Chapitre",
                                                    type=["xlsx","xls"], key="mc_replace_uploader")
                        if f_mc_new:
                            try:
                                df_mc_new = pd.read_excel(f_mc_new)
                                st.session_state.saham_mapping_chapitre = df_mc_new
                                save_cache("saham_mapping_chapitre", df_mc_new)
                                st.session_state.pop('show_upload_mc', None)
                                st.success(f"✅ Remplacé — {len(df_mc_new):,} chapitres")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Erreur : {e}")
                else:
                    f_mc = st.file_uploader("📂 Charger Mapping Chapitre",
                                            type=["xlsx","xls"], key="saham_mc_upload")
                    if f_mc:
                        try:
                            df_mc = pd.read_excel(f_mc)
                            st.session_state.saham_mapping_chapitre = df_mc
                            save_cache("saham_mapping_chapitre", df_mc)
                            st.success(f"✅ {len(df_mc):,} chapitres chargés")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Erreur : {e}")

            # ── Mapping Localités-Agences ─────────────────────────────────────
            with ref_col2:
                mla_loaded = st.session_state.saham_mapping_localites_agences is not None
                badge_mla  = "✅ Chargé" if mla_loaded else "❌ Non chargé"
                color_mla  = "#28a745" if mla_loaded else "#dc3545"
                st.markdown(f"""
                <div style="background:#f8f9fa;border-left:5px solid {color_mla};
                            border-radius:8px;padding:1rem;margin-bottom:0.8rem;">
                    <strong>🗺️ Mapping Localités-Agences</strong>
                    <span style="float:right;background:{color_mla};color:white;
                                 padding:2px 10px;border-radius:12px;font-size:0.8em;">{badge_mla}</span>
                    <br><small style="color:#6c757d;">Colonnes : Code_Agence · Code_Localite · Localite</small>
                </div>
                """, unsafe_allow_html=True)

                if mla_loaded:
                    df_mla_cur = st.session_state.saham_mapping_localites_agences
                    st.success(f"🗺️ {len(df_mla_cur):,} agences en mémoire")
                    with st.expander("👁️ Aperçu Mapping Localités"):
                        st.dataframe(df_mla_cur.head(8), use_container_width=True)
                    col_mla1, col_mla2 = st.columns(2)
                    with col_mla1:
                        if st.button("🔄 Remplacer", key="btn_replace_mla", use_container_width=True):
                            st.session_state['show_upload_mla'] = True
                            st.rerun()
                    with col_mla2:
                        if st.button("🗑️ Supprimer", key="btn_del_mla", use_container_width=True):
                            st.session_state.saham_mapping_localites_agences = None
                            clear_cache("saham_mapping_localites_agences")
                            st.session_state.pop('show_upload_mla', None)
                            st.rerun()
                    if st.session_state.get('show_upload_mla'):
                        f_mla_new = st.file_uploader("Nouveau fichier Mapping Localités",
                                                     type=["xlsx","xls"], key="mla_replace_uploader")
                        if f_mla_new:
                            try:
                                df_mla_new = pd.read_excel(f_mla_new)
                                st.session_state.saham_mapping_localites_agences = df_mla_new
                                save_cache("saham_mapping_localites_agences", df_mla_new)
                                st.session_state.pop('show_upload_mla', None)
                                st.success(f"✅ Remplacé — {len(df_mla_new):,} agences")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Erreur : {e}")
                else:
                    f_mla = st.file_uploader("📂 Charger Mapping Localités",
                                             type=["xlsx","xls"], key="saham_mla_upload")
                    if f_mla:
                        try:
                            df_mla = pd.read_excel(f_mla)
                            st.session_state.saham_mapping_localites_agences = df_mla
                            save_cache("saham_mapping_localites_agences", df_mla)
                            st.success(f"✅ {len(df_mla):,} agences chargées")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Erreur : {e}")

            st.divider()

            # ═════════════════════════════════════════════════════════════════
            # SECTION B — DONNÉES SAHAM (FICHIER PRINCIPAL)
            # ═════════════════════════════════════════════════════════════════
            st.markdown("""
            <div style="background:linear-gradient(135deg,#1a4d3e,#2d5f4e);border-radius:12px;
                        padding:1.2rem 1.5rem;margin-bottom:1.2rem;">
                <h4 style="color:#fff;margin:0 0 0.3rem 0;">📂 Étape 2 — Données Saham (Fichier Principal TXT)</h4>
                <p style="color:#c8e6c9;margin:0;font-size:0.9em;">
                    Choisissez le mode d'import selon la taille de votre fichier.
                </p>
            </div>
            """, unsafe_allow_html=True)

            mode_saham = st.radio(
                "Mode d'import",
                ["⚡ Fichier unique (tout en une fois)", "📅 Multi-années (par année, recommandé pour gros fichiers)"],
                horizontal=True, key="mode_saham_import"
            )

            st.markdown("<br>", unsafe_allow_html=True)

            # ─────────────────────────────────────────────────────────────────
            # MODE 1 : FICHIER UNIQUE
            # ─────────────────────────────────────────────────────────────────
            if mode_saham == "⚡ Fichier unique (tout en une fois)":
                st.markdown("""
                <div class="info-box">
                    <strong>⚡ Import unique</strong> — Uploadez un ou plusieurs fichiers TXT.
                    Idéal si le fichier est déjà filtré sur une période précise.
                    Pour les très gros fichiers (+4M lignes), préférez le mode Multi-années.
                </div>
                """, unsafe_allow_html=True)

                if st.session_state.saham_credits_depots_raw is not None:
                    df_loaded_s = st.session_state.saham_credits_depots_raw
                    st.success(f"✅ Données Saham en mémoire — **{len(df_loaded_s):,} lignes**")
                    col_si1, col_si2, col_si3 = st.columns(3)
                    cols_s = list(df_loaded_s.columns)
                    annee_col = cols_s[0] if cols_s else None
                    mois_col  = cols_s[1] if len(cols_s) > 1 else None
                    if annee_col:
                        nb_ann = df_loaded_s[annee_col].astype(str).str.strip().str.replace(' ','').str.replace('\u202f','').nunique()
                        col_si1.metric("Années", nb_ann)
                    col_si2.metric("Lignes", f"{len(df_loaded_s):,}")
                    col_si3.metric("Colonnes", len(cols_s))
                    with st.expander("👁️ Aperçu"):
                        st.dataframe(df_loaded_s.head(5), use_container_width=True)
                    c_r1, c_r2, c_r3 = st.columns([1,2,1])
                    with c_r2:
                        if st.button("🗑️ Réinitialiser les données Saham", use_container_width=True, key="reset_saham_unique"):
                            st.session_state.saham_credits_depots_raw = None
                            st.session_state.saham_years_data = {}
                            st.session_state.saham_pdm_localites = None
                            clear_cache("saham_credits_depots_raw")
                            clear_cache("saham_pdm_localites")
                            st.rerun()
                else:
                    f1_tab_list = st.file_uploader(
                        "📂 Sélectionnez un ou plusieurs fichiers TXT",
                        type=["txt", "csv"],
                        key="saham_f1_tab_unique",
                        accept_multiple_files=True,
                        help="Col 3=Code Agence · Col 4=Chapitre · Col 5=Débit · Dernière=Crédit"
                    )
                    if f1_tab_list:
                        try:
                            all_dfs_t = []
                            prog_t = st.progress(0)
                            for idx_t, f1_tab in enumerate(f1_tab_list):
                                msg_t = st.empty()
                                msg_t.info(f"⏳ Lecture de `{f1_tab.name}`...")
                                raw_bytes_t = f1_tab.read()
                                if chardet is not None:
                                    detected_enc_t = chardet.detect(raw_bytes_t).get('encoding','utf-8') or 'utf-8'
                                else:
                                    detected_enc_t = 'utf-8'
                                    for enc in ['utf-8','latin-1','cp1252']:
                                        try:
                                            raw_bytes_t.decode(enc); detected_enc_t = enc; break
                                        except Exception:
                                            pass
                                raw_str_t = raw_bytes_t.decode(detected_enc_t, errors='replace')
                                df_part_t = lire_fichier_principal_saham(raw_str_t)
                                all_dfs_t.append(df_part_t)
                                prog_t.progress((idx_t + 1) / len(f1_tab_list))
                                msg_t.empty()
                            prog_t.empty()
                            df_f1t = pd.concat(all_dfs_t, ignore_index=True) if len(all_dfs_t) > 1 else all_dfs_t[0]
                            st.session_state.saham_credits_depots_raw = df_f1t
                            st.session_state.saham_years_data = {}
                            save_cache("saham_credits_depots_raw", df_f1t)
                            st.success(f"✅ {len(f1_tab_list)} fichier(s) chargé(s) — {len(df_f1t):,} lignes")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Erreur lecture TXT : {e}")

            # ─────────────────────────────────────────────────────────────────
            # MODE 2 : MULTI-ANNÉES
            # ─────────────────────────────────────────────────────────────────
            else:
                st.markdown("""
                <div class="info-box">
                    <strong>📅 Import Multi-Années</strong> — Chargez les données année par année.
                    Chaque année est traitée indépendamment, puis combinée uniquement au moment du calcul.
                    <strong>Recommandé pour les fichiers &gt; 2 millions de lignes.</strong>
                </div>
                """, unsafe_allow_html=True)

                annees_saham_dispo = list(range(2015, 2027))
                annees_saham_sel = st.multiselect(
                    "Sélectionnez les années à importer",
                    options=annees_saham_dispo,
                    default=list(st.session_state.saham_years_data.keys()) or [2024, 2025],
                    key="saham_annees_multi"
                )

                if not annees_saham_sel:
                    st.warning("⚠️ Sélectionnez au moins une année.")
                else:
                    st.info(f"📅 {len(annees_saham_sel)} année(s) sélectionnée(s) : {', '.join(map(str, sorted(annees_saham_sel)))}")
                    st.markdown("<br>", unsafe_allow_html=True)

                    for annee_s in sorted(annees_saham_sel):
                        already_loaded = annee_s in st.session_state.saham_years_data
                        icon = "✅" if already_loaded else "📂"
                        with st.expander(f"{icon} Année **{annee_s}**", expanded=not already_loaded):

                            if already_loaded:
                                df_yr = st.session_state.saham_years_data[annee_s]
                                c_yr1, c_yr2, c_yr3 = st.columns(3)
                                c_yr1.metric("Lignes", f"{len(df_yr):,}")
                                cols_yr = list(df_yr.columns)
                                mois_col_yr = cols_yr[1] if len(cols_yr) > 1 else None
                                if mois_col_yr:
                                    mois_presents = sorted([
                                        int(m) for m in df_yr[mois_col_yr].astype(str)
                                        .str.strip().str.replace('"','').dropna().unique()
                                        if str(m).strip().replace('"','').isdigit() and 1 <= int(str(m).strip().replace('"','')) <= 12
                                    ])
                                    mois_nms = ['Jan','Fév','Mar','Avr','Mai','Jun','Jul','Aoû','Sep','Oct','Nov','Déc']
                                    c_yr2.metric("Mois", len(mois_presents))
                                    c_yr3.write("**Mois :** " + ", ".join([mois_nms[m-1] for m in mois_presents]))
                                col_ry1, col_ry2 = st.columns(2)
                                with col_ry1:
                                    if st.button(f"🔄 Recharger {annee_s}", key=f"reload_saham_{annee_s}", use_container_width=True):
                                        del st.session_state.saham_years_data[annee_s]
                                        st.session_state.saham_credits_depots_raw = None
                                        st.session_state.saham_pdm_localites = None
                                        clear_cache("saham_credits_depots_raw")
                                        st.rerun()
                                with col_ry2:
                                    with st.expander("👁️ Aperçu"):
                                        st.dataframe(df_yr.head(4), use_container_width=True)
                            else:
                                st.write(f"**📂 Fichier(s) TXT pour l'année {annee_s}**")
                                f_yr_list = st.file_uploader(
                                    f"Sélectionnez les fichiers TXT de {annee_s}",
                                    type=["txt","csv"],
                                    accept_multiple_files=True,
                                    key=f"saham_yr_{annee_s}"
                                )
                                if f_yr_list:
                                    if st.button(f"⚙️ Charger l'année {annee_s}", type="primary",
                                                 key=f"btn_load_yr_{annee_s}", use_container_width=True):
                                        with st.spinner(f"⏳ Lecture des fichiers {annee_s}..."):
                                            try:
                                                all_parts = []
                                                prog_yr = st.progress(0)
                                                for i_yr, f_yr in enumerate(f_yr_list):
                                                    raw_b = f_yr.read()
                                                    if chardet is not None:
                                                        enc_yr = chardet.detect(raw_b).get('encoding','utf-8') or 'utf-8'
                                                    else:
                                                        enc_yr = 'utf-8'
                                                        for enc in ['utf-8','latin-1','cp1252']:
                                                            try:
                                                                raw_b.decode(enc); enc_yr = enc; break
                                                            except Exception:
                                                                pass
                                                    raw_s = raw_b.decode(enc_yr, errors='replace')
                                                    all_parts.append(lire_fichier_principal_saham(raw_s))
                                                    prog_yr.progress((i_yr+1)/len(f_yr_list))
                                                prog_yr.empty()
                                                df_yr_combined = pd.concat(all_parts, ignore_index=True) if len(all_parts) > 1 else all_parts[0]
                                                st.session_state.saham_years_data[annee_s] = df_yr_combined
                                                # Reconstruire le df global combiné
                                                st.session_state.saham_credits_depots_raw = pd.concat(
                                                    list(st.session_state.saham_years_data.values()),
                                                    ignore_index=True
                                                )
                                                save_cache("saham_credits_depots_raw", st.session_state.saham_credits_depots_raw)
                                                st.session_state.saham_pdm_localites = None
                                                clear_cache("saham_pdm_localites")
                                                st.success(f"✅ {annee_s} chargé — {len(df_yr_combined):,} lignes")
                                                st.rerun()
                                            except Exception as e:
                                                st.error(f"❌ Erreur : {e}")

                    # ── Récapitulatif multi-années ────────────────────────────
                    if st.session_state.saham_years_data:
                        st.divider()
                        st.markdown("##### 📊 Récapitulatif des données chargées")
                        recap_rows = []
                        for yr, df_yr in sorted(st.session_state.saham_years_data.items()):
                            recap_rows.append({"Année": yr, "Lignes": f"{len(df_yr):,}"})
                        st.dataframe(pd.DataFrame(recap_rows), use_container_width=True, hide_index=True)
                        total_lignes = sum(len(v) for v in st.session_state.saham_years_data.values())
                        st.info(f"📦 Total combiné : **{total_lignes:,} lignes** sur **{len(st.session_state.saham_years_data)} année(s)**")

                        c_rst1, c_rst2, c_rst3 = st.columns([1,2,1])
                        with c_rst2:
                            if st.button("🗑️ Tout réinitialiser", use_container_width=True, key="reset_saham_multi"):
                                st.session_state.saham_years_data = {}
                                st.session_state.saham_credits_depots_raw = None
                                st.session_state.saham_pdm_localites = None
                                clear_cache("saham_credits_depots_raw")
                                clear_cache("saham_pdm_localites")
                                st.rerun()

            st.divider()

            # ═════════════════════════════════════════════════════════════════
            # SECTION C — CALCUL PDM
            # ═════════════════════════════════════════════════════════════════
            mappings_ok  = (st.session_state.saham_mapping_chapitre is not None and
                            st.session_state.saham_mapping_localites_agences is not None)
            donnees_ok   = st.session_state.saham_credits_depots_raw is not None
            all_ok_tab   = mappings_ok and donnees_ok

            if not all_ok_tab:
                missing_tab = []
                if st.session_state.saham_mapping_chapitre is None:
                    missing_tab.append("Mapping Chapitre")
                if st.session_state.saham_mapping_localites_agences is None:
                    missing_tab.append("Mapping Localités")
                if not donnees_ok:
                    missing_tab.append("Données Saham (fichier principal)")
                st.markdown(f"""
                <div class="warning-box">
                    <strong>⏳ En attente de :</strong> {' · '.join(missing_tab)}
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown("""
                <div style="background:linear-gradient(135deg,#1a4d3e,#2d5f4e);border-radius:12px;
                            padding:1.2rem 1.5rem;margin-bottom:1.2rem;">
                    <h4 style="color:#fff;margin:0 0 0.3rem 0;">⚙️ Étape 3 — Calcul des PDM</h4>
                    <p style="color:#c8e6c9;margin:0;font-size:0.9em;">
                        Sélectionnez les périodes Saham et BAM, puis lancez le calcul.
                    </p>
                </div>
                """, unsafe_allow_html=True)

                # ── Filtre période SAHAM ─────────────────────────────────────
                mois_noms_tab = {1:'Janvier',2:'Février',3:'Mars',4:'Avril',5:'Mai',
                                 6:'Juin',7:'Juillet',8:'Août',9:'Septembre',
                                 10:'Octobre',11:'Novembre',12:'Décembre'}

                # ── Préparer données brutes Saham ────────────────────────────
                df_saham_raw_tab = st.session_state.saham_credits_depots_raw.copy()
                cols_saham_tab = list(df_saham_raw_tab.columns)
                col_ann_s  = cols_saham_tab[0] if cols_saham_tab else None
                col_mois_s = cols_saham_tab[1] if len(cols_saham_tab) > 1 else None

                if col_ann_s:
                    df_saham_raw_tab['_ac'] = (
                        df_saham_raw_tab[col_ann_s].astype(str)
                        .str.strip()
                        .str.strip('"')
                        .str.replace(r'[\s\xa0\u202f\u00a0]', '', regex=True)
                        .str.replace(r'\.0$', '', regex=True)   # 2015.0 → 2015
                    )
                    def is_valid_year(v):
                        v2 = str(v).strip()
                        return v2.isdigit() and len(v2) == 4 and 2000 <= int(v2) <= 2035
                    annees_saham_tab = sorted([a for a in df_saham_raw_tab['_ac'].dropna().unique()
                                               if is_valid_year(a)])
                else:
                    annees_saham_tab = []

                if col_mois_s:
                    df_saham_raw_tab['_mc2'] = (
                        df_saham_raw_tab[col_mois_s].astype(str)
                        .str.strip()
                        .str.strip('"')
                        .str.replace(r'[\s\xa0\u202f\u00a0]', '', regex=True)
                        .str.replace(r'\.0$', '', regex=True)   # 1.0 → 1
                    )
                else:
                    pass

                # ── Préparer données BAM ─────────────────────────────────────
                df_bam_tab_global = st.session_state.combined_data_bam
                if df_bam_tab_global is not None:
                    bam_chk = normalize_bam_columns(df_bam_tab_global.copy())
                    annees_bam_tab = sorted(bam_chk['Annee'].dropna().unique().tolist()) if 'Annee' in bam_chk.columns else []
                else:
                    bam_chk = None
                    annees_bam_tab = []

                # ════════════════════════════════════════════════════════════
                # TABLEAU DE SÉLECTION CÔTE À CÔTE : SAHAM  |  BAM
                # ════════════════════════════════════════════════════════════
                st.markdown("""
                <div class="warning-box">
                    <strong>⚠️ Cohérence obligatoire</strong> — Sélectionnez
                    <strong>exactement la même période</strong> pour Saham et BAM.
                    Une incohérence fausse les PDM.
                </div>""", unsafe_allow_html=True)

                col_saham_per, col_sep, col_bam_per = st.columns([5, 1, 5])

                # ── Colonne SAHAM ────────────────────────────────────────────
                with col_saham_per:
                    st.markdown("##### 🏦 Période Saham")
                    mode_s_tab = st.radio("Mode", ["Année + Mois", "Année entière", "Tout"],
                                          key="mode_s_tab_new", horizontal=True)
                    if mode_s_tab != "Tout":
                        if annees_saham_tab:
                            ann_s_tab = st.selectbox("Année Saham", annees_saham_tab, key="ann_s_tab_new")
                        else:
                            ann_s_tab = None
                            raw_vals  = list(df_saham_raw_tab[col_ann_s].astype(str).unique()[:5]) if col_ann_s else []
                            clean_vals = list(df_saham_raw_tab['_ac'].unique()[:5]) if '_ac' in df_saham_raw_tab.columns else []
                            st.error(f"❌ Années non détectées. Valeurs brutes : {raw_vals} → nettoyées : {clean_vals}")
                    else:
                        ann_s_tab = None
                        st.info("Toutes les années Saham")

                    if mode_s_tab == "Année + Mois":
                        # Mois disponibles pour l'année sélectionnée uniquement
                        if ann_s_tab and '_mc2' in df_saham_raw_tab.columns:
                            mois_dispo_s = sorted([
                                int(m) for m in df_saham_raw_tab[df_saham_raw_tab['_ac'] == str(ann_s_tab)]['_mc2'].dropna().unique()
                                if str(m).strip().isdigit() and 1 <= int(str(m).strip()) <= 12
                            ])
                        else:
                            mois_dispo_s = []
                        if mois_dispo_s:
                            mois_s_tab = st.selectbox("Mois Saham", mois_dispo_s,
                                                      format_func=lambda x: mois_noms_tab.get(int(x), str(x)),
                                                      key="mois_s_tab_new")
                        else:
                            mois_s_tab = None
                            st.warning("Aucun mois disponible pour cette année")
                    else:
                        mois_s_tab = None

                # ── Séparateur visuel ────────────────────────────────────────
                with col_sep:
                    st.markdown("<div style='border-left:2px solid #ccc;height:180px;margin:2rem auto;'></div>",
                                unsafe_allow_html=True)

                # ── Colonne BAM ──────────────────────────────────────────────
                with col_bam_per:
                    st.markdown("##### 📊 Période BAM")
                    if df_bam_tab_global is not None:
                        mode_t = st.radio("Mode", ["Année + Mois", "Année entière", "Tout"],
                                          key="mode_t_tab_new", horizontal=True)
                        if mode_t != "Tout":
                            annee_t = st.selectbox("Année BAM", annees_bam_tab, key="annee_t_tab_new")
                        else:
                            annee_t = None
                            st.info("Toutes les années BAM")

                        if mode_t == "Année + Mois":
                            # Mois disponibles pour l'année BAM sélectionnée uniquement
                            if annee_t and 'Annee' in bam_chk.columns and 'mois' in bam_chk.columns:
                                mois_dispo_b = sorted([
                                    int(m) for m in bam_chk[bam_chk['Annee'] == annee_t]['mois'].dropna().unique()
                                    if str(m).strip().replace('.0','').isdigit() and 1 <= int(float(m)) <= 12
                                ])
                            else:
                                mois_dispo_b = []
                            if mois_dispo_b:
                                mois_t_sel = st.selectbox("Mois BAM", mois_dispo_b,
                                                          format_func=lambda x: mois_noms_tab.get(int(x), str(x)),
                                                          key="mois_t_tab_new")
                            else:
                                mois_t_sel = None
                                st.warning("Aucun mois disponible pour cette année")
                        else:
                            mois_t_sel = None
                    else:
                        st.warning("⚠️ Aucune donnée BAM chargée.")
                        mode_t = "Tout"; annee_t = None; mois_t_sel = None

                # ── Résumé période sélectionnée ──────────────────────────────
                def periode_label(mode, annee, mois, noms):
                    if mode == "Tout":
                        return "Toutes les données"
                    elif mode == "Année entière":
                        return f"Année {annee} (dernier mois)"
                    else:
                        m_nom = noms.get(int(mois), str(mois)) if mois else "—"
                        return f"{annee} — {m_nom}"

                label_s = periode_label(mode_s_tab, ann_s_tab, mois_s_tab, mois_noms_tab)
                label_b = periode_label(mode_t, annee_t, mois_t_sel, mois_noms_tab) if df_bam_tab_global is not None else "BAM non chargé"
                match_color = "#28a745" if label_s == label_b else "#dc3545"
                match_icon  = "✅" if label_s == label_b else "⚠️"
                st.markdown(f"""
                <div style="background:#f8f9fa;border:2px solid {match_color};border-radius:8px;
                            padding:0.8rem 1.2rem;margin-top:0.8rem;display:flex;gap:2rem;align-items:center;">
                    <span>🏦 <strong>Saham :</strong> {label_s}</span>
                    <span>📊 <strong>BAM :</strong> {label_b}</span>
                    <span style="color:{match_color};font-weight:bold;">{match_icon} {'Périodes identiques' if label_s == label_b else 'Périodes différentes !'}</span>
                </div>
                """, unsafe_allow_html=True)

                # ── Filtrer Saham ────────────────────────────────────────────
                df_saham_filtered_tab = df_saham_raw_tab.copy()
                if mode_s_tab != "Tout" and ann_s_tab and '_ac' in df_saham_filtered_tab.columns:
                    df_saham_filtered_tab = df_saham_filtered_tab[df_saham_filtered_tab['_ac'] == str(ann_s_tab)]
                if mode_s_tab == "Année + Mois" and mois_s_tab and '_mc2' in df_saham_filtered_tab.columns:
                    df_saham_filtered_tab = df_saham_filtered_tab[df_saham_filtered_tab['_mc2'] == str(mois_s_tab)]
                elif mode_s_tab == "Année entière" and '_mc2' in df_saham_filtered_tab.columns:
                    mois_dispo_tab = [int(m) for m in df_saham_filtered_tab['_mc2'].dropna().unique()
                                      if str(m).strip().isdigit() and 1 <= int(str(m).strip()) <= 12]
                    if mois_dispo_tab:
                        df_saham_filtered_tab = df_saham_filtered_tab[
                            df_saham_filtered_tab['_mc2'] == str(max(mois_dispo_tab))]
                df_saham_filtered_tab = df_saham_filtered_tab.drop(
                    columns=[c for c in ['_ac','_mc2'] if c in df_saham_filtered_tab.columns])

                # ── Aperçu lignes ────────────────────────────────────────────
                cs_inf1, cs_inf2 = st.columns(2)
                cs_inf1.info(f"📊 Saham filtré : **{len(df_saham_filtered_tab):,} lignes**")
                if df_bam_tab_global is not None:
                    bam_prev_t = bam_chk.copy()
                    if mode_t != "Tout" and annee_t:
                        bam_prev_t = bam_prev_t[bam_prev_t['Annee'] == annee_t]
                    if mode_t == "Année + Mois" and mois_t_sel:
                        bam_prev_t = bam_prev_t[bam_prev_t['mois'] == mois_t_sel]
                    elif mode_t == "Année entière" and annee_t:
                        mois_bam_dispo = sorted([int(float(m)) for m in bam_prev_t['mois'].dropna().unique()
                                                  if str(m).replace('.0','').isdigit() and 1 <= int(float(m)) <= 12])
                        if mois_bam_dispo:
                            bam_prev_t = bam_prev_t[bam_prev_t['mois'] == max(mois_bam_dispo)]
                    nb_bam_t = bam_prev_t['Localite'].nunique() if 'Localite' in bam_prev_t.columns else 0
                    cs_inf2.info(f"📊 BAM filtré : **{len(bam_prev_t):,} lignes** | **{nb_bam_t} localités**")

                st.markdown("<br>", unsafe_allow_html=True)

                # ── Bouton Calculer ──────────────────────────────────────────
                col_b1, col_b2, col_b3 = st.columns([1, 2, 1])
                with col_b2:
                    if st.button("🚀 Calculer les PDM par Localité", type="primary",
                                 use_container_width=True, key="traiter_saham_tab_new"):
                        with st.spinner("⏳ Traitement en cours..."):
                            try:
                                if df_bam_tab_global is not None:
                                    bam_pdm_t = normalize_bam_columns(df_bam_tab_global.copy())
                                    bam_pdm_t = clean_numeric_columns(bam_pdm_t)
                                    if mode_t != "Tout" and annee_t:
                                        bam_pdm_t = bam_pdm_t[bam_pdm_t['Annee'] == annee_t]
                                    if mode_t == "Année + Mois" and mois_t_sel:
                                        bam_pdm_t = bam_pdm_t[bam_pdm_t['mois'] == mois_t_sel]
                                    elif mode_t == "Année entière" and annee_t:
                                        mois_bam_calc = sorted([int(float(m)) for m in bam_pdm_t['mois'].dropna().unique()
                                                                  if str(m).replace('.0','').isdigit() and 1 <= int(float(m)) <= 12])
                                        if mois_bam_calc:
                                            bam_pdm_t = bam_pdm_t[bam_pdm_t['mois'] == max(mois_bam_calc)]
                                    if len(bam_pdm_t) == 0:
                                        st.error("❌ Aucune donnée BAM après filtrage.")
                                        st.stop()
                                else:
                                    bam_pdm_t = None

                                df_pdm_tab = process_saham_pdm(
                                    df_saham_filtered_tab,
                                    st.session_state.saham_mapping_chapitre,
                                    st.session_state.saham_mapping_localites_agences,
                                    bam_pdm_t
                                )
                                st.session_state.saham_pdm_localites = df_pdm_tab
                                save_cache("saham_pdm_localites", df_pdm_tab)
                                st.success(f"✅ PDM calculées pour **{len(df_pdm_tab)} localités**")
                                st.rerun()
                            except Exception as e:
                                st.error(f"❌ Erreur : {e}")
                                st.exception(e)

            # ── Redirection après calcul ──────────────────────────────────────
            if st.session_state.saham_pdm_localites is not None:
                st.divider()
                st.markdown("""
                <div class="success-box">
                    ✅ <strong>PDM calculées avec succès !</strong><br>
                    Rendez-vous sur <strong>Visualisations Saham Bank</strong> dans le menu latéral pour consulter tous les graphiques et tableaux.
                </div>
                """, unsafe_allow_html=True)
    
    # PAGE VISUALISATIONS BAM
    elif page == "Visualisations BAM":
        if st.session_state.combined_data_bam is not None:
            visualisations_bam_avancees()
        else:
            st.warning("Veuillez d'abord importer et traiter des données BAM")
    
    # PAGE VISUALISATIONS SAHAM BANK
    elif page == "Visualisations Saham Bank":
        create_saham_visualizations()
    
    # PAGE À PROPOS
    elif page == "À propos":
        st.header("À propos")
        st.markdown("""
        ### SAHAM BANK - Module BAM
        
        **Projet de Fin d'Études - 2026**
        
        #### Fonctionnalités :
        - Import multiple de fichiers Excel
        - Combinaison automatique avec mois
        - Détection des valeurs manquantes
        - Visualisations interactives
        - Analyses par Direction Régionale et Localité
        
        **Direction Financière de Saham Bank**  
        *Version 3.0 - Février 2026*
        """)

def main():
    # État 1 : Page d'accueil
    if st.session_state.show_welcome:
        welcome_page()
    # État 2 : Login en attente pour un module
    elif st.session_state.pending_module is not None:
        module_login_page(st.session_state.pending_module)
    # État 3 : Sélection du module (aucun module actif)
    elif st.session_state.selected_module is None:
        module_selection_page()
    # État 4 : Module BAM
    elif st.session_state.selected_module == "BAM":
        bam_module()
    # État 5 : Module Balance
    elif st.session_state.selected_module == "Balance":
        balance_module()

if __name__ == "__main__":
    main()
